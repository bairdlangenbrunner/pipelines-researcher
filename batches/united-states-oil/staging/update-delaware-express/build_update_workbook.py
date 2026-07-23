#!/usr/bin/env python3
"""Build the Delaware Express update workbook from the 2026-06-12 research.

Ad-hoc update-mode builder (2 rows, P7995 Targa NGL + P0354 Medallion->Plains crude)
per docs/reference/workbook_conventions.md: changed/filled cells red (FFCCCC fill /
CC0000 font), re-verified-unchanged cells blue (4472C4 fill, white font), README
first, no wrap anywhere (clip).

Usage: python batches/staging/delaware-express/build_update_workbook.py \
           --output batches/pipelines_batch_<stamp>_delaware-express_update.xlsx
"""
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
SNAP_OIL = ROOT / "data/GOIT_oil_ngl_snapshot_20260612.csv"
SNAP_OO = ROOT / "data/GEM_operators_owners_snapshot_20260612.csv"

PIDS = ["P7995", "P0354"]

RED_FILL = PatternFill("solid", start_color="FFCCCC")
RED_FONT = Font(color="CC0000", bold=True)
BLUE_FILL = PatternFill("solid", start_color="4472C4")
WHITE_FONT = Font(color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color="4472C4")
HDR_FONT = Font(color="FFFFFF", bold=True)
CLIP = Alignment(wrap_text=False, vertical="top")
HDR_ALIGN = Alignment(wrap_text=False, vertical="center", horizontal="center")

# refs (all passed scripts/url_verifier.py on 2026-06-12; PDF/SEC caveats in README)
TRGP_Q424 = "https://www.globenewswire.com/news-release/2025/02/20/3029407/14074/en/targa-resources-corp-reports-record-fourth-quarter-and-full-year-2024-financial-results-provides-growth-outlook-for-2025-and-announces-refinancing-of-badlands-preferred-equity.html"
TRGP_Q225 = "https://www.globenewswire.com/news-release/2025/08/07/3129009/14074/en/targa-resources-corp-reports-second-quarter-2025-financial-results.html"
TRGP_Q325 = "https://www.globenewswire.com/news-release/2025/11/05/3181203/0/en/targa-resources-corp-reports-record-third-quarter-2025-results-and-announces-expectation-for-a-25-increase-to-its-2026-common-dividend.html"
TRGP_Q425 = "https://www.globenewswire.com/news-release/2026/02/19/3240865/14074/en/Targa-Resources-Corp-Reports-Record-Fourth-Quarter-and-Full-Year-2025-Financial-Results-and-Provides-Outlook-for-Record-2026.html"
TRGP_Q126 = "https://www.globenewswire.com/news-release/2026/05/07/3289714/14074/en/Targa-Resources-Corp-Reports-Record-First-Quarter-2026-Financial-Results-and-Increases-Financial-Outlook-for-2026.html"
EASTDALEY = "https://eastdaley.com/ngl-insider/targa-adds-to-infrastructure-targeting-permian-ngls-3"
MED_2017 = "https://web.archive.org/web/20191114195243/https://www.medallionmidstream.com/news/medallion-announces-binding-open-season-new-delaware-basin-crude-oil-pipeline-joint-tariff"
MED_SITE_2023 = "https://web.archive.org/web/20230117104813/https://www.medallionmidstream.com/operations/delaware-basin"
RRC_26 = "https://www.rrc.texas.gov/media/jmxfrivr/2-6.pdf"
BW_2019 = "https://web.archive.org/web/20200302222702/https://www.businesswire.com/news/home/20190412005086/en/Medallion-Delaware-Express-Medallion-Pipeline-Announce-Binding"
OGJ_2019 = "https://www.ogj.com/general-interest/article/14036198/permian-crude-producers-push-gathering-capacity-optionality"
NGI_2019 = "https://naturalgasintel.com/news/medallion-delaware-express-launch-open-season-to-expand-permian-crude-capacity/"
FERC_OR1934 = "https://www.ferc.gov/sites/default/files/2020-06/G-2-OR19-34-000.pdf"
EMG_2025 = "https://emgtx.com/emg-closes-on-the-sale-of-its-crude-oil-business-in-the-delaware-basin/"
PAA_2025 = "https://www.globenewswire.com/news-release/2025/01/08/3005910/0/en/Plains-All-American-Announces-Bolt-on-Acquisitions-Capital-Structure-Optimization-Distribution-Increase.html"
TROUT_2025 = "https://www.troutman.com/insights/troutman-pepper-locke-advises-emg-and-medallion-in-sale-of-crude-oil-business-in-delaware-basin-to-plains-oryx-permian-basin.html"
PAA_8K_2021 = "https://www.sec.gov/Archives/edgar/data/0001070423/000110465921123178/tm2129264d1_ex99-1.htm"
STONEPEAK_2021 = "https://stonepeak.com/news/plains-all-american-and-oryx-midstream-complete-formation-of-permian-basin-joint-venture"

TODAY = "2026-06-12"


def refs(*urls):
    return ", ".join(urls)


NOTES = {
    "P7995": (
        "[2026-06-12 CB] Full re-research. NOT redundant with Delaware Express Pipeline (P0354) - unrelated "
        "same-named assets (P0354 is the Medallion/now-Plains Oryx 61-mi/16-in crude shuttle). Announced "
        "2025-02-20 with Targa Q4-2024 results (not Nov 2024 as previously noted): 'Delaware Express, a "
        "100-mile, 30-inch diameter pipeline expansion of its Grand Prix NGL Pipeline in the Permian Delaware' "
        "- new physical pipe (loop), so length/diameter filled rather than expansion-zeroed. Timeline: guided "
        "Q3 2026 at announcement; 'early completion' expected (Q2-2025 PR, 2025-08-07; 2025 capex raised to "
        "~$3.0B partly on acceleration); construction through 2025 (Q3-2025 + Q4-2025 PRs); Q2 2026 guidance "
        "(FY2025 results, 2026-02-19); Q1-2026 PR (2026-05-07): 'In May 2026, starting up operation of our "
        "Delaware Express NGL Pipeline expansion' - status proposed->operating (startup commenced May 2026; "
        "Reuters carried it same day; re-check for full-service confirmation next sweep). Capacity not "
        "disclosed for the line (Targa NGL system >1,000 MBbl/d into Mont Belvieu per FY2025 10-K); cost not "
        "separately disclosed (folded into 2025 growth capex). Tiers: length medium (East Daley repeats the "
        "company figure - one originating source); diameter/status company-stated; owner high. Parent filled "
        "Targa Resources Corp 100% (NYSE: TRGP; owns 100% of Grand Prix since Jan 2023 Blackstone buyout). Old "
        "targaresources.com ref link-rotted (timeouts 2026-06-12) - replaced on every ref cell. Distinct from "
        "Targa's Speedway NGL Pipeline (P7991, 500 mi/30 in, ISD 2027) - do not conflate."
    ),
    "P0354": (
        "[2026-06-12 CB] Ref fill + corrections (row had values but zero refs). NOT redundant with Delaware "
        "Express NGL Pipeline (P7995, Targa's 100-mi/30-in NGL loop of Grand Prix). Capacity 250,000->200,000 "
        "bpd: no source states 250k; Medallion site listed '200,000 bpd Crude Oil Pipeline Hydraulic Capacity' "
        "(archived 2023-01-17) and the Oct 2017 open-season release capped joint-tariff service at 'up to "
        "200,000 bpd'; initial mainline capacity 90,000 bpd - medium-high (operator-stated, two documents). "
        "Length 60->61 mi: 61-mile 16-inch mainline (Business Wire 2019-04-12, archived; OGJ 2019-07-01, "
        "independent) - high; planned as 63 mi in 2017; whole Delaware system incl. gathering ~130-140 mi. "
        "Flow direction corrected (Start/End swapped): origins are field points in Reeves/Pecos/Ward Cos incl. "
        "Eagle Eye Station, Ward Co. (RRC Texas Tariff No. 2.6, T-4 09805) with sole destination Crane Hub, "
        "Crane Co. (interconnects: EPIC Crude, Gray Oak, Medallion Midland, Plains/Cactus); Medallion 2017: "
        "'originating in Ward County and extending east to... Crane County'. 'Barstow' kept as the Ward Co. "
        "origin label but is itself map-derived/uncorroborated. StartYear 2019->2018 (inferred-medium): FERC "
        "order 170 FERC P61,048 fn.57 - implementing tariff filed 2018-07-01 (Docket IS18-661), effective by "
        "operation of law; operating by 2019-04-12 (BW). 2019 expansion: Reeves Co. gathering + mainline "
        "expansion (FERC OR19-28; Federal Register 2019-06-27). Ownership: EMG-backed Medallion Delaware "
        "Express LLC sold via EMG Medallion 2 Holdings LLC to POPB DE Crude Holdings LLC (Plains Oryx Permian "
        "Basin subsidiary), closed 2025-01-07, ~$160M (EMG + Plains + Troutman + OGJ - high). Medallion "
        "Pipeline Co LLC dropped from Owner: it owns the separate Midland-Basin system (acquired by ONEOK Oct "
        "2024) and was only the joint-tariff partner, never an owner of this line. Parent = Plains All "
        "American Pipeline LP 65% / Oryx Midstream Holdings LLC 35% (Plains Oryx JV split per PAA 8-K Ex-99.1 "
        "+ Stonepeak release, 2021-10-05; Plains operates the JV; Oryx is Stonepeak-backed)."
    ),
}

CHANGES = {
    "P7995": {
        "Status": "operating",
        "Status [ref]": refs(TRGP_Q126),
        "Fuel [ref]": refs(TRGP_Q424, EASTDALEY),
        "PipelineType [ref]": refs(TRGP_Q424),
        "Start [ref]": refs(TRGP_Q126, TRGP_Q425),
        "LengthKnown": "100",
        "LengthKnownUnits": "mi",
        "LengthKnownKm": "160.93",
        "LengthMergedKm": "160.93",
        "Length [ref]": refs(TRGP_Q424, EASTDALEY),
        "Diameter": "30.00",
        "DiameterUnits": "in",
        "DiameterInMm": "762.00",
        "Diameter [ref]": refs(TRGP_Q424),
        "StartLocation [ref]": refs(TRGP_Q424),
        "EndLocation": "Grand Prix NGL Pipeline interconnect",
        "EndLocation [ref]": refs(TRGP_Q424),
        "Parent": "Targa Resources Corp [100.00%]",
        "ResearcherNotes": NOTES["P7995"],
        "LastUpdated": TODAY,
    },
    "P0354": {
        "Capacity": "200,000.00",
        "CapacityBOEd": "200000.00",
        "Capacity [ref]": refs(MED_SITE_2023, MED_2017),
        "LengthKnown": "61",
        "LengthKnownKm": "98.17",
        "LengthMergedKm": "98.17",
        "Length [ref]": refs(BW_2019, OGJ_2019),
        "Diameter [ref]": refs(BW_2019, OGJ_2019, NGI_2019),
        "StartYear1": "2018",
        "StartYearEarliest": "2018",
        "Start [ref]": refs(FERC_OR1934, BW_2019),
        "Status [ref]": refs(EMG_2025, PAA_2025),
        "Fuel [ref]": refs(MED_2017, RRC_26),
        "FuelSource [ref]": refs(MED_2017, OGJ_2019),
        "StartLocation": "Barstow",
        "StartLocation [ref]": refs(RRC_26, MED_2017),
        "EndLocation": "Crane",
        "EndLocation [ref]": refs(RRC_26, MED_2017),
        "Owner": "Medallion Delaware Express LLC [100.%]",
        "OwnerEntityIDs": "E100001015292",
        "Parent": "Plains All American Pipeline LP [65.00%]; Oryx Midstream Holdings LLC [35.00%]",
        "OtherEnglishNames": "Medallion Delaware Express",
        "ResearcherNotes": NOTES["P0354"],
        "LastUpdated": TODAY,
    },
}

# value cells confirmed unchanged this batch (blue)
REVERIFIED = {
    "P7995": ["Fuel", "PipelineType", "StartYear1", "StartLocation", "Owner"],
    "P0354": ["Status", "Fuel", "Diameter", "FuelSource"],
}

OO_CHANGES = {
    "P0354": {
        "AggregateOwners": "Medallion Delaware Express LLC [100.%]",
        "Owner1": "Medallion Delaware Express LLC",
        "Owner1%": "100.00%",
        "Owner2": "",
        "Owner [ref]": refs(EMG_2025, PAA_2025, TROUT_2025),
        "Operator": "Medallion Delaware Express LLC",
        "Operator [ref]": refs(RRC_26, STONEPEAK_2021),
        "LastUpdated": "2026/06/12",
    },
    "P7995": {
        "Owner [ref]": refs(TRGP_Q424),
        "Operator": "Targa Resources Corp",
        "Operator [ref]": refs(TRGP_Q126),
        "LastUpdated": "2026/06/12",
    },
}

SUMMARY_ROWS = [
    # pid, asset, field, current, proposed, action, tier, evidence, urls
    ("-", "both", "REDUNDANCY VERDICT", "two same-named rows", "keep both - unrelated assets", "confirm", "high",
     "P0354 = Medallion (now Plains Oryx) crude shuttle, 61 mi/16 in, Ward Co->Crane Hub, in service ~2018. P7995 = Targa NGL loop of Grand Prix, 100 mi/30 in, announced 2025-02-20, started up May 2026. Different commodity, owner, size, vintage, route; no source connects them.",
     refs(TRGP_Q424, RRC_26)),
    ("P7995", "Targa NGL", "Status", "proposed", "operating", "change", "medium-high",
     "Targa Q1-2026 PR (2026-05-07): 'In May 2026, starting up operation of our Delaware Express NGL Pipeline expansion'; Reuters same day. Startup commenced - re-check for full-service confirmation next sweep.",
     TRGP_Q126),
    ("P7995", "Targa NGL", "LengthKnown", "(blank)", "100 mi (160.93 km)", "fill", "medium",
     "'100-mile, 30-inch diameter pipeline expansion of its Grand Prix NGL Pipeline' (2025-02-20 PR); East Daley '100-mile NGL pipe' (derives from same announcement -> one originating source). New physical pipe - expansion-zero convention does NOT apply.",
     refs(TRGP_Q424, EASTDALEY)),
    ("P7995", "Targa NGL", "Diameter", "(blank)", "30 in (762 mm)", "fill", "medium",
     "Company-stated (PR + same-day 8-K Ex-99.1); no independent spec found.", TRGP_Q424),
    ("P7995", "Targa NGL", "EndLocation", "Targa NGL pipeline system interconnect", "Grand Prix NGL Pipeline interconnect", "change", "medium",
     "Only disclosed as an intra-Delaware expansion of Grand Prix; no counties/plants/tie-ins published. NOT a Mont Belvieu long-haul (that is Speedway P7991).", TRGP_Q424),
    ("P7995", "Targa NGL", "Parent", "unknown", "Targa Resources Corp [100.00%]", "fill", "high",
     "TRGP describes the line as 'its/our' pipeline in every filing; owns 100% of Grand Prix since Jan 2023. OwnerEntityIDs still needs a GEM-assigned E-ID (none exists for Targa in GOIT).", refs(TRGP_Q424, TRGP_Q126)),
    ("P7995", "Targa NGL", "all [ref] cells", "targaresources.com segment page", "globenewswire PRs + East Daley", "change", "-",
     "Old ref link-rotted (ReadTimeout from two networks 2026-06-12).", ""),
    ("P7995", "Targa NGL", "Capacity / Cost", "(blank)", "(no change - FLAG)", "review", "-",
     "Never disclosed for the line. Context: NGL system >1,000 MBbl/d to Mont Belvieu (FY2025 10-K); cost folded into 2025 capex $2.6-2.8B -> ~$3.0B.", refs(TRGP_Q424, TRGP_Q225)),
    ("P0354", "Medallion crude", "Capacity", "250,000 bpd", "200,000 bpd", "change", "medium-high",
     "No source states 250k. Medallion site (arch. 2023): '200,000 bpd... Hydraulic Capacity'; 2017 release: joint tariff 'up to 200,000 bpd'; initial mainline 90,000 bpd.", refs(MED_SITE_2023, MED_2017)),
    ("P0354", "Medallion crude", "LengthKnown", "60 mi", "61 mi (98.17 km)", "change", "high",
     "61-mile, 16-inch mainline: BW 2019-04-12 (archived) + OGJ 2019-07-01 (independent). Planned 63 mi in 2017.", refs(BW_2019, OGJ_2019)),
    ("P0354", "Medallion crude", "StartLocation / EndLocation", "Crane -> Barstow", "Barstow -> Crane (SWAPPED)", "change", "high",
     "GEM had the flow reversed. RRC Tariff 2.6: origins = Reeves/Pecos/Ward Co field points (Eagle Eye Station, Ward Co); sole destination Crane Hub, Crane Co. 2017 release: 'originating in Ward County and extending east'. 'Barstow' label itself uncorroborated (Ward Co town) - alt: 'Eagle Eye Station (Ward County)'.", refs(RRC_26, MED_2017)),
    ("P0354", "Medallion crude", "StartYear1", "2019", "2018", "change", "medium",
     "FERC 170 FERC P61,048 fn.57: implementing tariff filed 2018-07-01 (IS18-661), effective by operation of law -> service ~Aug 2018; operating by 2019-04-12 (BW). Inferred - revert to 2019 if judged too thin.", refs(FERC_OR1934, BW_2019)),
    ("P0354", "Medallion crude", "Owner", "Medallion Delaware Express LLC; Medallion Pipeline Co LLC", "Medallion Delaware Express LLC [100.%]", "change", "high",
     "Medallion Pipeline Co LLC = the separate Midland system (ONEOK since Oct 2024), only the joint-tariff partner here, never an owner of this line.", refs(EMG_2025, TROUT_2025)),
    ("P0354", "Medallion crude", "Parent", "Medallion Midstream LLC; unknown", "Plains All American Pipeline LP [65.00%]; Oryx Midstream Holdings LLC [35.00%]", "change", "high",
     "Sold to Plains Oryx Permian Basin subsidiary (POPB DE Crude Holdings LLC), closed 2025-01-07, ~$160M - buyer+seller+counsel+OGJ. JV split 65/35 per PAA 8-K Ex-99.1 + Stonepeak release (2021-10-05). NOT ONEOK (that deal = Midland Medallion).", refs(EMG_2025, PAA_2025, PAA_8K_2021, STONEPEAK_2021)),
    ("P0354", "Medallion crude", "all blank [ref] cells", "(blank)", "filled (Fuel/Status/Capacity/Length/Diameter/Start/FuelSource/StartLoc/EndLoc)", "fill", "high",
     "Row had values but zero refs; every filled ref passed url_verifier 2026-06-12.", ""),
    ("P0354", "Medallion crude", "Operator (operators/owners tab)", "(blank)", "Medallion Delaware Express LLC", "fill", "medium",
     "Carrier of record per RRC Tariff 2.6 (T-4 09805) - tariff predates the Jan 2025 sale, hence medium. Plains operates the Plains Oryx JV day-to-day.", refs(RRC_26, STONEPEAK_2021)),
]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
    ws.freeze_panes = "A2"


def autowidth(ws, widths=None, default=14, cap=55):
    widths = widths or {}
    for c in range(1, ws.max_column + 1):
        hdr = ws.cell(row=1, column=c).value or ""
        ws.column_dimensions[get_column_letter(c)].width = widths.get(hdr, min(max(default, len(str(hdr)) + 2), cap))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    oil = pd.read_csv(SNAP_OIL, header=2, low_memory=False, dtype=str)
    oo = pd.read_csv(SNAP_OO, header=1, low_memory=False, dtype=str)
    oil_rows = {r["ProjectID"]: r for _, r in oil[oil["ProjectID"].isin(PIDS)].iterrows()}
    oo_rows = {r["ProjectID"]: r for _, r in oo[oo["ProjectID"].isin(PIDS)].iterrows()}

    wb = Workbook()

    # README
    ws = wb.active
    ws.title = "README"
    readme = [
        ["Delaware Express - update batch (both same-named assets)"],
        [""],
        ["Mode", "update"],
        ["Scope", "Delaware Express NGL Pipeline (P7995, Targa) + Delaware Express Pipeline (P0354, Medallion->Plains Oryx), GOIT oil/NGL tracker, United States"],
        ["Researched", TODAY],
        ["GEM snapshot", SNAP_OIL.name],
        ["Staging", "batches/staging/delaware-express/staged_updates.json"],
        ["URL verification", "Every URL passed scripts/url_verifier.py on 2026-06-12. Caveats: sec.gov 403s the verifier UA (PAA 8-K confirmed with declared-contact UA); rrc.texas.gov + ferc.gov PDFs pass HTTP but bodies confirmed via pypdf; businesswire bot-walls -> archive.org snapshot used; medallionmidstream.com link-rotted -> archive.org."],
        [""],
        ["REDUNDANCY VERDICT", "NOT redundant - keep both rows. Same name, unrelated assets: P0354 is the Medallion (now Plains Oryx) 61-mi/16-in CRUDE shuttle (Ward Co -> Crane Hub, ~2018); P7995 is Targa's 100-mi/30-in NGL loop of Grand Prix (announced 2025-02-20, started up May 2026). Different commodity/owner/size/vintage/route; no source connects them. Also do not conflate P7995 with Targa's Speedway NGL Pipeline (P7991)."],
        [""],
        ["Color key"],
        ["red fill / red font", "changed or newly filled cell - the thing to review and paste"],
        ["blue fill / white font", "value unchanged but re-verified this batch (>=2 sources unless noted)"],
        [""],
        ["Sheets"],
        ["Oil_Updated", "Both rows, full GOIT column layout (paste-ready). Changed/filled cells red; re-verified values blue. ResearcherNotes carries tier + evidence per row."],
        ["Oil_OperatorsOwners", "Both rows as on the 'Pipeline operators/owners' tab (GID 1489950650), ProjectID-keyed. Operator/Owner [ref] fills + the P0354 owner cleanup in red."],
        ["Changes_Summary", "One line per finding: field, current vs proposed, action, tier, key evidence, sources."],
        [""],
        ["Headline findings"],
        ["1", "P7995 status proposed -> operating: Targa Q1-2026 results (2026-05-07) - 'In May 2026, starting up operation of our Delaware Express NGL Pipeline expansion'."],
        ["2", "P7995 specs filled: 100 mi / 30 in (announcement PR 2025-02-20 + East Daley). New physical pipe (loop of Grand Prix) - expansion-zero convention does not apply. Capacity/cost never disclosed (flagged, left blank)."],
        ["3", "P0354 capacity corrected 250,000 -> 200,000 bpd (no source supports 250k; operator says 200k hydraulic; initial 90k)."],
        ["4", "P0354 flow direction corrected: Start/End swapped to Barstow (Ward Co) -> Crane (Crane Hub = sole tariff destination). Length 60 -> 61 mi."],
        ["5", "P0354 ownership: sold to Plains Oryx Permian Basin subsidiary, closed 2025-01-07. Parent = PAA 65% / Oryx Midstream Holdings 35%. NOT ONEOK - ONEOK bought the separate Midland-Basin Medallion (Oct 2024). Medallion Pipeline Co LLC dropped from Owner."],
        ["6", "P0354 StartYear 2019 -> 2018 (inferred-medium: FERC tariff filed 2018-07-01, effective by operation of law; operating by Apr 2019). Revert if judged too thin."],
        ["7", "P0354 had values but ZERO refs - all ref cells now filled and verified."],
        ["8", "P7995 old targaresources.com ref is link-rotted (timeouts) - replaced on every ref cell."],
        [""],
        ["Follow-ups", "(a) P7995 OwnerEntityIDs needs a GEM-assigned E-ID for Targa Resources Corp (none in GOIT; P7991 same). (b) Divestiture sweep: any other GOIT row owned by Medallion Pipeline Co LLC / Medallion Midstream LLC now parents to ONEOK Inc. (c) P7995 route still unmapped; no route published. (d) Re-check P7995 for full-service confirmation after Targa's Q2-2026 results (~Aug 2026)."],
        ["Note", "Researcher initials left as-is on the rows; flip to CB on apply if preferred."],
        ["Escalation gates", "None tripped (2-row targeted batch)."],
    ]
    for r in readme:
        ws.append(r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 160
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = CLIP
    ws["A1"].font = Font(bold=True, size=14)

    # Oil_Updated
    ws = wb.create_sheet("Oil_Updated")
    cols = list(oil.columns)
    ws.append(cols)
    style_header(ws, len(cols))
    for i, pid in enumerate(PIDS, start=2):
        base = oil_rows[pid]
        changed = CHANGES[pid]
        for j, col in enumerate(cols, start=1):
            val = changed.get(col, base.get(col))
            val = "" if (val is None or (isinstance(val, float) and pd.isna(val)) or str(val) == "nan") else str(val)
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CLIP
            if col in changed:
                cell.fill = RED_FILL
                cell.font = RED_FONT
            elif col in REVERIFIED[pid]:
                cell.fill = BLUE_FILL
                cell.font = WHITE_FONT
    autowidth(ws, {"PipelineName": 38, "Status [ref]": 55, "Owner": 45, "Parent": 55,
                   "ResearcherNotes": 55, "OtherEnglishNames": 30, "Capacity [ref]": 55,
                   "Length [ref]": 55, "Start [ref]": 55})

    # Oil_OperatorsOwners
    ws = wb.create_sheet("Oil_OperatorsOwners")
    oo_cols = list(oo.columns)
    ws.append(oo_cols)
    style_header(ws, len(oo_cols))
    for i, pid in enumerate(PIDS, start=2):
        base = oo_rows[pid]
        changed = OO_CHANGES[pid]
        for j, col in enumerate(oo_cols, start=1):
            val = changed.get(col, base.get(col))
            val = "" if (val is None or (isinstance(val, float) and pd.isna(val)) or str(val) == "nan") else str(val)
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CLIP
            if col in changed:
                cell.fill = RED_FILL
                cell.font = RED_FONT
    autowidth(ws, {"PipelineName": 38, "Operator [ref]": 55, "Owner [ref]": 55,
                   "AggregateOwners": 40, "Operator": 30})

    # Changes_Summary
    ws = wb.create_sheet("Changes_Summary")
    hdr = ["ProjectID", "Asset", "Field", "Current value", "Proposed value", "Action",
           "Confidence", "Key evidence", "Source URLs"]
    ws.append(hdr)
    style_header(ws, len(hdr))
    for i, row in enumerate(SUMMARY_ROWS, start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CLIP
            if j == 6 and row[5] in ("change", "fill"):
                cell.fill = RED_FILL
                cell.font = RED_FONT
    autowidth(ws, {"Field": 30, "Current value": 32, "Proposed value": 40, "Key evidence": 80,
                   "Source URLs": 60, "Confidence": 12, "Action": 9, "Asset": 15})

    out = ROOT / args.output if not Path(args.output).is_absolute() else Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
