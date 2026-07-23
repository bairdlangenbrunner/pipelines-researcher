#!/usr/bin/env python3
"""Build P8000 deliverables: tracker-format xlsx row + GEM wiki page txt."""
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SCRATCH = "/private/tmp/claude-501/-Users-baird-Dropbox--git-ALL--github-repos-gem-goit-ggit-pipeline-routes/fbd222df-7ae9-4cca-be04-c971f6f659af/scratchpad"
GEOJSON = "/Users/baird/Dropbox/_git_ALL/_github-repos-gem/goit-ggit-pipeline-routes/data/individual-routes/gas-pipelines/P8000.geojson"
OUT_XLSX = "/Users/baird/Downloads/P8000_leviathan_egypt_offshore_gas_pipeline_row.xlsx"
OUT_TXT = "/Users/baird/Downloads/P8000_leviathan_egypt_offshore_gas_pipeline_wiki.txt"

# refs (all url_verifier-checked 2026-07-22; AN via Wayback due to bot-wall 403; MEES live but paywalled)
AN = "https://web.archive.org/web/20250612230425/https://www.arabnews.com/node/1813231/middle-east"
TOI = "https://www.timesofisrael.com/israel-to-build-new-natural-gas-pipeline-from-its-offshore-rig-direct-to-egypt/"
JP = "https://www.jpost.com/israel-news/egyptian-energy-minister-in-israel-to-talk-gas-cooperation-659690"
WO = "https://www.worldoil.com/news/2021/2/20/israel-and-egypt-discuss-natural-gas-pipeline-link"
OT = "https://www.offshore-technology.com/news/israel-pipeline-exports-egypt/"
JPT = "https://jpt.spe.org/chevron-greenlights-construction-of-israels-nitzana-gas-pipeline-to-egypt"
MEES = "https://www.mees.com/2023/5/12/oil-gas/israel-egypt-onshore-gas-pipeline-advances-with-cabinet-approval/b3b61270-f0cc-11ed-b6ad-21fb4845845e"
RZ = "https://www.rigzone.com/news/new_pipeline_paves_way_for_increased_israeli_gas_exports_to_egypt-10-jul-2026-184106-article/"

rows = list(csv.reader(open(f"{SCRATCH}/gas_tab.csv")))
NCOL = len(rows[2])

notes = (
    "Status inferred via dormancy rule (>4 yr): proposed Feb 2021 (Steinitz/El-Molla "
    "ministerial agreement), no development reported since; Israel-Egypt export expansion "
    "proceeded instead via the Ashdod-El Arish third line (P3620/P3657, completed 2026) and "
    "the onshore Nitzana Pipeline (P7864, FID Sep 2025). No capacity/length/diameter/cost "
    "ever published for the subsea option - the US$200m / 3-5 bcm/y figures in Oct 2021 "
    "reporting refer to the onshore Sinai link (later Nitzana), not this line. "
    "arabnews.com returns 403 to bots - Wayback snapshot cited. mees.com ref is live but "
    "paywalled (headline supports value). Route digitized from Delek investor presentation "
    "map (main line ~485 km per digitization; unsourced, not entered as length)."
)
background = (
    "Leviathan's partners explored expanded export options including a floating LNG "
    "facility or a subsea pipeline to Egypt's underused LNG terminals. On 21 February 2021 "
    "the Israeli and Egyptian energy ministers agreed on construction of an offshore "
    "pipeline from the Leviathan field to the liquefaction facilities at Idku and Damietta "
    "to increase gas exports to Europe. The project did not advance further."
)

vals = {
    2: "Leviathan–Egypt Offshore Gas Pipeline",
    5: "P8000",
    6: "Leviathan–Idku Gas Pipeline",
    7: "cancelled",
    8: f"{JPT}, {RZ}, {MEES}",
    9: "BL",
    10: "2026-07-22",
    11: "Gas",
    12: AN,
    13: "transmission",
    14: AN,
    18: "Israel, Egypt",
    19: notes,
    20: background,
    21: f"{AN}, {WO}",
    22: "--",
    23: "unknown [unknown %]",
    24: "unknown",
    25: 2021,
    26: 2,
    27: f"{AN}, {JP}",
    44: 2025,
    45: f"{JPT}, {RZ}, {MEES}",
    48: "inferred",
    70: "Leviathan platform",
    73: "Israel",
    74: "Asia",
    75: "Western Asia",
    76: "Idku LNG terminal",
    78: "Beheira",
    79: "Egypt",
    80: "Africa",
    81: "Northern Africa",
    82: 2,
    83: f"{AN}, {OT}",
    105: "yes",
    106: "Mapped route (at any accuracy)",
    107: "low",
    108: ("Digitized from Delek investor presentation map; includes ~107 km spur toward "
         "the Aphrodite field (Cyprus), cf. P0473"),
}

GREEN = PatternFill("solid", start_color="C6EFCE")   # >=2 independent working sources
YELLOW = PatternFill("solid", start_color="FFEB9C")  # single source
ref_tiers = {8: GREEN, 12: YELLOW, 14: YELLOW, 21: GREEN, 27: GREEN, 45: GREEN, 83: GREEN}

wb = Workbook()
ws = wb.active
ws.title = "Gas (new row P8000)"
for r in rows[:3]:
    ws.append(r + [""] * (NCOL - len(r)))
new_row = ["" ] * NCOL
for i, v in vals.items():
    new_row[i] = v
ws.append(new_row)
for c in ws[3]:
    c.font = Font(bold=True)
for idx, fill in ref_tiers.items():
    ws.cell(row=4, column=idx + 1).fill = fill
ws.freeze_panes = "A4"
for col, w in {"C": 34, "F": 10, "H": 12, "T": 50, "U": 50}.items():
    ws.column_dimensions[col].width = w
wb.save(OUT_XLSX)
print("wrote", OUT_XLSX)

# ---- wiki page ----
gj = json.load(open(GEOJSON))
lines = gj["features"][0]["geometry"]["coordinates"]
map_lines = ";\n".join(":".join(f"{lat},{lon}" for lon, lat in line) for line in lines)

wiki = f"""{{{{Navbar-Global Gas Infrastructure Tracker}}}}
'''Leviathan–Egypt Offshore Gas Pipeline''' (also known as the Leviathan–Idku Gas Pipeline) was a proposed offshore natural gas pipeline running from Israel to Egypt. The project is considered cancelled.<ref name=":5" /><ref name=":6" />

==Location==
The pipeline was proposed to run from the Leviathan gas field platform, offshore Israel, to the liquefied natural gas (LNG) export terminals at Idku and Damietta on Egypt's Mediterranean coast.<ref name=":0">{{{{Cite web|url={AN}|title=Israel to link Leviathan gas field to Egypt LNG plants, minister says|date=2021-02-21|website=Arab News|access-date=2026-07-22}}}}</ref><ref name=":4">{{{{Cite web|url={OT}|title=Israel plans to build $200m pipeline to boost gas exports to Egypt|date=2021-10-22|website=Offshore Technology|access-date=2026-07-22}}}}</ref>

{{{{#display_map:
| lines=
{map_lines}
| center=
32.0,32.5
| width=400
| height=300
| type=normal
| zoom=6}}}}

==Project details==
*'''Operator: '''
*'''Owner: '''
*'''Parent company: '''
*'''Capacity: '''
*'''Length: '''
*'''Diameter: '''
*'''Status: '''Cancelled<ref name=":5" /><ref name=":6" />
*'''Start year: '''
*'''Cost: '''

==Background==
The partners in the Leviathan gas field — Chevron, NewMed Energy (formerly Delek Drilling), and Ratio — had explored options for expanding exports from the field, including a floating LNG facility or a subsea pipeline to link up with Egypt's two LNG terminals, which had been idled or running below capacity.<ref name=":0" /><ref name=":1">{{{{Cite web|url={WO}|title=Israel and Egypt discuss natural gas pipeline link|date=2021-02-20|website=World Oil|access-date=2026-07-22}}}}</ref>

On February 21, 2021, Israeli energy minister Yuval Steinitz and Egyptian petroleum minister Tarek El-Molla agreed on the construction of an offshore gas pipeline from the Leviathan field to the liquefaction facilities in Egypt, "in order to increase the gas exports to Europe through the liquefaction facilities in Egypt."<ref name=":0" /><ref name=":2">{{{{Cite web|url={TOI}|title=Israel to build new natural gas pipeline from its offshore rig direct to Egypt|date=2021-02-22|website=The Times of Israel|access-date=2026-07-22}}}}</ref><ref name=":3">{{{{Cite web|url={JP}|title=Egyptian energy minister in Israel to talk gas cooperation|date=2021-02-21|website=The Jerusalem Post|access-date=2026-07-22}}}}</ref>

As of October 2021, the subsea line was still described as planned alongside a cheaper onshore link through the northern Sinai Peninsula (estimated at US$200 million and 3–5 billion cubic meters per year); those onshore plans later advanced as the [[Nitzana Pipeline]].<ref name=":4" /> The offshore project, however, was not developed further. Expansion of Israel–Egypt export capacity instead proceeded through the [[Ashdod-El Arish Gas Pipeline]], completed in 2026, and the onshore [[Nitzana Pipeline]], which reached a final investment decision in September 2025.<ref name=":5">{{{{Cite web|url={JPT}|title=Chevron Greenlights Construction of Israel's Nitzana Gas Pipeline to Egypt|date=2025-09-16|website=Journal of Petroleum Technology|access-date=2026-07-22}}}}</ref><ref name=":6">{{{{Cite web|url={RZ}|title=New Pipeline Paves Way for Increased Israeli Gas Exports to Egypt|date=2026-07-10|website=Rigzone|access-date=2026-07-22}}}}</ref>

As of July 2026, no development on the Leviathan–Egypt Offshore Gas Pipeline has been reported for over 4 years, and it is considered cancelled.

==Articles and resources==
===References===
{{{{reflist}}}}
===Related GEM.wiki articles===
*[[Ashdod-El Arish Gas Pipeline]]
*[[Nitzana Pipeline]]
*[[El Arish–Ashkelon Pipeline]]
*[[Cyprus–Egypt Gas Pipeline]]
*[[Leviathan Subsea Gas Pipelines]]
*[[Egyptian LNG Terminal]]
*[[Damietta Segas LNG Terminal]]
===Additional data===
To access additional data, including an interactive map of gas pipelines, a downloadable dataset, and summary data, please visit the [https://globalenergymonitor.org/projects/global-gas-infrastructure-tracker/ Global Gas Infrastructure Tracker] on the Global Energy Monitor website.

[[Category:Global Gas Infrastructure Tracker]]
[[Category:Gas and Hydrogen Pipelines]]
"""
open(OUT_TXT, "w").write(wiki)
print("wrote", OUT_TXT)
