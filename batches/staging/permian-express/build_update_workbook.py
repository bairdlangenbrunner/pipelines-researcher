#!/usr/bin/env python3
"""Build the Permian Express update workbook from the 2026-06-11 re-research.

Ad-hoc update-mode builder (4 rows, P0113/P2581/P2660/P2661) per
docs/reference/workbook_conventions.md: changed/filled cells red
(FFCCCC fill / CC0000 font), re-verified-unchanged cells blue (4472C4 fill,
white font), README first, no wrap anywhere (clip).

Usage: python batches/staging/permian-express/build_update_workbook.py \
           --output batches/pipelines_batch_<stamp>_permian-express_update.xlsx
"""
import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
SNAP_OIL = ROOT / "data/GOIT_oil_ngl_snapshot_20260611.csv"
SNAP_OO = ROOT / "data/GEM_operators_owners_snapshot_20260611.csv"
STAGED = Path(__file__).with_name("staged_updates.json")

PIDS = ["P0113", "P2581", "P2660", "P2661"]

RED_FILL = PatternFill("solid", start_color="FFCCCC")
RED_FONT = Font(color="CC0000", bold=True)
BLUE_FILL = PatternFill("solid", start_color="4472C4")
WHITE_FONT = Font(color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color="4472C4")
HDR_FONT = Font(color="FFFFFF", bold=True)
CLIP = Alignment(wrap_text=False, vertical="top")
HDR_ALIGN = Alignment(wrap_text=False, vertical="center", horizontal="center")

# refs used below (all passed scripts/url_verifier.py on 2026-06-11)
ET_10K_25 = "https://www.sec.gov/Archives/edgar/data/1276187/000127618726000013/et-20251231.htm"
XOM_EX21 = "https://www.sec.gov/Archives/edgar/data/34088/000003408826000045/xomexhibit21123125.htm"
RRC_TARIFF = "https://www.rrc.texas.gov/media/iqrly2vr/5-11-0.pdf"
SXL_10K_16 = "https://www.sec.gov/Archives/edgar/data/1161154/000116115417000014/a2016form10-k.htm"
EIA_XLSX = "https://www.eia.gov/petroleum/xls/EIA_LiqPipProject.xlsx"
OGJ_2012 = "https://www.ogj.com/pipelines-transportation/pipelines/article/17273965/sunoco-begins-permian-express-oil-line-open-season"
HART_2012 = "http://web.archive.org/web/20250818112533/https://www.hartenergy.com/exclusives/more-one-way-use-pipeline-4206"
SXL_8K_Q312 = "https://www.sec.gov/Archives/edgar/data/1161154/000119312512457643/d434020dex991.htm"
SXL_8K_Q213 = "https://www.sec.gov/Archives/edgar/data/1161154/000116115413000009/q22013pressrelease8-kexhib.htm"
NASM_WB = "http://web.archive.org/web/20170624010832/http://northamericanshalemagazine.com:80/articles/1949/energy-transfer-partners-to-build-permian-express-3-pipeline"
GD_PE1 = "https://www.offshore-technology.com/marketdata/permian-express-phase-i-oil-pipeline-the-us/"
GD_PE2 = "https://www.offshore-technology.com/marketdata/permian-express-phase-ii-oil-pipeline-the-us/"
SXL_10K_15 = "https://www.sec.gov/Archives/edgar/data/1161154/000116115416000062/a2015form10-k.htm"
SXL_8K_Q313 = "https://www.sec.gov/Archives/edgar/data/1161154/000116115413000024/q32013pressrelease8-kexhib.htm"
SXL_8K_Q215 = "https://www.sec.gov/Archives/edgar/data/1161154/000116115415000031/q22015pressrelease8-kexhib.htm"
RBN_PE2_WB = "https://web.archive.org/web/20231205223128/https://rbnenergy.com/node/4536"
ET_8K_Q318 = "https://www.sec.gov/Archives/edgar/data/1276187/000127618718000064/ex991eterq32018.htm"
ET_DECK_AUG18 = "https://www.sec.gov/Archives/edgar/data/1161154/000119312518249842/d607097d425.htm"
ET_CALL_Q218 = "https://www.sec.gov/Archives/edgar/data/1161154/000119312518245266/d609237d425.htm"
BW_PE3_WB = "http://web.archive.org/web/20210121091919/https://www.businesswire.com/news/home/20180529005848/en/Energy-Transfer-Launches-Marketing-Process-Additional-Commitments"
ET_DECK_SEP18 = "https://www.sec.gov/Archives/edgar/data/1161154/000119312518267597/d620886d425.htm"
ET_8K_Q219 = "https://www.sec.gov/Archives/edgar/data/1276187/000127618719000030/ex991eterq22019.htm"
ET_8K_Q319 = "https://www.sec.gov/Archives/edgar/data/1276187/000127618719000047/ex991eterq32019.htm"
ET_10K_19 = "https://www.sec.gov/Archives/edgar/data/1276187/000127618720000011/et12-31x201910k.htm"
RBN_PE4_WB = "http://web.archive.org/web/20250619182512/https://rbnenergy.com/node/25491"
FOOL_Q319 = "https://www.fool.com/earnings/call-transcripts/2019/11/07/energy-transfer-lp-et-q3-2019-earnings-call-transc.aspx"

PARENT_NEW = "Energy Transfer LP [87.70%]; Exxon Mobil Corp [12.30%]"
TODAY = "2026-06-11"


def refs(*urls):
    return ", ".join(urls)


NOTES = {
    "P0113": (
        "[2026-06-11 CB] Phase-level re-research. Capacity corrected 200,000->150,000 bpd: PE1 entered service "
        "Q2 2013 at 90,000 bpd, ramping to full 150,000 bpd late 2013/early 2014 (SXL Q3 2012 + Q2 2013 8-Ks; "
        "EIA Liquids Pipeline Projects DB: +90k 2013 Q2 plus PE1 expansion +60k 2014 Q1 = 150k; NASM 2017 "
        "retrospective). The 200,000 bpd figure belongs to PE2 (SXL FY2013 10-K; Alon USA Partners 10-K). Tier: "
        "high (2+ independent). PE1 is not a single new-build: reversed Wichita Falls-Wortham line + excess "
        "capacity on West Texas Gulf southern leg, Wichita Falls->Nederland/Beaumont (OGJ 2012-06-27; Hart "
        "Energy 2012-10-01). Length 300 mi UNVERIFIED: only published figure is GlobalData 611 km (~380 mi, "
        "single source); GEM route-estimate 641 km - review recommended, not changed (no 2-source figure). "
        "Diameter left blank (GlobalData max 16 in, single source, low tier). Parent corrected to ET 87.7% / "
        "XOM 12.3% per both partners' FY2025 10-Ks (independent filers, sum exactly 100); operator Sunoco "
        "Pipeline L.P. (TX RRC tariff, P-5 ID 829627) staged on operators/owners tab."
    ),
    "P2581": (
        "[2026-06-11 CB] StartLocation corrected Wichita Falls->Midland: PE2 origins are Midland, Garden City "
        "and Colorado City (SXL FY2015 10-K; EIA), connected via the SunVit lateral from Midland (OGJ "
        "2016-11-10); Wichita Falls is PE1's origin and appears in no PE2 description in any SXL/ET filing "
        "FY2013-FY2025. New-build mainline Garden City->Colorado City (20 in)->Corsicana (24 in), 334 mi (EIA; "
        "GlobalData 537 km); deliveries reach Nederland over the existing Sunoco/ET system. Capacity 230,000 "
        "bpd confirmed (EIA; RBN) - operator announced ~200,000 bpd at the July 2015 startup, 230k is the "
        "post-startup nameplate carried by both independent compilers. In service July 2015 (SXL Q2 2015 8-K; "
        "EIA Q3 2015). Tier: high. Parent corrected to ET 87.7% / XOM 12.3% (both partners' FY2025 10-Ks)."
    ),
    "P2660": (
        "[2026-06-11 CB] All values confirmed as-is. Expansion treatment (LengthKnown=0, Diameter blank) "
        "CONFIRMED: ET Aug 2018 investor deck (SEC Form 425) - PE3 provides takeaway 'utilizing existing "
        "pipelines'; EIA carries no mileage or diameter for PE3 (its 'New' project-type label = new "
        "service/capacity, not new pipe); no source anywhere assigns PE3 a length or diameter. Capacity "
        "140,000 bpd (ETP May 2018 release, archived; ET Aug 2018 call transcript, SEC 425; EIA). Partial "
        "service ~100,000 bpd Q4 2017 (ET Q3 2018 8-K; EIA); final tranche: EIA says Q3 2018 and ET FY2018 "
        "10-K says 'fully operational in September 2018', while ET's Q4 2018 release books the PE3 expansion "
        "in service in Q4 2018 - StartYear1=2017 unaffected. Tier: high (capacity/status/start), medium-high "
        "(no-new-pipe: one explicit primary statement + consistent independent indirect evidence). Parent "
        "corrected to ET 87.7% / XOM 12.3% (both partners' FY2025 10-Ks)."
    ),
    "P2661": (
        "[2026-06-11 CB] Reclassified as expansion with no new mainline pipe: LengthKnown 400 mi->0, Diameter "
        "24 in->blank (tracker convention). ET deck Sept 2018 (SEC Form 425): 'Permian Express 4 Expansion "
        "Project (formerly PE3 Phase II)'; ET COO McCrea, Q2 2018 call (SEC 425), in a discussion of squeezing "
        "capacity from existing Permian pipes with drag-reducing agents: 'we are looking at expanding Perm "
        "Express 3, which would be Permian Express 4'; every ET quarterly release calls it 'the Permian "
        "Express 4 expansion'. EIA lists project type 'Expansion' and EIA's own Definitions sheet says that "
        "for pump-station expansions the Miles column shows 'length of pipeline affected' - so EIA's 400 mi "
        "(the evident origin of the previous GEM value) is affected-mainline length, not new pipe; the 24-in "
        "is the existing Colorado City-Corsicana mainline. RBN lists 334 mi (= the PE2 mainline); the "
        "334-vs-400 disagreement between compilers is itself a tell. No construction contract, open season "
        "for new pipe, or route filing found. +120,000 bpd from Colorado City to Nederland (ET CFO Long, Q3 "
        "2019 call); partial service May 2019, full service 2019-10-01 (ET Q3 2019 8-K; FY2019 10-K; EIA). "
        "Tier: high."
    ),
}

# tracker-tab cell changes/fills (red). values written as paste-ready strings
CHANGES = {
    "P0113": {
        "Capacity": "150,000.00",
        "CapacityBOEd": "150000.00",
        "Capacity [ref]": refs(SXL_8K_Q312, NASM_WB, EIA_XLSX),
        "OtherEnglishNames": "Permian Express 1",
        "EndLocation": "Nederland",
        "EndLocation [ref]": refs(OGJ_2012, HART_2012),
        "StartLocation [ref]": refs(OGJ_2012, HART_2012),
        "Status [ref]": refs(ET_10K_25, RRC_TARIFF),
        "Start [ref]": refs(SXL_8K_Q213, GD_PE1),
        "FuelSource [ref]": refs(HART_2012, ET_10K_25),
        "Parent": PARENT_NEW,
        "ResearcherNotes": NOTES["P0113"],
        "LastUpdated": TODAY,
    },
    "P2581": {
        "StartLocation": "Midland",
        "StartLocation [ref]": refs(SXL_10K_15, EIA_XLSX),
        "OtherEnglishNames": "Permian Express 2",
        "EndLocation": "Nederland",
        "EndLocation [ref]": refs(SXL_8K_Q313, RBN_PE2_WB),
        "Status [ref]": refs(ET_10K_25, GD_PE2),
        "Start [ref]": refs(SXL_8K_Q215, EIA_XLSX),
        "Capacity [ref]": refs(EIA_XLSX, RBN_PE2_WB),
        "Length [ref]": refs(EIA_XLSX, GD_PE2),
        "Diameter [ref]": refs(EIA_XLSX, GD_PE2),
        "FuelSource [ref]": refs(SXL_10K_15, EIA_XLSX),
        "Parent": PARENT_NEW,
        "ResearcherNotes": NOTES["P2581"],
        "LastUpdated": TODAY,
    },
    "P2660": {
        "OtherEnglishNames": "Permian Express 3",
        "Status [ref]": refs(ET_10K_25, EIA_XLSX),
        "Start [ref]": refs(ET_8K_Q318, EIA_XLSX),
        "Capacity [ref]": refs(BW_PE3_WB, EIA_XLSX),
        "Length [ref]": refs(ET_DECK_AUG18, EIA_XLSX),
        "Parent": PARENT_NEW,
        "ResearcherNotes": NOTES["P2660"],
        "LastUpdated": TODAY,
    },
    "P2661": {
        "LengthKnown": "0",
        "LengthKnownUnits": "km",
        "LengthKnownKm": "0.00",
        "LengthMergedKm": "0.00",
        "Length [ref]": refs(ET_DECK_SEP18, EIA_XLSX),
        "Diameter": "",
        "DiameterUnits": "",
        "DiameterInMm": "--",
        "StartLocation": "Colorado City",
        "StartLocation [ref]": refs(FOOL_Q319),
        "EndLocation": "Nederland",
        "EndLocation [ref]": refs(FOOL_Q319, RBN_PE4_WB),
        "FuelSource": "Permian Basin",
        "FuelSource [ref]": refs(ET_8K_Q219, RBN_PE4_WB),
        "OtherEnglishNames": "Permian Express 4",
        "Status [ref]": refs(ET_8K_Q319, RBN_PE4_WB),
        "Start [ref]": refs(ET_10K_19, EIA_XLSX),
        "Capacity [ref]": refs(ET_8K_Q219, EIA_XLSX),
        "Parent": PARENT_NEW,
        "ResearcherNotes": NOTES["P2661"],
        "LastUpdated": TODAY,
    },
}

# value cells confirmed unchanged this batch (blue)
REVERIFIED = {
    "P0113": ["Status", "StartYear1", "StartLocation", "FuelSource", "Owner"],
    "P2581": ["Status", "StartYear1", "Capacity", "LengthKnown", "Diameter", "FuelSource", "Owner"],
    "P2660": ["Status", "StartYear1", "Capacity", "LengthKnown", "Owner"],
    "P2661": ["Status", "StartYear1", "Capacity", "Owner"],
}

# operators/owners tab changes (red)
OO_CHANGES = {
    pid: {
        "Operator": "Sunoco Pipeline L.P.",
        "Operator [ref]": refs(RRC_TARIFF, SXL_10K_16),
        "Owner [ref]": refs(ET_10K_25, XOM_EX21),
    }
    for pid in PIDS
}

SUMMARY_ROWS = [
    # pid, segment, field, current, proposed, action, tier, evidence
    ("P0113", "Phase I", "Capacity", "200,000 bpd", "150,000 bpd", "change", "high",
     "PE1 = 90k bpd at Q2 2013 startup, full 150k late 2013/early 2014 (SXL 8-Ks; EIA: +90k 2013 + 60k expansion 2014; NASM). 200k is PE2's figure (SXL FY2013 10-K; Alon USA 10-K).",
     refs(SXL_8K_Q312, NASM_WB, EIA_XLSX)),
    ("P0113", "Phase I", "OtherEnglishNames", "Permian Express 1; Permian Express 2", "Permian Express 1", "change", "-",
     "PE1/PE2 are separate rows; dual naming on both rows likely caused the capacity mix-up.", ""),
    ("P0113", "Phase I", "EndLocation", "(blank)", "Nederland", "fill", "high",
     "OGJ 2012: 'continuous pipeline service from Wichita Falls, Tex., to Nederland and Beaumont'; Hart Energy 2012.", refs(OGJ_2012, HART_2012)),
    ("P0113", "Phase I", "Parent", "ET 88.00% / XOM 12.00%", "ET 87.70% / XOM 12.30%", "change", "high",
     "ET FY2025 10-K: 'Permian Express Pipelines 87.7%'; XOM FY2025 Ex-21: 'Permian Express Partners LLC 12.3'. Independent filers, sum exactly 100. (Applies to all 4 rows.)",
     refs(ET_10K_25, XOM_EX21)),
    ("P0113", "Phase I", "LengthKnown", "300 mi", "(no change - FLAG)", "review", "low",
     "300 mi not found in any non-GEM source. GlobalData: 611 km (~380 mi, single source); GEM route-estimate 641 km. PE1 = reversed Wichita Falls-Wortham + WTG southern leg, so length depends on segments counted.",
     GD_PE1),
    ("P0113", "Phase I", "Diameter", "(blank)", "(no change - keep blank)", "review", "low",
     "GlobalData max 16 in is the only source (single, low tier). No operator filing gives PE1 diameter.", GD_PE1),
    ("P2581", "Phase II", "StartLocation", "Wichita Falls", "Midland", "change", "high",
     "SXL FY2015 10-K: PE2 'origins in multiple locations in Western Texas: Midland, Garden City and Colorado City'; EIA new-build = Garden City->Colorado City->Corsicana. Wichita Falls is PE1's origin; appears in no PE2 filing.",
     refs(SXL_10K_15, EIA_XLSX)),
    ("P2581", "Phase II", "OtherEnglishNames", "Permian Express 1; Permian Express 2", "Permian Express 2", "change", "-",
     "Same naming cleanup as P0113.", ""),
    ("P2581", "Phase II", "EndLocation", "(blank)", "Nederland", "fill", "high",
     "SXL Q3 2013 8-K (Gulf Coast markets 'including our Nederland terminal'); RBN: 'Destination: Nederland, TX'. New-build mainline ends Corsicana; Nederland reached over existing system (noted).",
     refs(SXL_8K_Q313, RBN_PE2_WB)),
    ("P2581", "Phase II", "Capacity / Length / Diameter / Status / StartYear1", "(unchanged)", "confirmed", "confirm", "high",
     "230,000 bpd (EIA; RBN; initial ~200k at July 2015 startup); 334 mi (EIA; GlobalData 537 km); 20/24 in (EIA segment-explicit; GlobalData max 24 in); operating; 2015.",
     refs(EIA_XLSX, RBN_PE2_WB, GD_PE2)),
    ("P2660", "Phase III", "all fields", "(unchanged)", "confirmed", "confirm", "high",
     "Expansion treatment (length 0, no diameter) confirmed: ET deck 'utilizing existing pipelines'; EIA has no pipe specs for PE3. 140,000 bpd; partial 100k Q4 2017 + final 40k 2018; operating.",
     refs(ET_DECK_AUG18, EIA_XLSX)),
    ("P2661", "Phase IV", "LengthKnown", "400 mi", "0", "change", "high",
     "PE4 was an expansion ('formerly PE3 Phase II', ET deck Sept 2018; 'Permian Express 4 expansion' in every ET release; EIA type 'Expansion'). EIA's 400 mi = 'length of pipeline affected' per EIA's own definition, not new pipe. RBN says 334 mi (= PE2 mainline).",
     refs(ET_DECK_SEP18, EIA_XLSX)),
    ("P2661", "Phase IV", "Diameter", "24 in", "(blank)", "change", "high",
     "No new pipe -> blank per convention. 24-in is the existing Colorado City-Corsicana mainline (EIA PE2 note).", refs(ET_DECK_SEP18, EIA_XLSX)),
    ("P2661", "Phase IV", "StartLocation", "Permian", "Colorado City", "change", "medium",
     "ET CFO Long, Q3 2019 call: PE4 added 120k bpd 'from Colorado City to Nederland, Texas'. Single explicit source for corridor start (RBN says Midland as system receipt origin) -> medium; keep 'Permian' if system-origin framing preferred.",
     FOOL_Q319),
    ("P2661", "Phase IV", "EndLocation", "(blank)", "Nederland", "fill", "high",
     "ET CFO quote; RBN: 'Destination: Energy Transfer Nederland Terminal, TX'.", refs(FOOL_Q319, RBN_PE4_WB)),
    ("P2661", "Phase IV", "FuelSource", "(blank)", "Permian Basin", "fill", "high",
     "ET Q2 2019 8-K: 'adds 120,000 barrels per day of capacity from the Permian Basin'; RBN.", refs(ET_8K_Q219, RBN_PE4_WB)),
    ("P2661", "Phase IV", "OtherEnglishNames", "(blank)", "Permian Express 4", "fill", "-",
     "Naming consistency.", ""),
    ("ALL", "I-IV", "Operator (operators/owners tab)", "(blank)", "Sunoco Pipeline L.P.", "fill", "high",
     "TX RRC tariff TX No. 5.11.0 (eff. 2024-07-01): 'Operated under PEP's T-4 Permit No 09001 and Sunoco Pipeline L.P.'s P5 ID 829627'; SXL FY2016 10-K: 'we... are the operator of all of the assets' (SPLP = wholly owned ET subsidiary per ET FY2025 10-K).",
     refs(RRC_TARIFF, SXL_10K_16)),
    ("ALL", "I-IV", "Owner [ref] (operators/owners tab)", "(blank)", "ET FY2025 10-K + XOM Ex-21", "fill", "high",
     "Owner = Permian Express Partners LLC [100%] confirmed; PEP active in both partners' Feb 2026 filings, PE 1-4 enumerated in the JV.", refs(ET_10K_25, XOM_EX21)),
]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
    ws.freeze_panes = "A2"


def autowidth(ws, widths=None, default=14, cap=55):
    widths = widths or {}
    for c in range(1, ws.max_column + 1):
        hdr = ws.cell(row=1, column=c).value or ""
        ws.column_dimensions[get_column_letter(c)].width = widths.get(hdr, min(max(default, len(str(hdr)) + 2), cap))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    oil = pd.read_csv(SNAP_OIL, header=2, low_memory=False, dtype=str)
    oo = pd.read_csv(SNAP_OO, header=1, low_memory=False, dtype=str)
    oil_rows = {r["ProjectID"]: r for _, r in oil[oil["ProjectID"].isin(PIDS)].iterrows()}
    oo_rows = {r["ProjectID"]: r for _, r in oo[oo["ProjectID"].isin(PIDS)].iterrows()}

    wb = Workbook()

    # README
    ws = wb.active
    ws.title = "README"
    readme = [
        ["Permian Express Oil Pipeline - update batch"],
        [""],
        ["Mode", "update"],
        ["Scope", "Permian Express Oil Pipeline, all 4 phases (P0113, P2581, P2660, P2661), GOIT oil tracker, United States"],
        ["Researched", TODAY],
        ["GEM snapshot", SNAP_OIL.name],
        ["Staging", "batches/staging/permian-express/staged_updates.json"],
        ["URL verification", "Every URL passed scripts/url_verifier.py on 2026-06-11. Note: sec.gov 403s the verifier's default User-Agent; verified with a declared-contact UA per SEC fair-access policy."],
        [""],
        ["Color key"],
        ["red fill / red font", "changed or newly filled cell - the thing to review and paste"],
        ["blue fill / white font", "value unchanged but re-verified this batch (>=2 sources unless noted)"],
        [""],
        ["Sheets"],
        ["Oil_Updated", "All 4 rows, full GOIT column layout (paste-ready). Changed/filled cells red; re-verified values blue. ResearcherNotes carries tier + evidence per row."],
        ["Oil_OperatorsOwners", "The 4 rows as they appear on the 'Pipeline operators/owners' tab (GID 1489950650), ProjectID-keyed. Operator + Operator [ref] + Owner [ref] fills in red. Paste by ProjectID."],
        ["Changes_Summary", "One line per finding: field, current vs proposed, action, confidence tier, key evidence, sources. Includes the two FLAG-only items (PE1 length, PE1 diameter) where no change is proposed."],
        [""],
        ["Headline findings"],
        ["1", "P0113 Phase I capacity 200,000 -> 150,000 bpd (200k is PE2's figure; PE1 = 90k at 2013 startup -> 150k full)."],
        ["2", "P2581 Phase II StartLocation Wichita Falls -> Midland (PE2 originates Midland/Garden City/Colorado City; Wichita Falls is PE1's origin)."],
        ["3", "P2661 Phase IV reclassified as expansion: LengthKnown 400 mi -> 0, Diameter 24 in -> blank ('formerly PE3 Phase II'; EIA's 400 mi = 'length of pipeline affected')."],
        ["4", "Parent split 88/12 -> 87.7/12.3 on all 4 rows (ET + XOM FY2025 10-Ks, independent filers, sum exactly 100)."],
        ["5", "Operator filled = Sunoco Pipeline L.P. on all 4 rows (TX RRC tariff P-5 ID 829627 + SEC operator language) - operators/owners tab."],
        ["6", "P2660 Phase III fully confirmed as-is (incl. its 0-length expansion treatment)."],
        ["7", "FLAGS (no change staged): PE1 length 300 mi unverifiable (GlobalData says 611 km, single source; route-estimate 641 km); PE1 diameter (GlobalData max 16 in, single source)."],
        [""],
        ["Note", "Researcher initials left as-is (BL) on the rows; flip to CB on apply if preferred."],
        ["Escalation gates", "None tripped (4-row targeted batch; conflicts are corrections with 2+ independent sources)."],
    ]
    for r in readme:
        ws.append(r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 160
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = CLIP
    ws["A1"].font = Font(bold=True, size=14)

    # Oil_Updated
    ws = wb.create_sheet("Oil_Updated")
    cols = list(oil.columns)
    ws.append(cols)
    style_header(ws, len(cols))
    for i, pid in enumerate(PIDS, start=2):
        base = oil_rows[pid]
        changed = CHANGES[pid]
        for j, col in enumerate(cols, start=1):
            val = changed.get(col, base.get(col))
            val = "" if (val is None or (isinstance(val, float) and pd.isna(val)) or str(val) == "nan") else str(val)
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CLIP
            if col in changed:
                cell.fill = RED_FILL
                cell.font = RED_FONT
            elif col in REVERIFIED[pid]:
                cell.fill = BLUE_FILL
                cell.font = WHITE_FONT
    autowidth(ws, {"PipelineName": 45, "SegmentName": 50, "Status [ref]": 55, "Owner": 55,
                   "ResearcherNotes": 55, "OtherEnglishNames": 40, "Parent": 50})

    # Oil_OperatorsOwners
    ws = wb.create_sheet("Oil_OperatorsOwners")
    oo_cols = list(oo.columns)
    ws.append(oo_cols)
    style_header(ws, len(oo_cols))
    for i, pid in enumerate(PIDS, start=2):
        base = oo_rows[pid]
        changed = OO_CHANGES[pid]
        for j, col in enumerate(oo_cols, start=1):
            val = changed.get(col, base.get(col))
            val = "" if (val is None or (isinstance(val, float) and pd.isna(val)) or str(val) == "nan") else str(val)
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CLIP
            if col in changed:
                cell.fill = RED_FILL
                cell.font = RED_FONT
    autowidth(ws, {"PipelineName": 45, "SegmentName": 20, "Operator [ref]": 55, "Owner [ref]": 55,
                   "AggregateOwners": 40, "Operator": 28})

    # Changes_Summary
    ws = wb.create_sheet("Changes_Summary")
    hdr = ["ProjectID", "Segment", "Field", "Current value", "Proposed value", "Action",
           "Confidence", "Key evidence", "Source URLs"]
    ws.append(hdr)
    style_header(ws, len(hdr))
    for i, row in enumerate(SUMMARY_ROWS, start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CLIP
            if j == 5 and row[5] in ("change", "fill"):
                cell.fill = RED_FILL
                cell.font = RED_FONT
    autowidth(ws, {"Field": 30, "Current value": 32, "Proposed value": 32, "Key evidence": 80,
                   "Source URLs": 60, "Confidence": 11, "Action": 9})

    out = ROOT / args.output if not Path(args.output).is_absolute() else Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
