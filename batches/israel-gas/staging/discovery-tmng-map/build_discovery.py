#!/usr/bin/env python3
"""Build the Israel-gas INGL-map discovery + validation deliverable.

Reads the staged JSON (canonical pending state) and the fresh GGIT snapshot
header, and writes:
  * a discovery xlsx with
      - Gas_NewRows  : full backend mirror (every gas column in sheet order,
                       NO leading SheetRow locator) for P8001 + P8003, so each
                       column aligns 1:1 with the live sheet for paste-at-bottom;
                       [ref] cells tier-colored (green >=2 independent, yellow 1).
      - Gas_EditFlags: the Phase-4 validation candidates (P0462/P0479/P5276/
                       P7604/P7606) as one flag per proposed change — candidates
                       routed through the Update SOP, never auto-applied.
  * per-row GEM wiki-page text for P8001 and P8003.

Column mapping is by NAME off the fresh header (drift-safe). All cells clip
(wrap_text=False) per workbook conventions.
"""
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path("/Users/baird/Dropbox/_git_ALL/_github-repos-gem/pipelines-researcher")
HERE = REPO / "batches/israel-gas/staging/discovery-tmng-map"
VALID = REPO / "batches/israel-gas/staging/validation-tmng-map"
SNAP = REPO / "data/GGIT_gas_snapshot_20260723.csv"
ROUTE = REPO / "batches/israel-gas/staging/route-creation-tmng-map/candidate_routes/P8003.geojson"
DELIV = REPO / "batches/israel-gas/deliverables"
STAMP = "20260723_1606_ET"

XLSX = DELIV / f"pipelines_batch_{STAMP}_israel-gas_discovery.xlsx"
WIKI_P8001 = DELIV / f"pipelines_batch_{STAMP}_israel-gas_p8001-mari-b-ashdod-wiki.txt"
WIKI_P8003 = DELIV / f"pipelines_batch_{STAMP}_israel-gas_p8003-karish-tanin-wiki.txt"

GREEN = PatternFill("solid", start_color="C6EFCE")   # >=2 independent working sources
YELLOW = PatternFill("solid", start_color="FFEB9C")  # single source
CLIP = Alignment(wrap_text=False)


def col_index(header):
    return {h.strip(): i for i, h in enumerate(header) if h.strip()}


def n_sources(cell):
    return len([u for u in str(cell).split(",") if u.strip().startswith("http")])


def main():
    rows = list(csv.reader(open(SNAP)))
    header = rows[2]
    NCOL = len(header)
    idx = col_index(header)

    staged = json.loads((HERE / "staged_new.json").read_text())["candidates"]
    edits = json.loads((VALID / "staged_edits.json").read_text())["edits"]

    wb = Workbook()

    # ---- Gas_NewRows : backend mirror ----
    ws = wb.active
    ws.title = "Gas_NewRows"
    for r in rows[:3]:
        ws.append(r + [""] * (NCOL - len(r)))
    for c in ws[3]:
        c.font = Font(bold=True)

    ref_cols = {name for name in idx if name.endswith("[ref]")}
    for cand in staged:
        vals = cand["values"]
        row = [""] * NCOL
        for name, v in vals.items():
            if name in idx:
                row[idx[name]] = v
        ws.append(row)
        rnum = ws.max_row
        for name, v in vals.items():
            if name in ref_cols and name in idx and str(v).strip():
                ws.cell(row=rnum, column=idx[name] + 1).fill = (
                    GREEN if n_sources(v) >= 2 else YELLOW)
    ws.freeze_panes = "A4"
    for c in ws.iter_rows():
        for cell in c:
            cell.alignment = CLIP
    for col, w in {"C": 30, "F": 9, "G": 34, "T": 60, "U": 60}.items():
        ws.column_dimensions[col].width = w

    # ---- Gas_EditFlags : Phase-4 validation candidates ----
    we = wb.create_sheet("Gas_EditFlags")
    cols = ["ProjectID", "PipelineName", "Confidence", "Field", "Type",
            "From / current", "To / note", "[ref]"]
    we.append(cols)
    for c in we[1]:
        c.font = Font(bold=True)
    for e in edits:
        for ch in e["changes"]:
            to = ch.get("to", ch.get("note", ch.get("value", "")))
            frm = ch.get("from", ch.get("value", "") if ch.get("type") == "no_change" else "")
            we.append([e["project_id"], e["pipeline_name"], e.get("confidence", ""),
                       ch["field"], ch["type"], frm, to, ch.get("ref", "")])
            rnum = we.max_row
            if ch.get("type") == "value_change":
                we.cell(row=rnum, column=7).fill = GREEN if n_sources(ch.get("ref", "")) >= 2 else YELLOW
            elif ch.get("type") in ("value_add", "flag"):
                we.cell(row=rnum, column=7).fill = YELLOW
    we.freeze_panes = "A2"
    for c in we.iter_rows():
        for cell in c:
            cell.alignment = CLIP
    for col, w in {"A": 10, "B": 34, "C": 11, "D": 22, "E": 13, "F": 40, "G": 80, "H": 60}.items():
        we.column_dimensions[col].width = w

    # staged_store-conformant mirror of the validation edits (so staged_summary
    # counts them). staged_edits.json stays the richer human source of record.
    write_updates_mirror(edits)

    DELIV.mkdir(exist_ok=True)
    wb.save(XLSX)
    print("wrote", XLSX.name)

    # ---- wiki texts ----
    p8001 = next(c for c in staged if c["project_id"] == "P8001")["values"]
    WIKI_P8001.write_text(wiki_p8001(p8001))
    print("wrote", WIKI_P8001.name)

    p8003 = next(c for c in staged if c["project_id"] == "P8003")["values"]
    WIKI_P8003.write_text(wiki_p8003(p8003))
    print("wrote", WIKI_P8003.name)


def write_updates_mirror(edits):
    rows = {}
    for e in edits:
        row = {"pipeline": e["pipeline_name"], "changes": {}}
        for ch in e["changes"]:
            col = ch["field"]
            k, i = col, 2
            while k in row["changes"]:
                k, i = f"{col} ({i})", i + 1
            typ = ch.get("type", "")
            is_val = typ in ("value_change", "value_add")
            row["changes"][k] = {
                "old": ch.get("from", ""),
                "new": (ch.get("to", ch.get("value", "")) if is_val else ""),
                "action": typ,
                "tier": e.get("confidence", ""),
                "refs": [u.strip() for u in str(ch.get("ref", "")).split(",")
                         if u.strip().startswith("http")],
                "evidence": ("" if is_val else (ch.get("note", "") or ch.get("to", ""))),
            }
        rows[e["project_id"]] = row
    out = {"meta": {"scope": {"tracker": "gas", "country": "Israel"},
                    "mode": "update", "leg": "validation",
                    "batch": "israel-gas / validation-tmng-map",
                    "note": ("INGL/TMNG-map Phase-4 validation. Candidate edits for the "
                             "Update SOP — none auto-applied. Human-readable source of "
                             "record is staged_edits.json (richer per-change typing); "
                             "this file is the staged_store-conformant mirror that "
                             "staged_summary.py counts."),
                    "generated": "2026-07-23"},
           "rows": rows}
    p = VALID / "staged_updates.json"
    p.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print("wrote", p.name, f"({len(rows)} rows,",
          sum(len(r['changes']) for r in rows.values()), "changes)")


def wiki_p8001(v):
    OM = "https://www.offshore-mag.com/subsea/article/16761306/tamar-a-triumph-for-both-noble-and-israel"
    ERM = "https://www3.dfc.gov/Environment/EIA/nobletamar/Summary_Disclosure_Document_Tamar.pdf"
    HE = "https://www.hebrewenergy.com/yam-tethys-ltd-yam-tethys-project/"
    TOI = "https://www.timesofisrael.com/natural-gas-from-tamar-field-reaches-israel/"
    YN = "https://www.ynetnews.com/business/article/bjjbbl5a11e"
    return f"""{{{{Navbar-Global Gas Infrastructure Tracker}}}}
'''Mari-B–Ashdod Gas Pipeline''' (also known as the Yam Tethys Pipeline, Mari-B Pipeline, or Tamar–Ashdod Pipeline) is an operating offshore natural gas pipeline in Israel carrying gas from offshore production platforms to the Ashdod Onshore Terminal.<ref name=":0" /><ref name=":1" />

==Location==
The pipeline runs from the Mari-B production platform, offshore Israel, northeast to the Ashdod Onshore Terminal (AOT), where it feeds Israel's national gas transmission grid.<ref name=":1" /><ref name=":0" />

==Project details==
*'''Operator: '''
*'''Owner: '''
*'''Parent company: '''
*'''Capacity: '''~6 billion cubic meters per year (original Mari-B design)<ref name=":2" />
*'''Length: '''
*'''Diameter: '''30 inches (762 mm)<ref name=":0" /><ref name=":1" />
*'''Status: '''Operating<ref name=":0" /><ref name=":3" />
*'''Start year: '''2004<ref name=":2" /><ref name=":4" />

==Background==
The pipeline was built for the Yam Tethys project (the Mari-B, Noa and Pinnacles fields, about 20 km offshore Israel) under a transmission licence issued by Israel's Minister of Energy on 29 April 2002 for a gas pipeline from the production platform to the Ashdod Onshore Terminal. The line was completed at the end of 2003 and Mari-B delivered first gas in February 2004.<ref name=":2">{{{{Cite web|url={HE}|title=Yam Tethys Ltd / Yam Tethys Project|website=HebrewEnergy|access-date=2026-07-23}}}}</ref>

When the much larger Tamar field was developed, its platform was installed next to the Mari-B platform in 2013 and tied into this existing 30-inch export line via a subsea hot-tap, rather than building a new export pipeline; the line has since carried Tamar's gas to Ashdod.<ref name=":0">{{{{Cite web|url={OM}|title=Tamar a triumph for both Noble and Israel|website=Offshore Magazine|date=2013|access-date=2026-07-23}}}}</ref><ref name=":1">{{{{Cite web|url={ERM}|title=Tamar Expansion: Disclosure Summary of Environmental and Social Assessments|website=ERM/Noble Energy for OPIC|date=2015|access-date=2026-07-23}}}}</ref><ref name=":4">{{{{Cite web|url={TOI}|title=Natural gas from Tamar field reaches Israel|website=The Times of Israel|date=2013-03-31|access-date=2026-07-23}}}}</ref>

The Mari-B field itself depleted around 2013–2015 and its wells were plugged and abandoned in 2021; the Mari-B platform now serves as a logistics and crew-lodging facility for Chevron's Tamar operations, while the export pipeline remains in active service.<ref name=":3">{{{{Cite web|url={YN}|title=Mari-B platform repurposed for Tamar operations|website=Ynetnews|date=2026|access-date=2026-07-23}}}}</ref>

==Articles and resources==
===References===
{{{{reflist}}}}
===Additional data===
To access additional data, including an interactive map of gas pipelines, a downloadable dataset, and summary data, please visit the [https://globalenergymonitor.org/projects/global-gas-infrastructure-tracker/ Global Gas Infrastructure Tracker] on the Global Energy Monitor website.

[[Category:Global Gas Infrastructure Tracker]]
[[Category:Gas and Hydrogen Pipelines]]
"""


def wiki_p8003(v):
    gj = json.loads(ROUTE.read_text())
    coords = []
    for f in gj["features"]:
        g = f["geometry"]
        if g["type"] == "LineString":
            coords = g["coordinates"]
    # The candidate carries the FULL drawn line Tanin ⊕ -> Karish ⊕ (a FUTURE field-
    # tieback leg) -> Dor ○ (the operating export). This discovery row is the OPERATING
    # export line, so the display map shows only the Karish-FPSO -> Dor segment; the
    # future Tanin leg is staged in the route candidate for a human trim decision.
    KARISH = (34.243301, 33.236991)
    ki = min(range(len(coords)),
             key=lambda i: (coords[i][0] - KARISH[0]) ** 2 + (coords[i][1] - KARISH[1]) ** 2)
    export = coords[ki:]
    map_lines = ";\n".join(f"{lat},{lon}" for lon, lat in export)
    OT = "https://www.offshore-technology.com/projects/karish-tanin-field-development-mediterranean-sea/"
    OM = "https://www.offshore-mag.com/regional-reports/middle-east/article/14234861/karish-offshore-gas-field-connected-to-israeli-distribution-system"
    NSE = "https://www.nsenergybusiness.com/projects/karish-gas-field-development/"
    WP = "https://en.wikipedia.org/wiki/Karish_gas_field"
    TOI = "https://www.timesofisrael.com/israel-gives-final-approval-for-start-of-production-at-karish-gas-field/"
    OED = "https://www.oedigital.com/news/485590-energean-takes-full-ownership-of-energean-israel"
    WO = "https://www.worldoil.com/news/2019/6/25/energean-ingl-agree-to-the-transfer-of-near-shore-and-onshore-infrastructure"
    return f"""{{{{Navbar-Global Gas Infrastructure Tracker}}}}
'''Karish–Tanin FPSO Gas Export Pipeline''' (also known as the Karish gas sales pipeline or Energean Power FPSO export pipeline) is an operating subsea natural gas pipeline off the coast of Israel, carrying gas from the Energean Power FPSO to a landfall at Dor.<ref name=":0" /><ref name=":1" />

==Location==
The pipeline runs from the Energean Power FPSO, moored over the Karish and Tanin fields about 90 km offshore northern Israel, to a landfall at Dor, where it connects to Israel's national gas transmission grid.<ref name=":0" /><ref name=":2" />

{{{{#display_map:
| lines=
{map_lines}
| center=
32.9,34.5
| width=400
| height=300
| type=normal
| zoom=8}}}}

==Project details==
*'''Operator: '''Energean<ref name=":0" />
*'''Owner: '''Energean Israel Limited (100%)<ref name=":3" />
*'''Parent company: '''Energean plc<ref name=":3" />
*'''Capacity: '''8 billion cubic meters per year (FPSO nameplate)<ref name=":2" /><ref name=":4" />
*'''Length: '''90.3 km<ref name=":5" /><ref name=":2" />
*'''Diameter: '''30 inches and 24 inches (dual-diameter)<ref name=":5" /><ref name=":2" /><ref name=":1" />
*'''Status: '''Operating<ref name=":6" /><ref name=":7" />
*'''Start year: '''2022<ref name=":6" />

==Background==
Gas from the Karish field (and later Karish North and Tanin) is produced and processed on the Energean Power FPSO. Treated sales gas is exported via a roughly 90 km subsea pipeline, installed by Allseas and completed in June 2020, to a landfall at Dor, where it enters Israel's national transmission grid.<ref name=":0">{{{{Cite web|url={OT}|title=Karish and Tanin Field Development, Mediterranean Sea|website=Offshore Technology|access-date=2026-07-23}}}}</ref><ref name=":5">{{{{Cite web|url={WP}|title=Karish gas field|website=Wikipedia|access-date=2026-07-23}}}}</ref><ref name=":2">{{{{Cite web|url={NSE}|title=Karish Gas Field Development, Mediterranean Sea, Israel|website=NS Energy Business|access-date=2026-07-23}}}}</ref>

First gas was achieved on 26 October 2022, with flow initially constrained to about 6.5 bcm/y until the Karish North tie-back allowed full use of the FPSO's 8 bcm/y design capacity.<ref name=":6">{{{{Cite web|url={TOI}|title=Israel gives final approval for start of production at Karish gas field|website=The Times of Israel|date=2022-10|access-date=2026-07-23}}}}</ref><ref name=":1">{{{{Cite web|url={OM}|title=Karish offshore gas field connected to Israeli distribution system|website=Offshore Magazine|access-date=2026-07-23}}}}</ref>

Energean plc holds 100% of the development through its subsidiary Energean Israel Limited, having acquired Kerogen Capital's former 30% interest in late 2020.<ref name=":3">{{{{Cite web|url={OED}|title=Energean Takes Full Ownership of Energean Israel|website=Offshore Engineer|date=2021|access-date=2026-07-23}}}}</ref> Under a 2019 agreement, Energean transferred title of the nearshore (about 10 km) and onshore facilities, including the Dor receiving station, to Israel Natural Gas Lines (INGL), which owns and operates that segment.<ref name=":7">{{{{Cite web|url={WO}|title=Energean, INGL agree to the transfer of near shore and onshore infrastructure|website=World Oil|date=2019-06-25|access-date=2026-07-23}}}}</ref><ref name=":4">{{{{Cite web|url=https://www.washingtoninstitute.org/policy-analysis/israels-karish-offshore-gas-field-facts-and-figures|title=Israel's Karish Offshore Gas Field: Facts and Figures|website=The Washington Institute|date=2022|access-date=2026-07-23}}}}</ref>

==Articles and resources==
===References===
{{{{reflist}}}}
===Related GEM.wiki articles===
*[[Leviathan Subsea Gas Pipelines]]
*[[Israel Cyprus Gas Pipeline]]
===Additional data===
To access additional data, including an interactive map of gas pipelines, a downloadable dataset, and summary data, please visit the [https://globalenergymonitor.org/projects/global-gas-infrastructure-tracker/ Global Gas Infrastructure Tracker] on the Global Energy Monitor website.

[[Category:Global Gas Infrastructure Tracker]]
[[Category:Gas and Hydrogen Pipelines]]
"""


if __name__ == "__main__":
    main()
