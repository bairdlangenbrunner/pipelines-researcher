#!/usr/bin/env python3
"""Sanity-check a built workbook before delivery: it opens, and no cell carries a
spreadsheet error string. Reconciliation/research workbooks have no formulas, so this
mainly catches corruption or stray error text.

    python scripts/recalc.py batches/pipelines_batch_<...>.xlsx
"""
from __future__ import annotations

import argparse
import sys

from openpyxl import load_workbook

ERRORS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "#SPILL!")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx")
    args = ap.parse_args()
    try:
        wb = load_workbook(args.xlsx)
    except Exception as e:
        sys.exit(f"FAIL: cannot open {args.xlsx}: {e}")
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in ERRORS:
                    bad.append((ws.title, cell.coordinate, cell.value))
    if bad:
        print(f"FAIL: {len(bad)} error cell(s):")
        for t, coord, v in bad[:20]:
            print(f"  {t}!{coord} = {v}")
        sys.exit(1)
    print(f"ok: {args.xlsx} — {len(wb.sheetnames)} sheets, no error cells "
          f"({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
