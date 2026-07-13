#!/usr/bin/env python3
"""Discovery deliverable: build the reviewable xlsx from staged_new.json. Nothing is
auto-applied — Baird pastes qualifying new rows into the live Sheet manually.

    python scripts/build_discovery_workbook.py --staging batches/staging/annual-gas-iraq/ \
        --output batches/pipelines_batch_<stamp>_<scope>_discovery.xlsx
    # <stamp> from: TZ=America/New_York date "+%Y%m%d_%H%M_ET"   (never overwrite)

Sheets (commodity-prefixed; empty omitted; README first):
  <Cmdty>_NewRows          PRIMARY paste-ready view — the EXACT tracker header (from the
                           snapshot named in discovery_context.json); one green-tinted row
                           per qualifying candidate, values + verified [ref]s in place.
  <Cmdty>_MonitorList      below the add-threshold (Discovery SOP §3) — watch, don't add.
  <Cmdty>_MatchedExisting  candidates that matched an existing GEM row under another name —
                           OtherEnglishNames suggestions, NOT new rows.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_recon_workbook import (  # noqa: E402
    CONF_FILL, HEADER_FILL, J, _style_header, _write_sheet,
)

NEW_FILL = PatternFill("solid", fgColor="E2EFDA")   # green tint = candidate new row

# The tracker consolidates many attribute refs into a handful of [ref] columns, but
# discovery agents often stage granular per-attribute ref keys. Remap them onto the real
# tracker column so a researched value never lands without its [ref] (and so nothing is
# silently dropped from the mirror).
REF_ALIAS = {
    "LengthKnown [ref]": "Length [ref]",
    "LengthKnownKm [ref]": "Length [ref]",
    "LengthEstimateKm [ref]": "Length [ref]",
    "StartLocation [ref]": "Location [ref]",
    "EndLocation [ref]": "Location [ref]",
    "StartState/Province [ref]": "Location [ref]",
    "EndState/Province [ref]": "Location [ref]",
    "StartCountryOrArea [ref]": "Location [ref]",
    "EndCountryOrArea [ref]": "Location [ref]",
    "CountriesOrAreas [ref]": "Location [ref]",
    "ProposalYear [ref]": "Proposal [ref]",
    "ProposalMonth [ref]": "Proposal [ref]",
    "ConstructionYear [ref]": "Construction [ref]",
    "ConstructionMonth [ref]": "Construction [ref]",
    "StartYear1 [ref]": "Start [ref]",
    "StartMonth1 [ref]": "Start [ref]",
    "DiameterInMm [ref]": "Diameter [ref]",
}
# Owner/Parent/Operator refs have NO column in the main tracker header — by convention
# they live on the ProjectID-keyed operators/owners tab. Route them to a dedicated sheet
# rather than dropping them (and never let a ref key overwrite an Owner/Parent value cell).
OWNER_REF_KEYS = {"Owner [ref]", "Parent [ref]", "Operator [ref]", "ParentEntityIDs [ref]"}


def _nonempty(v) -> bool:
    return v is not None and str(v).strip() != ""


def _dedup(seq):
    out = []
    for x in seq:
        if x not in out:
            out.append(x)
    return out


def _tracker_header(scope: dict) -> list[str]:
    csv_name = scope.get("csv")
    if not csv_name:
        sys.exit("staged_new.json meta.scope.csv missing — cannot mirror the tracker header")
    cand = Path(csv_name)
    if not cand.exists():
        cand = Path(__file__).resolve().parent.parent / "data" / Path(csv_name).name
    if not cand.exists():
        sys.exit(f"snapshot {csv_name} not found in data/ — the NewRows tab mirrors its header")
    with cand.open(newline="") as f:
        return list(csv.reader(f))[2]   # tracker header at CSV row index 2


def _new_rows_view(wb, title, header, rows):
    """PRIMARY paste-ready mirror. Only *non-empty* staged values/refs are written and
    green-tinted — an empty cell is never coloured. A ref may only ever land in a `[ref]`
    column (a value column is never overwritten by a ref), and granular ref keys are
    consolidated onto the tracker's real ref columns via REF_ALIAS. Owner/Parent refs
    (which have no tracker column) are returned for a separate operators/owners tab."""
    ws = wb.create_sheet(title)
    ws.append(header)
    col_of = {h: i + 1 for i, h in enumerate(header) if h}
    unmapped = set()
    oo_rows = []   # (candidate, {owner_ref_key: [urls]}) → OperatorsOwners tab
    for c in rows:
        cells = {}                       # value-column index -> value
        for k, v in (c.get("values") or {}).items():
            if not _nonempty(v):
                continue                 # looked, found nothing → leave white, do not colour
            (cells.__setitem__(col_of[k], v) if k in col_of else unmapped.add(k))
        ref_urls = {}                    # ref-column index -> [urls]
        oo = {}
        for rc, urls in (c.get("refs") or {}).items():
            urls = [u for u in (urls if isinstance(urls, list) else [urls]) if _nonempty(u)]
            if not urls:
                continue
            target = REF_ALIAS.get(rc, rc)
            # a ref key MUST name a [ref] column; a bare value-column name (e.g. a
            # mis-keyed "Owner") is coerced so it can never be written into that value cell.
            if not target.endswith(" [ref]"):
                target += " [ref]"
            if target in OWNER_REF_KEYS:
                oo.setdefault(target, []).extend(urls)
            elif target in col_of:
                ref_urls.setdefault(col_of[target], []).extend(urls)
            else:
                unmapped.add(rc)
        for ci, us in ref_urls.items():
            cells[ci] = J(_dedup(us))
        notes = c.get("researcher_notes", "")
        if _nonempty(notes) and "ResearcherNotes" in col_of and col_of["ResearcherNotes"] not in cells:
            cells[col_of["ResearcherNotes"]] = notes
        ws.append([cells.get(i, "") for i in range(1, len(header) + 1)])
        rn = ws.max_row
        for i, v in cells.items():
            if not _nonempty(v):
                continue
            cell = ws.cell(rn, i)
            cell.fill = NEW_FILL
            cell.alignment = Alignment(wrap_text=False, vertical="top")
        if oo:
            oo_rows.append((c, oo))
    _style_header(ws, len(header))
    for i, h in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(i)].width = 40 if h.endswith(" [ref]") else 16
    ws.freeze_panes = "A2"
    if unmapped:
        print(f"  WARN: {title}: staged keys not in the tracker header (dropped from the "
              f"mirror; still in staged_new.json): {sorted(unmapped)}")
    return ws, oo_rows


def _owner_refs_view(wb, title, oo_rows):
    """Owner/operator [ref]s for the new rows. The main tracker has no Owner/Parent [ref]
    column — these apply on the ProjectID-keyed operators/owners tab once the row has an
    ID. Per that tab's convention the [ref] precedes its value. Only non-empty cells are
    written/tinted."""
    ws = wb.create_sheet(title)
    header = ["Candidate name", "Owner [ref]", "Owner", "Parent [ref]", "Parent",
              "Operator [ref]", "Operator"]
    ws.append(header)
    for c, oo in oo_rows:
        vals = c.get("values", {})
        row = [
            c.get("name") or vals.get("PipelineName", ""),
            J(_dedup(oo.get("Owner [ref]", []))), vals.get("Owner", ""),
            J(_dedup(oo.get("Parent [ref]", []))), vals.get("Parent", ""),
            J(_dedup(oo.get("Operator [ref]", []))), vals.get("Operator", ""),
        ]
        ws.append(row)
        rn = ws.max_row
        for ci in range(1, len(header) + 1):
            if _nonempty(ws.cell(rn, ci).value):
                ws.cell(rn, ci).fill = NEW_FILL
                ws.cell(rn, ci).alignment = Alignment(wrap_text=False, vertical="top")
    _style_header(ws, len(header))
    for i, w in enumerate((34, 52, 28, 52, 28, 52, 28), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def _compact_columns(extra):
    g = lambda k: (lambda r: r.get(k, ""))
    cols = [
        ("Candidate name", g("name"), 34),
        *extra,
        ("Key values", lambda r: J([f"{k}={v}" for k, v in r.get("values", {}).items()]), 44),
        ("Verified ref(s)", lambda r: J(sorted({u for us in r.get("refs", {}).values() for u in us})), 52),
        ("Corroboration tier", g("tier"), 13),
        ("ResearcherNotes", g("researcher_notes"), 64),
    ]
    return cols


def _fill_readme(ws, meta, sheet_defs):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    scope = meta.get("scope", {})
    rows = [
        ("GEM pipeline discovery", ""),
        ("Tracker / commodity", scope.get("tracker", "")),
        ("Country", scope.get("country", "")),
        ("GEM CSV", scope.get("csv", "")),
        ("Counts", J([f"{k}={v}" for k, v in meta.get("class_counts", {}).items()])),
        ("", ""),
        ("Work from", "the _NewRows tab (exact tracker header; green cells = staged values/refs). "
                      "MonitorList = below the add-threshold, watch only. MatchedExisting = "
                      "OtherEnglishNames suggestions for existing rows, NOT new pipelines."),
        ("Add-threshold", "new_row requires ALL of: identified sponsor; country + region/endpoints; "
                          "a concrete step (MOU / FEED / permit / tender / FID). Below -> monitor."),
        ("Standing rules", "Never cite gem.wiki/globalenergymonitor/theodora/wikidot (rule 1). No "
                           "fabricated URLs (rule 2). Every ref passed url_verifier. No orphan "
                           "values/refs. Nothing auto-applied — paste manually."),
        ("", ""),
        ("Sheets", ""),
    ]
    for r in rows:
        ws.append(r)
    for name, desc in sheet_defs:
        ws.append((name, desc))
    ws["A1"].font = Font(bold=True, size=13)
    for c in (1, 2):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF", size=13)
    for rn in range(2, ws.max_row + 1):
        ws.cell(rn, 1).font = Font(bold=True)
        ws.cell(rn, 2).alignment = Alignment(wrap_text=False, vertical="top")
    ws.freeze_panes = "A2"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--staging", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    out = Path(args.output)
    if out.exists():
        sys.exit(f"refusing to overwrite existing {out} (use a fresh <stamp>)")

    data = json.loads((Path(args.staging) / "staged_new.json").read_text())
    meta = data.get("meta", {})
    candidates = data.get("candidates", [])
    prefix = (meta.get("scope", {}).get("tracker") or "gas").capitalize()

    new_rows = [c for c in candidates if c.get("class") == "new_row"]
    monitor = [c for c in candidates if c.get("class") == "monitor"]
    matched = [c for c in candidates if c.get("class") == "matched_existing"]

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    sheet_defs = []

    if new_rows:
        title = f"{prefix}_NewRows"
        _, oo_rows = _new_rows_view(wb, title, _tracker_header(meta.get("scope", {})), new_rows)
        sheet_defs.append((title,
                           f"{len(new_rows)} — PRIMARY paste-ready: exact tracker header, one green row per "
                           "candidate that clears the add-threshold, values + verified [ref]s in place."))
        if oo_rows:
            oo_title = f"{prefix}_OperatorsOwners"
            _owner_refs_view(wb, oo_title, oo_rows)
            sheet_defs.append((oo_title,
                               f"{len(oo_rows)} — owner/operator [ref]s for the new rows. The main tracker has "
                               "no Owner/Parent [ref] column; apply these on the ProjectID-keyed operators/owners "
                               "tab once each row has an ID. [ref] precedes its value."))
    def _flag_col2(color):
        def styler(ws, rn, r):
            ws.cell(rn, 2).fill = CONF_FILL[color]
        return styler

    if monitor:
        cols = _compact_columns([("Why monitor (threshold leg failed)",
                                  lambda r: r.get("monitor_reason", ""), 40)])
        title = f"{prefix}_MonitorList"
        _write_sheet(wb, title, cols, monitor, _flag_col2("yellow"))
        sheet_defs.append((title,
                           f"{len(monitor)} — below the add-threshold: watch for the concrete step, "
                           "do NOT add yet."))
    if matched:
        cols = _compact_columns([("Matched ProjectID",
                                  lambda r: r.get("matched_project_id", ""), 16)])
        title = f"{prefix}_MatchedExisting"
        _write_sheet(wb, title, cols, matched, _flag_col2("green"))
        sheet_defs.append((title,
                           f"{len(matched)} — same physical pipe as an existing GEM row under another "
                           "name: add the candidate name to that row's OtherEnglishNames, no new row."))

    _fill_readme(readme, meta, sheet_defs)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")
    print("  next: python scripts/recalc.py " + str(out))


if __name__ == "__main__":
    main()
