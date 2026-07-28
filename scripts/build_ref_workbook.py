#!/usr/bin/env python3
"""Reference-sweep step 4: build the reviewable deliverable from staged_resolutions.json
(the agent's output). Nothing here is auto-applied — Baird pastes verified refs into the
live Sheet manually. Every Proposed ref(s) cell has passed url_verifier (HTTP 200 + value
present, no GEM/theodora).

    python scripts/build_ref_workbook.py --staging batches/saudi-arabia-oil/staging/ref-sweep/ \
        --output batches/pipelines_batch_<stamp>_saudi-arabia_refsweep.xlsx
    # <stamp> from: TZ=America/New_York date "+%Y%m%d_%H%M_ET"   (never overwrite)

Sheets (commodity-prefixed; empty omitted; README first):
  <Cmdty>_StatusReview     ANNUAL UPDATE only — one verdict per in-dev segment row: confirm /
                           change (evidence-based, with proposed Status + date cols) / stale
                           (dormancy rule -> Presumed) / unclear. Leads the packet when present.
  <Cmdty>_Backend          PRIMARY paste-ready view — a 1:1 mirror of the GEM tracker backend:
                           the FULL backend column set in exact sheet order, current values
                           prefilled, with proposed ref(s)/values overlaid on touched cells
                           (colored by corroboration tier). Leading SheetRow = row locator.
  <Cmdty>_OperatorsOwners  paste-ready mirror of the separate "Pipeline operators/owners"
                           backend tab (ProjectID-keyed; [ref] PRECEDES its values) — the
                           Operator [ref] / Owner [ref] cells, colored by tier.
  <Cmdty>_Validity         DEEP SWEEP only — existence/duplicate/classification/attribution/
                           spec concerns per pipeline (read-and-flag, never auto-applied).
  <Cmdty>_Fills            DEEP SWEEP only — blank value fields researched + filled with a
                           paired ref (or left blank when not corroborated).
  <Cmdty>_RouteSuggestions DEEP SWEEP only — sourced endpoint coords + corridor for weak-
                           RouteAccuracy rows (feeds the routes repo via a separate human PR).
  <Cmdty>_GulfPub          DEEP SWEEP only — GulfPub (PE World Map, Tier 2) cross-comparison:
                           overlaps / additions / ambiguous, present when gulfpub_crosswalk.json
                           was generated (build_gulfpub_crosswalk.py) for the staging dir.
  <Cmdty>_WikiAlignment    QC PACKET only (meta.mode "qc") — per-field sheet↔wiki diffs
                           (WIKI_UPDATE / SHEET_SUSPECT / WIKI_STALE_VS_STAGED / UNPARSED).
  <Cmdty>_RouteIntegrity   QC PACKET only — drawn GeoJSON vs the row's own attributes
                           (NOT the permanently-dropped route/WKT-format check).
  <Cmdty>_Flags            QC PACKET only — country-scoped mechanical checks, rendered from
                           the qc_flags.json sidecar. In QC mode the Validity tab is retitled
                           <Cmdty>_Findings, and instead of a Backend mirror the packet leads
                           with <Cmdty>_AllFillsBackend: ALL corroborated fills (carried +
                           this packet's own) unified in the exact backend layout, values +
                           [ref] overlaid tier-colored (no SheetRow locator — columns align
                           1:1 with the sheet for copy-paste).
  <Cmdty>_Refs_Added       MISSING_REF resolved — green ≥2 independent / yellow single
  <Cmdty>_Refs_Reverified  HAS_REF, links live + contain value (blue)
  <Cmdty>_Refs_DeadLinks   HAS_REF with a dead/value-missing link + proposed replacement
  <Cmdty>_Refs_Unresolved  couldn't reach 2 working corroborating links → manual review

The four *_Refs_* bucket tabs are supporting detail; the <Cmdty>_Backend tab is the one
Baird works from. Route/geometry `[ref]` cells are out of scope (reconciled separately).

HANDOFF mode (staged_actions.json present in the staging dir) writes TWO files instead —
derived from --output: <stem>-actions.xlsx (ONLY suggested changes + open issues, tab
order = work order: Decisions → StatusChanges → AllFillsBackend [fills AND paste-ready
refs unified on the full backend mirror] → OperatorsOwners → NewRows/NewRowRefs/
MatchedExisting → WikiUpdates → RouteSuggestions → OpenFlags) and <stem>-evidence.xlsx
(the audit trail: ConfirmedAudit, FillDetail, RefWorkDetail, non-action WikiAlignment,
covered RouteIntegrity/Flags, MonitorList, GulfPub). Legacy qc dirs without the sidecar
keep the single-workbook layout above.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_recon_workbook import (  # noqa: E402
    CONF_FILL, HEADER_FILL, J, _style_header, _write_sheet,
)
from ref_pairs import OO_PRIMARY  # noqa: E402  (oo-tab ref cols: Owner/Operator [ref])
from build_discovery_workbook import (  # noqa: E402  (discovery tabs in handoff mode)
    _compact_columns, _new_rows_view, _owner_refs_view,
)

BLUE_FILL = PatternFill("solid", fgColor="DDEBF7")   # re-verified (blue), per confidence_tiers
VALIDITY_REF = "__VALIDITY__"   # synthetic ref_col sentinel on deep-sweep validity records
STATUS_REF = "__STATUS__"       # synthetic ref_col sentinel on annual-update status reviews
ROUTE_REF = "__ROUTE__"         # synthetic ref_col sentinel on deep-sweep route suggestions
WIKIDIFF_REF = "__WIKIDIFF__"   # QC workflow: sheet↔wiki alignment diffs (wiki_alignment.py)
ROUTEQC_REF = "__ROUTEQC__"     # QC workflow: route-integrity flags (route_integrity.py)

# class_out -> (sheet suffix, readme blurb)
_BUCKETS = {
    "REFS_ADDED": ("Refs_Added", "blank [ref] filled. Tier cell green = ≥2 independent working "
                   "sources; yellow = single source (still needs a 2nd)."),
    "REVERIFIED": ("Refs_Reverified", "existing [ref] re-checked: all links live AND still contain "
                   "the value, ≥2 independent. Blue = verified, no action needed."),
    "DEAD_LINK": ("Refs_DeadLinks", "existing [ref] has a dead / value-missing link (red). Proposed "
                  "ref(s) = a verified replacement to swap in."),
    "UNRESOLVED": ("Refs_Unresolved", "could not reach 2 working, independent, value-containing links "
                   "(red). Manual review — no fabricated URLs (standing rule 2)."),
}
_ORDER = ["REFS_ADDED", "REVERIFIED", "DEAD_LINK", "UNRESOLVED"]


def _verif_summary(r: dict) -> str:
    vs = r.get("verifications", [])
    if not vs:
        return ""
    live = sum(1 for v in vs if v.get("ok"))
    val = sum(1 for v in vs if v.get("contains_value"))
    return f"{live}/{len(vs)} live, {val} contain value"


def _ref_columns(with_source=False):
    g = lambda k: (lambda r: r.get(k, ""))
    cols = [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 32),
        ("SegmentName", g("segment_name"), 22),
        ("Ref column", lambda r: r.get("ref_col") or "(Owner → Operators/Owners tab)", 24),
        ("Data point(s)", lambda r: r.get("primary_value_col") or J(list(r.get("values", {}).keys())), 18),
        ("Current value", lambda r: r.get("primary_value") or J([f"{k}={v}" for k, v in r.get("values", {}).items()]), 22),
        ("Current ref", g("current_ref"), 30),
        ("Proposed ref(s)", lambda r: J(r.get("proposed_refs", [])), 52),
        ("Verification status", _verif_summary, 22),
        ("Corroboration tier", g("tier"), 13),
        ("Independent?", lambda r: "yes" if r.get("independent") else ("no" if r.get("proposed_refs") else ""), 11),
        ("Source language", g("source_language"), 12),
        ("ResearcherNotes", g("researcher_notes"), 50),
        ("Wiki (visited, not cited)", g("wiki"), 34),
    ]
    if with_source:
        cols.append(("Source packet", g("source_dir"), 28))
    return cols


def _tier_color(r: dict) -> str:
    if r.get("tier_color"):
        return r["tier_color"]
    return {"high": "green", "medium": "yellow"}.get(r.get("tier", ""), "red")


def _make_styler(columns, bucket: str):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}
    tier_c = idx["Corroboration tier"]
    verif_c = idx["Verification status"]
    cur_ref_c = idx["Current ref"]

    def styler(ws, rn, r):
        if bucket == "REVERIFIED":
            ws.cell(rn, tier_c).fill = BLUE_FILL
            ws.cell(rn, verif_c).fill = BLUE_FILL
            return
        ws.cell(rn, tier_c).fill = CONF_FILL.get(_tier_color(r), PatternFill())
        if bucket == "DEAD_LINK":
            ws.cell(rn, cur_ref_c).fill = CONF_FILL["red"]
        elif bucket == "UNRESOLVED":
            ws.cell(rn, tier_c).fill = CONF_FILL["red"]
    return styler


# --- validity + fills (deep-sweep extensions) ------------------------------ #
# Ordered most-important-first so the strongest signal wins; the Finding column is
# authoritative — Concern type is only a sort/filter hint when the record carries no
# explicit `concern_type` (older staging predates the structured field).
_CONCERN_RULES = [
    ("existence", ("does not exist", "doesn't exist", "no evidence", "cannot find",
                   "can't find", "hallucinat", "phantom", "fabricat", "not a real",
                   "non-existent", "nonexistent", "no independent evidence")),
    ("duplicate", ("duplicate", "relabel", "same pipe", "same physical", "same as ",
                   "already tracked", "redundant", "double-count", "double count")),
    ("classification", ("not a transmission", "not dry gas", " ngl", "(ngl", "reclassif",
                        "gathering line", "process line", "feeder line", "misclassif",
                        "should be classified", "not a pipeline", "may be ngl",
                        "commodity may be")),
    ("attribution", ("wrong owner", "wrong operator", "wrong province", "wrong endpoint",
                     "misattribut", "endpoint", "fuelsource", "start state", "province error",
                     "naming", "operator")),
    ("spec", ("length", "diameter", "capacity", "discrepancy", " km vs", "does not reconcile",
              "conflict", "mismatch", "uncorroborated")),
]


# Phrasing that signals the agent CONFIRMED the pipeline is real (often while raising a
# lesser caveat). Used to keep negated mentions ("NOT a duplicate/non-existent flag",
# "exists and is correctly attributed") from being miscoded as existence/duplicate reds.
_CONFIRM_SIGNALS = ("genuine", "genuinely exist", "exists and is correctly", "correctly attributed",
                    "correctly identified", "not a duplicate", "non-existent flag", "existence ok",
                    "existence is", "does exist", "is real", "well-sourced", "not an error",
                    "segment exists", "pipeline exists", "plausibly exists")


def _validity_verdict(r: dict) -> str:
    v = (r.get("verdict") or "").strip()
    if v:
        return v
    note = (r.get("researcher_notes") or "").lower()
    return "confirmed (caveat)" if any(s in note for s in _CONFIRM_SIGNALS) else "concern"


def _concern_type(r: dict) -> str:
    explicit = (r.get("concern_type") or "").strip()
    if explicit:
        return explicit
    note = (r.get("researcher_notes") or "").lower()
    rules = _CONCERN_RULES
    # If existence/identity is affirmatively confirmed, drop the existence/duplicate buckets
    # (their keywords are usually negated here) — a confirmed pipeline can still be a spec,
    # attribution, or classification (e.g. NGL-vs-gas) concern, so keep those.
    if any(s in note for s in _CONFIRM_SIGNALS):
        rules = [(l, k) for (l, k) in _CONCERN_RULES if l not in ("existence", "duplicate")]
    for label, kws in rules:
        if any(k in note for k in kws):
            return label
    return "review"


_HIGH_CONCERN = {"existence", "duplicate", "classification"}


def _validity_columns():
    g = lambda k: (lambda r: r.get(k, ""))
    return [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 22),
        ("Verdict", _validity_verdict, 16),
        ("Concern type", _concern_type, 16),
        ("Finding", g("researcher_notes"), 72),
        ("Recommendation", g("recommendation"), 40),
        ("Sources", lambda r: J(r.get("proposed_refs", [])), 52),
        ("Corroboration tier", g("tier"), 13),
        ("Independent?", lambda r: "yes" if r.get("independent") else ("no" if r.get("proposed_refs") else ""), 11),
        ("Source language", g("source_language"), 12),
        ("Wiki (visited, not cited)", g("wiki"), 34),
    ]


def _validity_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}
    tier_c = idx["Corroboration tier"]
    concern_c = idx["Concern type"]
    verdict_c = idx["Verdict"]

    def styler(ws, rn, r):
        ws.cell(rn, tier_c).fill = CONF_FILL.get(_tier_color(r), PatternFill())
        is_concern = _validity_verdict(r) == "concern"
        ws.cell(rn, verdict_c).fill = CONF_FILL["red"] if is_concern else CONF_FILL["green"]
        # red the concern type only for an OPEN high-severity concern (phantom / duplicate /
        # misclassified) — not when the pipeline was confirmed real with a lesser caveat.
        if is_concern and ws.cell(rn, concern_c).value in _HIGH_CONCERN:
            ws.cell(rn, concern_c).fill = CONF_FILL["red"]
    return styler


def _fills_columns(with_source=False):
    g = lambda k: (lambda r: r.get(k, ""))
    cols = [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 22),
        ("Field", lambda r: r.get("primary_value_col") or J(r.get("value_cols", [])), 18),
        ("Target tab", lambda r: "operators/owners" if r.get("tab") == "operators_owners" else "tracker", 15),
        ("Proposed value", lambda r: r.get("primary_value") or J([f"{k}={v}" for k, v in r.get("values", {}).items()]), 24),
        ("Proposed ref(s)", lambda r: J(r.get("proposed_refs", [])), 52),
        ("Verification status", _verif_summary, 22),
        ("Outcome", lambda r: {"REFS_ADDED": "filled (corroborated)",
                               "UNRESOLVED": "not corroborated / dropped"}.get(r.get("class_out", ""), r.get("class_out", "")), 24),
        ("Corroboration tier", g("tier"), 13),
        ("Independent?", lambda r: "yes" if r.get("independent") else ("no" if r.get("proposed_refs") else ""), 11),
        ("Source language", g("source_language"), 12),
        ("ResearcherNotes", g("researcher_notes"), 50),
        ("Wiki (visited, not cited)", g("wiki"), 34),
    ]
    if with_source:
        cols.append(("Source packet", g("source_dir"), 28))
    return cols


def _fills_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}
    tier_c = idx["Corroboration tier"]

    def styler(ws, rn, r):
        if r.get("class_out") == "UNRESOLVED":
            ws.cell(rn, tier_c).fill = CONF_FILL["red"]
        else:
            ws.cell(rn, tier_c).fill = CONF_FILL.get(_tier_color(r), PatternFill())
    return styler


# --- status review (annual-update extension) ------------------------------- #
_STATUS_VERDICT_FILL = {"confirm": "green", "change": "yellow", "stale": "red", "unclear": "red"}


def _status_columns(with_source=False):
    g = lambda k: (lambda r: r.get(k, ""))
    cols = [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 22),
        ("Current status", g("current_status"), 13),
        ("Verdict", g("verdict"), 10),
        ("Proposed status", g("proposed_status"), 13),
        ("Proposed changes", lambda r: J([f"{k}={v}" for k, v in r.get("values", {}).items()]), 36),
        ("Evidence date", g("evidence_date"), 12),
        ("Staleness rule", g("staleness_rule"), 13),
        ("Proposed ref(s)", lambda r: J(r.get("proposed_refs", [])), 52),
        ("Verification status", _verif_summary, 22),
        ("Corroboration tier", g("tier"), 13),
        ("Independent?", lambda r: "yes" if r.get("independent") else ("no" if r.get("proposed_refs") else ""), 11),
        ("Source language", g("source_language"), 12),
        ("ResearcherNotes", g("researcher_notes"), 60),
        ("Wiki (visited, not cited)", g("wiki"), 34),
    ]
    if with_source:
        cols.append(("Source packet", g("source_dir"), 28))
    return cols


def _status_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}
    verdict_c = idx["Verdict"]
    prop_c = idx["Proposed status"]
    tier_c = idx["Corroboration tier"]

    def styler(ws, rn, r):
        v = (r.get("verdict") or "").lower()
        ws.cell(rn, verdict_c).fill = CONF_FILL.get(_STATUS_VERDICT_FILL.get(v, "red"), PatternFill())
        if r.get("proposed_status"):
            ws.cell(rn, prop_c).fill = CONF_FILL["yellow"]
        if r.get("proposed_refs"):
            ws.cell(rn, tier_c).fill = CONF_FILL.get(_tier_color(r), PatternFill())
    return styler


def _ref_cell_text(r: dict) -> str:
    """What to paste into the `[ref]` cell: the proposed ref(s); for a re-verified unit
    with no proposed change, the existing (re-checked) ref."""
    refs = r.get("proposed_refs") or []
    if refs:
        return J(refs)
    if r.get("class_out") == "REVERIFIED":
        return r.get("current_ref", "")
    return ""


def _ref_cell_fill(r: dict):
    """Corroboration color for a `[ref]` cell: blue=re-verified, else tier green/yellow/red."""
    if r.get("class_out") == "REVERIFIED":
        return BLUE_FILL
    return CONF_FILL.get(_tier_color(r), PatternFill())


def _int_or_zero(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def _backend_snapshot(meta: dict):
    """Load the GEM tracker backend's FULL header + every data row keyed by
    (ProjectID, SheetRow), so the _Backend tab can reproduce the complete backend column
    layout in exact sheet order with current values prefilled. The tracker header is at CSV
    row index 2; data rows follow, and SheetRow = data-row index + 4 (verified against the
    snapshot). Returns (header, rows) — or ([], {}) if the snapshot named in meta.scope.csv
    can't be located, in which case the caller falls back to a structure-only mirror."""
    csv_name = (meta.get("scope") or {}).get("csv")
    if not csv_name:
        return [], {}
    # snapshots live in the repo data/ dir; tolerate either an absolute path or a bare name
    cand = Path(csv_name)
    if not cand.exists():
        cand = Path(__file__).resolve().parent.parent / "data" / Path(csv_name).name
    if not cand.exists():
        return [], {}
    try:
        with cand.open(newline="") as f:
            allrows = list(csv.reader(f))
    except OSError:
        return [], {}
    if len(allrows) < 3:
        return [], {}
    header = allrows[2]
    try:
        pid_i = header.index("ProjectID")
    except ValueError:
        return header, {}
    rows: dict[tuple, dict] = {}
    for di, raw in enumerate(allrows[3:]):        # data starts after the header at CSV index 2
        pid = raw[pid_i] if pid_i < len(raw) else ""
        sheet_row = di + 4                         # SheetRow = data-row index + 4
        rows[(pid, sheet_row)] = {col: (raw[ci] if ci < len(raw) else "")
                                  for ci, col in enumerate(header)}
    return header, rows


def _backend_view(wb, title, resolutions, backend_header, snapshot_rows, color_values=False,
                  sheet_row_col=True):
    """The PRIMARY tab: a 1:1 paste-ready mirror of the GEM tracker backend. Reproduces the
    tracker's FULL column set in exact sheet order (every backend column, including computed
    ones), one row per in-scope segment, with current values prefilled from the snapshot.
    On each touched cluster the proposed ref(s) are overlaid on the `[ref]` cell — color-
    coded by corroboration tier (same green ≥2-independent / yellow single / red low-or-none
    / blue re-verified key as the bucket tabs) — and any proposed value is overlaid on its
    value cell (tier-colored too when color_values, e.g. the handoff AllFillsBackend tab).
    With sheet_row_col (the sweep Backend mirror) a single leading SheetRow locator column
    (the tracker's own row number, not a backend field) rides in front so Baird can find
    each scattered row; the AllFillsBackend paste tabs pass sheet_row_col=False so EVERY
    column aligns 1:1 with the sheet and blocks copy-paste with no offset (rows located by
    ProjectID). Owner/Parent refs are staged on the separate
    Operators/Owners tab, so they never overlay here (there is no Owner [ref] backend column).
    Falls back to identity columns only if the snapshot header can't be loaded."""
    # group resolutions by segment (ProjectID, SheetRow), first-seen order; index by ref_col
    seg_order: list[tuple] = []
    segs: dict[tuple, dict] = {}
    for r in resolutions:
        sk = (r.get("project_id", ""), r.get("sheet_row", ""))
        if sk not in segs:
            segs[sk] = {"base": r, "by_ref": {}}
            seg_order.append(sk)
        rc = r.get("ref_col")
        if rc and rc not in (VALIDITY_REF, STATUS_REF):
            segs[sk]["by_ref"].setdefault(rc, r)

    # fall back to a minimal identity header if the snapshot couldn't be loaded
    header = list(backend_header) if backend_header else \
        ["ProjectID", "PipelineName", "SegmentName"]

    ws = wb.create_sheet(title)
    headers = (["SheetRow"] if sheet_row_col else []) + header
    ws.append(headers)

    # 1-based sheet-column index of each backend column (offset by the SheetRow col if present)
    off = 2 if sheet_row_col else 1
    col_idx = {h: i + off for i, h in enumerate(header)}
    ref_idx = {h: ci for h, ci in col_idx.items() if h.endswith(" [ref]")}

    for sk in seg_order:
        pid, srow = sk
        current = snapshot_rows.get((pid, srow), {})
        b = segs[sk]["base"]
        # prefill from the snapshot; if the row is missing, seed the identity cells we know
        if not current:
            current = {"ProjectID": b.get("project_id", ""),
                       "PipelineName": b.get("pipeline_name", ""),
                       "SegmentName": b.get("segment_name", "")}
        ws.append(([srow] if sheet_row_col else []) + [current.get(h, "") for h in header])
        rn = ws.max_row
        for rc, r in segs[sk]["by_ref"].items():
            # overlay proposed value(s) onto their backend value cells (skip cols off-schema).
            # In color_values mode (handoff paste tabs) a non-FILL unit is ref-only work —
            # its values are the CURRENT sheet values, so leave the prefilled value cell
            # untinted (only a genuinely proposed value earns a tier color).
            for vc, vv in (r.get("values") or {}).items():
                ci = col_idx.get(vc)
                if ci and str(vv).strip():
                    if color_values and r.get("class_in") != "FILL":
                        continue
                    vcell = ws.cell(rn, ci, vv)
                    if color_values:
                        vcell.fill = _ref_cell_fill(r)
            # overlay proposed ref text onto the [ref] cell + color by corroboration tier
            ci = ref_idx.get(rc)
            if ci:
                cell = ws.cell(rn, ci, _ref_cell_text(r))
                cell.fill = _ref_cell_fill(r)
                cell.alignment = Alignment(wrap_text=False, vertical="top")

    if sheet_row_col:
        ws.column_dimensions["A"].width = 9
    for h, ci in col_idx.items():
        ws.column_dimensions[get_column_letter(ci)].width = 46 if h.endswith(" [ref]") else 16
    _style_header(ws, len(headers))
    # keep the locator/identity columns (through ProjectID) visible while scrolling
    anchor = get_column_letter(col_idx["ProjectID"] + 1) if "ProjectID" in col_idx else "B"
    ws.freeze_panes = f"{anchor}2"
    return ws


def _prune_trailing_empty(value_cols, resolutions, ref_col):
    """Drop trailing value cols that are blank across every in-scope row for this ref
    (keeps the paste left-aligned; unused Owner2..Owner11 slots stay off-sheet). Always
    keep at least the first value col."""
    used = set()
    for r in resolutions:
        if r.get("ref_col") != ref_col:
            continue
        for c, v in r.get("values", {}).items():
            if str(v).strip():
                used.add(c)
    last = -1
    for i, c in enumerate(value_cols):
        if c in used:
            last = i
    return value_cols[: last + 1] if last >= 0 else value_cols[:1]


def _operators_owners_view(wb, title, resolutions):
    """Paste-ready mirror of the backend "Pipeline operators/owners" tab (GID 1489950650),
    where the `[ref]` column PRECEDES its values. ProjectID-keyed (one row per ProjectID);
    for each ref unit the `[ref]` cell — carrying the proposed ref(s), color-coded by
    corroboration tier — comes FIRST, then its value columns for context. Trailing all-empty
    owner slots are pruned. Baird pastes each `[ref]` cell back onto that tab by ProjectID."""
    # ref units in native order (Operator [ref] before Owner [ref]), by first appearance;
    # value cols union across records (carried fills may lack value_cols — fall back to
    # their values keys so the fill's value still renders next to its [ref])
    ref_order: list[str] = []
    ref_vcols: dict[str, list[str]] = {}
    for r in resolutions:
        rc = r.get("ref_col")
        if not rc:
            continue
        if rc not in ref_vcols:
            ref_order.append(rc)
            ref_vcols[rc] = []
        for c in (r.get("value_cols") or list((r.get("values") or {}).keys())):
            if c not in ref_vcols[rc]:
                ref_vcols[rc].append(c)
    for rc in ref_order:
        ref_vcols[rc] = _prune_trailing_empty(ref_vcols[rc], resolutions, rc)

    # group by ProjectID (the OO tab is one row per ProjectID)
    proj_order: list[str] = []
    projs: dict[str, dict] = {}
    for r in resolutions:
        pid = r.get("project_id", "")
        if pid not in projs:
            projs[pid] = {"base": r, "by_ref": {}}
            proj_order.append(pid)
        projs[pid]["by_ref"][r.get("ref_col")] = r

    ws = wb.create_sheet(title)
    base = ["ProjectID", "PipelineName", "SegmentName"]
    headers = list(base)
    ref_pos: dict[str, int] = {}      # ref_col -> 1-based column index of its [ref] cell
    for rc in ref_order:
        headers.append(rc)            # ref FIRST (mirrors the tab)
        ref_pos[rc] = len(headers)
        headers.extend(ref_vcols[rc])  # then its value cols
    ws.append(headers)

    for i, w in enumerate([12, 30, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col, h in enumerate(headers[len(base):], start=len(base) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 46 if h.endswith(" [ref]") else 18

    for pid in proj_order:
        b = projs[pid]["base"]
        by_ref = projs[pid]["by_ref"]
        rowvals = [pid, b.get("pipeline_name", ""), b.get("segment_name", "")]
        for rc in ref_order:
            r = by_ref.get(rc)
            rowvals.append(_ref_cell_text(r) if r else "")
            for vc in ref_vcols[rc]:
                rowvals.append(r.get("values", {}).get(vc, "") if r else "")
        ws.append(rowvals)
        rn = ws.max_row
        for rc in ref_order:
            r = by_ref.get(rc)
            if not r:
                continue
            cell = ws.cell(rn, ref_pos[rc])   # the [ref] cell
            cell.fill = _ref_cell_fill(r)
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            # a FILL unit proposes a NEW value — tier-color the value cells too (value and
            # its paired [ref] travel together); never color an empty cell
            if r.get("class_in") == "FILL":
                for off in range(1, len(ref_vcols[rc]) + 1):
                    vcell = ws.cell(rn, ref_pos[rc] + off)
                    if str(vcell.value or "").strip():
                        vcell.fill = _ref_cell_fill(r)

    _style_header(ws, len(headers))
    ws.freeze_panes = "D2"   # keep ProjectID..SegmentName visible while scrolling
    return ws


# --- route suggestions (deep-sweep extension) ------------------------------ #
def _fmt_coord(v):
    """A lat/lon for display — blank when unsourced (coordinates are never fabricated)."""
    return "" if v is None or v == "" else v


def _fmt_waypoints(r: dict) -> str:
    wps = r.get("waypoints") or []
    parts = []
    for w in wps:
        nm = (w.get("name") or "").strip()
        lat, lon = w.get("lat"), w.get("lon")
        coord = f" ({lat},{lon})" if lat is not None and lon is not None else ""
        parts.append(f"{nm}{coord}".strip())
    txt = "; ".join(p for p in parts if p)
    note = (r.get("waypoint_note") or "").strip()
    return (txt + (f" — {note}" if note else "")) if (txt or note) else ""


def _route_columns(with_source=False):
    g = lambda k: (lambda r: r.get(k, ""))
    cols = [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 28),
        ("SegmentName", g("segment_name"), 22),
        ("Current RouteAccuracy", g("current_route_accuracy"), 16),
        ("Suggested RouteAccuracy", g("suggested_route_accuracy"), 16),
        ("Start", g("start_name"), 30),
        ("Start lat", lambda r: _fmt_coord(r.get("start_lat")), 10),
        ("Start lon", lambda r: _fmt_coord(r.get("start_lon")), 10),
        ("End", g("end_name"), 30),
        ("End lat", lambda r: _fmt_coord(r.get("end_lat")), 10),
        ("End lon", lambda r: _fmt_coord(r.get("end_lon")), 10),
        ("Waypoints", _fmt_waypoints, 40),
        ("Corridor description", g("corridor_desc"), 72),
        ("Proposed ref(s)", lambda r: J(r.get("proposed_refs", [])), 52),
        ("Verification status", _verif_summary, 20),
        ("Corroboration tier", g("tier"), 13),
        ("Independent?", lambda r: "yes" if r.get("independent") else ("no" if r.get("proposed_refs") else ""), 11),
        ("Source language", g("source_language"), 12),
        ("ResearcherNotes", g("researcher_notes"), 50),
        ("Wiki (visited, not cited)", g("wiki"), 34),
    ]
    if with_source:
        cols.append(("Source packet", g("source_dir"), 28))
    return cols


def _route_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}
    tier_c = idx["Corroboration tier"]
    coord_cs = [idx[h] for h in ("Start lat", "Start lon", "End lat", "End lon")]

    def styler(ws, rn, r):
        ws.cell(rn, tier_c).fill = CONF_FILL.get(_tier_color(r), PatternFill())
        # corridor-only suggestion (endpoints not both coordinated) → yellow the empty
        # coord cells so it reads as "needs a sourced trace", not "ready to digitize"
        if r.get("class_out") == "ROUTE_PARTIAL":
            for cc in coord_cs:
                if not str(ws.cell(rn, cc).value or "").strip():
                    ws.cell(rn, cc).fill = CONF_FILL["yellow"]
    return styler


def _route_candidate_columns():
    """§8 route-creation candidates: staged <PID>.geojson geometry + provenance.
    Destination is the ROUTES REPO (human branch+PR), never the sheet or an auto-edit."""
    g = lambda k: (lambda r: r.get(k, ""))
    src = lambda k: (lambda r: (r.get("source") or {}).get(k, ""))

    def _license(r):
        s = r.get("source") or {}
        return s.get("license") or ("ODbL" if s.get("odbl") else "")

    def _gr(k):
        return lambda r: (r.get("georef") or {}).get(k, "") if r.get("georef") else ""

    def _sig(k):
        return lambda r: (r.get("geometry_signals") or {}).get(k, "")

    return [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 10),
        ("PipelineName", g("pipeline_name"), 28),
        ("SegmentName", g("segment_name"), 22),
        ("Current RouteAccuracy", g("current_route_accuracy"), 16),
        ("Suggested RouteAccuracy", g("suggested_route_accuracy"), 16),
        ("Method", g("method"), 18),
        ("Geometry file", g("geometry_file"), 34),
        ("Length km", g("length_km"), 10),
        ("Sheet km", g("sheet_length_km"), 10),
        ("Ratio", g("length_ratio"), 8),
        ("Source", src("name"), 24),
        ("License", _license, 14),
        ("Georef RMSE km", _gr("rmse_km"), 12),
        ("Georef GCPs", _gr("n_gcps"), 10),
        ("QC result", lambda r: "pass" if r.get("qc_passed") else "FAIL", 10),
        ("Replacement?", lambda r: "yes" if r.get("replacement") else "", 12),
        ("Route IoU", _sig("iou"), 9),
        ("Route g_score", _sig("g_score"), 11),
        ("Packet?", lambda r: "yes" if (r.get("packet") or (r.get("georef") and not (r.get("georef") or {}).get("pass", True))) else "", 8),
        ("Proposed ref(s)", lambda r: J(r.get("proposed_refs", [])), 52),
        ("Verification status", _verif_summary, 20),
        ("Corroboration tier", g("tier"), 13),
        ("Independent?", lambda r: "yes" if r.get("independent") else ("no" if r.get("proposed_refs") else ""), 11),
        ("Source URL", src("url"), 40),
        ("ResearcherNotes", g("researcher_notes"), 50),
    ]


def _route_candidate_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}

    def styler(ws, rn, r):
        ws.cell(rn, idx["Corroboration tier"]).fill = CONF_FILL.get(_tier_color(r), PatternFill())
        # a route-replacement candidate: yellow the Replacement? cell (routes-repo convention)
        if r.get("replacement"):
            ws.cell(rn, idx["Replacement?"]).fill = CONF_FILL["yellow"]
        # gate failure is loud, not dropped — red the QC cell so the reviewer sees it
        if not r.get("qc_passed"):
            ws.cell(rn, idx["QC result"]).fill = CONF_FILL["red"]
    return styler


# --- GulfPub cross-comparison (deep-sweep extension) ------------------------ #
def _gulfpub_view(wb, title, crosswalk: dict):
    """Flat cross-comparison of the scoped GulfPub (PE World Map) reconciliation against GEM:
    one row per overlap (matched pair), then GulfPub-only additions, then ambiguous matches.
    GulfPub is a Tier-2 source — a single value here NEVER reaches green on its own; conflicts
    route to Update's normal ≥2-independent source search. Read-and-flag only, nothing applied.
    Reads the crosswalk produced by build_gulfpub_crosswalk.py (from reconcile's match_diff.json)."""
    cols = [
        ("Kind", 18), ("GEM ProjectID", 13), ("GEM name", 28), ("GEM segment", 20),
        ("GulfPub name", 28), ("Match conf", 11), ("Composite", 10),
        ("GEM status", 12), ("GP status", 12), ("Status conflict", 14),
        ("GEM diam", 10), ("GP diam", 10), ("Diam flag", 12),
        ("GEM len km", 11), ("GP len km", 11), ("Len flag", 12),
        ("GEM route acc", 14), ("GP start", 26), ("GP end", 26), ("GP has geom", 11),
        ("GP operator", 24), ("GP owners", 24), ("GP capacity", 11), ("GP startyear", 11),
        ("Candidates", 30), ("GP description", 60),
    ]
    field = {
        "Kind": "kind", "GEM ProjectID": "gem_pid", "GEM name": "gem_name",
        "GEM segment": "gem_segment", "GulfPub name": "gulfpub_name", "Match conf": "match_conf",
        "Composite": "composite", "GEM status": "gem_status", "GP status": "gp_status",
        "Status conflict": "status_conflict", "GEM diam": "gem_diam", "GP diam": "gp_diam",
        "Diam flag": "diam_flag", "GEM len km": "gem_len_km", "GP len km": "gp_len_km",
        "Len flag": "len_flag", "GEM route acc": "gem_route_acc", "GP start": "gp_start",
        "GP end": "gp_end", "GP has geom": "gp_has_geom", "GP operator": "gp_operator",
        "GP owners": "gp_owners", "GP capacity": "gp_capacity", "GP startyear": "gp_startyear",
        "Candidates": "candidates", "GP description": "gp_desc",
    }
    ws = wb.create_sheet(title)
    headers = [h for h, _ in cols]
    ws.append(headers)
    for i, (_, wdt) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    idx = {h: i + 1 for i, (h, _) in enumerate(cols)}

    rows = (crosswalk.get("overlaps") or []) + (crosswalk.get("additions") or []) \
        + (crosswalk.get("ambiguous") or [])
    n = 0
    for rec in rows:
        vals = []
        for h, _ in cols:
            v = rec.get(field[h], "")
            vals.append("" if v is None else v)
        ws.append(vals)
        rn = ws.max_row
        n += 1
        # match confidence cell → tier color
        mc = str(rec.get("match_conf", "")).lower()
        if mc in ("green", "yellow", "red"):
            ws.cell(rn, idx["Match conf"]).fill = CONF_FILL[mc]
        # status/spec disagreements → red/yellow flags
        if str(rec.get("status_conflict", "")).upper() == "CONFLICT":
            ws.cell(rn, idx["Status conflict"]).fill = CONF_FILL["red"]
        for flag_col, cell_col in (("diam_flag", "Diam flag"), ("len_flag", "Len flag")):
            fv = str(rec.get(flag_col, "")).lower()
            if fv and fv not in ("ok", ""):
                ws.cell(rn, idx[cell_col]).fill = CONF_FILL["yellow"]

    _style_header(ws, len(headers))
    ws.freeze_panes = "E2"   # keep Kind..GulfPub name visible while scrolling
    return ws, n


# --- wiki-alignment / route-integrity / mechanical flags (QC-workflow extensions) --- #
def _wikidiff_columns():
    g = lambda k: (lambda r: r.get(k, ""))
    # wiki link leftmost — the researcher's first move on every row is opening the page
    return [
        ("Wiki page (visited, not cited)", g("wiki"), 34),
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 20),
        ("Field", g("field"), 18),
        ("Wiki key", g("wiki_key"), 20),
        ("Sheet value", g("sheet_value"), 26),
        ("Sheet (norm)", g("sheet_value_norm"), 20),
        ("Wiki value", lambda r: str(r.get("wiki_value", ""))[:250], 40),
        ("Wiki (norm)", g("wiki_value_norm"), 20),
        ("Staged value", g("staged_value"), 18),
        ("Staged source", g("staged_source"), 26),
        ("Class", g("class_out"), 21),
        ("Severity", g("severity"), 9),
        ("Action", g("action"), 90),
        ("Known — staged", g("staged_note"), 44),
    ]


def _wikidiff_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}

    def _tint(ws, rn, header, color):
        cell = ws.cell(rn, idx[header])
        if str(cell.value or "").strip():   # never color an empty cell
            cell.fill = CONF_FILL[color]

    def styler(ws, rn, r):
        cls = r.get("class_out", "")
        cls_cell = ws.cell(rn, idx["Class"])
        note = ws.cell(rn, idx["Known — staged"])
        if str(note.value or "").strip():   # a prior staged packet already covers it
            note.fill = BLUE_FILL
        if r.get("severity") != "flag":     # info = review context, not a hard flag
            cls_cell.fill = CONF_FILL["yellow"]
            return
        if cls == "WIKI_UPDATE":            # wiki lags the sheet → edit the wiki
            cls_cell.fill = CONF_FILL["red"]
            _tint(ws, rn, "Wiki value", "red")
        elif cls == "SHEET_SUSPECT":        # sheet side is the suspect → verify + fix sheet
            cls_cell.fill = CONF_FILL["red"]
            _tint(ws, rn, "Sheet value", "red")
        elif cls == "WIKI_STALE_VS_STAGED":  # a staged sheet edit supersedes both
            cls_cell.fill = CONF_FILL["yellow"]
            _tint(ws, rn, "Staged value", "yellow")
        elif cls == "UNPARSED":
            cls_cell.fill = CONF_FILL["red"]
    return styler


def _routeqc_columns():
    g = lambda k: (lambda r: r.get(k, ""))
    return [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 20),
        ("RouteAccuracy", g("route_accuracy"), 13),
        ("Check", g("check"), 16),
        ("Measured (route)", lambda r: str(r.get("measured", "")), 36),
        ("Expected (sheet)", lambda r: str(r.get("expected", "")), 36),
        ("Detail", g("detail"), 90),
        ("Severity", g("severity"), 9),
        ("Known — staged", g("staged_note"), 44),
    ]


def _routeqc_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}

    def styler(ws, rn, r):
        color = "red" if r.get("severity") == "flag" else "yellow"
        ws.cell(rn, idx["Check"]).fill = CONF_FILL[color]
        note = ws.cell(rn, idx["Known — staged"])
        if str(note.value or "").strip():
            note.fill = BLUE_FILL
    return styler


def _qc_flags_view(wb, title, qc_flags: dict):
    """Mechanical-check flags (build_qc_workbook.run_checks, country-scoped) from the
    qc_flags.json sidecar. Red Detail = open flag; blue staged note = a prior staged
    packet already covers it (don't re-research)."""
    cols = [("Check", 20), ("ProjectID", 12), ("SheetRow", 9), ("PipelineName", 30),
            ("SegmentName", 20), ("CountriesOrAreas", 22), ("Detail", 72),
            ("Known — staged", 44)]
    field = {"Check": "check", "ProjectID": "project_id", "SheetRow": "sheet_row",
             "PipelineName": "pipeline_name", "SegmentName": "segment_name",
             "CountriesOrAreas": "countries", "Detail": "detail",
             "Known — staged": "staged_note"}
    ws = wb.create_sheet(title)
    ws.append([h for h, _ in cols])
    for i, (_, wdt) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    idx = {h: i + 1 for i, (h, _) in enumerate(cols)}
    n = 0
    for f in qc_flags.get("flags", []):
        ws.append([f.get(field[h], "") for h, _ in cols])
        rn = ws.max_row
        n += 1
        if str(f.get("staged_note") or "").strip():
            ws.cell(rn, idx["Known — staged"]).fill = BLUE_FILL
        else:
            ws.cell(rn, idx["Detail"]).fill = CONF_FILL["red"]
    _style_header(ws, len(cols))
    ws.freeze_panes = "A2"
    return ws, n


def _concerns_columns():
    """The handoff GATEKEEPER view: every validity concern for the scope — carried from
    prior staged packets (staged_actions.json) and, in the split handoff, this packet's
    own Leg-3 findings, ALL types. Read-and-review only, nothing applied."""
    g = lambda k: (lambda r: r.get(k, ""))
    return [
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 20),
        ("Concern", g("concern_type"), 14),
        ("Verdict", g("verdict"), 18),
        ("Recommendation", g("recommendation"), 44),
        ("Finding", g("researcher_notes"), 70),
        ("Sources", lambda r: J(r.get("proposed_refs", [])), 52),
        ("Corroboration tier", g("tier"), 13),
        ("Also flagged (this packet)", lambda r: J(r.get("also_flagged", [])), 34),
        ("Source packet", g("source_dir"), 30),
    ]


def _concerns_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}

    def styler(ws, rn, r):
        ct = r.get("concern_type", "")
        ws.cell(rn, idx["Concern"]).fill = \
            CONF_FILL["red" if ct in _HIGH_CONCERN else "yellow"]
        if r.get("proposed_refs"):
            ws.cell(rn, idx["Corroboration tier"]).fill = \
                CONF_FILL.get(_tier_color(r), PatternFill())
        also = ws.cell(rn, idx["Also flagged (this packet)"])
        if str(also.value or "").strip():   # cross-referenced by this run's own checks
            also.fill = BLUE_FILL
    return styler


def _refwork_styler(columns):
    """Mixed-bucket styler for the handoff <Cmdty>_RefWork tab (REFS_ADDED /
    DEAD_LINK / UNRESOLVED in one sheet): dispatch to the per-bucket styler."""
    stylers = {b: _make_styler(columns, b) for b in _ORDER}

    def styler(ws, rn, r):
        s = stylers.get(r.get("class_out", ""))
        if s:
            s(ws, rn, r)
    return styler


# --- handoff split: actions workbook + evidence workbook -------------------- #
# The handoff deliverable is TWO files (decision 2026-07-16): <stem>-actions.xlsx
# holds ONLY what the researcher must act on, grouped by destination (sheet paste /
# wiki edit / routes-repo PR / open issues); <stem>-evidence.xlsx holds the audit
# trail (confirmed audits, verification detail, covered flags). No confirmed /
# known-staged / info-only row ever reaches the actions file.

_THIS_PACKET = "(this packet)"


def _norm_decision(r: dict) -> dict:
    """Normalize a carried concern (staged_actions.json) or this packet's own Leg-3
    validity finding into one Decisions/ConfirmedAudit record: explicit
    verdict/concern_type win; own rows fall back to the keyword classifiers and mark
    Source packet as (this packet)."""
    d = dict(r)
    d["verdict"] = (r.get("verdict") or "").strip() or _validity_verdict(r)
    d["concern_type"] = (r.get("concern_type") or "").strip() or _concern_type(r)
    if not str(d.get("source_dir") or "").strip():
        d["source_dir"] = _THIS_PACKET
    return d


def _is_open_verdict(v: str) -> bool:
    return not str(v or "").strip().lower().startswith("confirm")


def _confirmed_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}

    def styler(ws, rn, r):
        ws.cell(rn, idx["Verdict"]).fill = CONF_FILL["green"]
        if r.get("proposed_refs"):
            ws.cell(rn, idx["Corroboration tier"]).fill = \
                CONF_FILL.get(_tier_color(r), PatternFill())
    return styler


def _wikiupdates_columns():
    g = lambda k: (lambda r: r.get(k, ""))
    # wiki link leftmost — the researcher's first move on every row is opening the page
    return [
        ("Wiki page (visited, not cited)", g("wiki"), 34),
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 20),
        ("Field", g("field"), 18),
        ("Wiki says (current)", lambda r: str(r.get("wiki_value", ""))[:250], 40),
        ("Write this (sheet value)", g("sheet_value"), 26),
        ("Action", g("action"), 90),
    ]


def _wikiupdates_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}

    def styler(ws, rn, r):
        stale = ws.cell(rn, idx["Wiki says (current)"])
        if str(stale.value or "").strip():   # never color an empty cell
            stale.fill = CONF_FILL["red"]
    return styler


def _openflags_columns():
    g = lambda k: (lambda r: r.get(k, ""))
    return [
        ("Issue type", g("issue_type"), 22),
        ("ProjectID", g("project_id"), 12),
        ("SheetRow", g("sheet_row"), 9),
        ("PipelineName", g("pipeline_name"), 30),
        ("SegmentName", g("segment_name"), 20),
        ("Field / check", g("field_check"), 20),
        ("Detail", g("detail_txt"), 90),
        ("Suggested next step", g("next_step"), 56),
        ("Source packet", g("source_dir"), 28),
    ]


def _openflags_styler(columns):
    idx = {h: i + 1 for i, (h, _, _) in enumerate(columns)}

    def styler(ws, rn, r):
        ws.cell(rn, idx["Detail"]).fill = CONF_FILL["red"]
    return styler


def _collect_open_flags(qc_flags, routeqc_res, wd_unparsed, unresolved_refs):
    """Normalize the genuinely OPEN residue — flags no staged packet covers and ref
    units that could not be resolved — into the actions workbook's OpenFlags rows."""
    base = lambda r: {k: r.get(k, "") for k in
                      ("project_id", "sheet_row", "pipeline_name", "segment_name")}
    rows = []
    for f in qc_flags:
        if str(f.get("staged_note") or "").strip():
            continue
        rows.append({**base(f), "issue_type": "mechanical",
                     "field_check": f.get("check", ""),
                     "detail_txt": f.get("detail", ""),
                     "next_step": "verify independently, then fix on the sheet "
                                  "(QC detects, Update fixes)",
                     "source_dir": _THIS_PACKET})
    for r in routeqc_res:
        if str(r.get("staged_note") or "").strip() or r.get("severity") != "flag":
            continue
        rows.append({**base(r), "issue_type": "route integrity",
                     "field_check": r.get("check", ""),
                     "detail_txt": f"{r.get('detail', '')} (measured: {r.get('measured', '')}; "
                                   f"expected: {r.get('expected', '')})",
                     "next_step": "review the drawn route vs the row's attributes; a route "
                                  "fix goes via a human branch+PR against "
                                  "GOIT-GGIT-pipeline-routes",
                     "source_dir": _THIS_PACKET})
    for r in wd_unparsed:
        rows.append({**base(r), "issue_type": "wiki unparsed",
                     "field_check": r.get("field", ""),
                     "detail_txt": r.get("action", "") or "wiki page missing or unparseable",
                     "next_step": "open the wiki page manually; fix the Wiki link or the page",
                     "source_dir": _THIS_PACKET})
    for r in unresolved_refs:
        rows.append({**base(r), "issue_type": "ref unresolved",
                     "field_check": r.get("ref_col", ""),
                     "detail_txt": r.get("researcher_notes", "") or
                                   "could not reach 2 working, independent, "
                                   "value-containing links",
                     "next_step": "manual source search — never fabricate a URL "
                                  "(standing rule 2)",
                     "source_dir": r.get("source_dir") or _THIS_PACKET})
    return rows


def _split_readme(ws, meta, sheet_defs, actions_file: bool, companion: str):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    scope = meta.get("scope", {})
    counts = meta.get("counts", {})
    cap = (meta.get("commodity") or scope.get("tracker") or "gas").capitalize()
    if actions_file:
        title = "GEM handoff packet — ACTIONS (suggested changes + open issues)"
        work_from = (
            f"Work top to bottom. 1) {cap}_Decisions — settle existence / duplicate / "
            f"classification / attribution doubts FIRST (don't invest work in a row GEM "
            f"may stop tracking). 2) {cap}_StatusChanges — status edits (verdict != "
            f"confirm; confirms are counts-only). 3) {cap}_AllFillsBackend — THE paste "
            f"surface: every corroborated fill AND paste-ready reference for the scope, "
            f"overlaid on the full backend layout; paste the colored cells only. "
            f"4) {cap}_OperatorsOwners — same, for the ProjectID-keyed operators/owners "
            f"tab. 5) {cap}_NewRows / _NewRowRefs / _MatchedExisting — discovery output. "
            f"6) {cap}_WikiUpdates — gem.wiki edits (different destination: the wiki, not "
            f"the sheet). 7) {cap}_RouteSuggestions — candidates for a human routes-repo "
            f"branch+PR. 8) {cap}_OpenFlags — unresolved issues needing human judgment. "
            "Confirmed audits, per-fill/per-ref verification detail, and already-covered "
            "flags live in the companion evidence workbook.")
        color_key = (
            "[ref]/value cell color = corroboration tier: green=≥2 independent working "
            "sources / yellow=single / red=low or none / blue=re-verified existing ref. "
            "Decisions: red Concern = existence/duplicate/classification (settle before "
            "any other work on the row), yellow = attribution/spec. OpenFlags: red Detail "
            "= open, uncovered issue. On the backend paste tabs, NEVER paste the "
            "computed/formula columns — colored cells only.")
    else:
        title = "GEM handoff packet — EVIDENCE & audit trail (no action required)"
        work_from = (
            "Nothing in this file asks for an edit. It records what was checked and "
            "cleared (ConfirmedAudit), the per-fill and per-ref verification detail "
            "behind the actions workbook's paste cells (FillDetail / RefWorkDetail), the "
            "sheet↔wiki diff context that needs no wiki edit, flags a staged packet "
            "already covers, and the discovery monitor list. Read it when you want to "
            "know WHY an action row says what it says — Source packet names the staging "
            "dir holding the canonical record.")
        color_key = (
            "Blue 'Known — staged' = a prior staged packet already covers it (apply that "
            "packet; no new research); blue 'audited — no existence concern' = the row "
            "already passed a deep-sweep existence audit. Green Verdict = confirmed. "
            "WikiAlignment: yellow Class = info-severity review context, or a staged "
            "correction supersedes the diff (WIKI_STALE_VS_STAGED — apply that packet "
            "first, then align the wiki). Tier colors as in the actions file.")
    rows = [
        (title, ""),
        ("Tracker / commodity", meta.get("commodity", scope.get("tracker", ""))),
        ("Country", scope.get("country", meta.get("country", ""))),
        ("GEM CSV", scope.get("csv", "")),
        ("Prior staged packets", J(meta.get("staged_dirs", []))),
        ("ESCALATIONS", J([f"{e.get('title','')} — {e.get('summary','')} [{e.get('memo','')}]"
                           for e in meta.get("escalations", [])])),
        ("Companion file", companion),
        ("", ""),
        ("Counts", J([f"{k}={v}" for k, v in counts.items()])),
        ("", ""),
        ("Work from", work_from),
        ("Color key", color_key),
        ("Standing rules", "gem.wiki is VISITED for the diff but NEVER cited as a source (rule 1). "
                           "Never fabricate a URL (rule 2). Every researched ref passed url_verifier. "
                           "Nothing here is auto-applied — sheet edits are pasted manually, wiki "
                           "edits made manually, and a route is NEVER auto-replaced (routes-repo "
                           "fixes go via a separate human branch+PR)."),
        ("", ""),
        ("Sheets", ""),
    ]
    for r in rows:
        ws.append(r)
    for name, desc in sheet_defs:
        ws.append((name, desc))
    ws["A1"].font = Font(bold=True, size=13)
    for c in (1, 2):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF", size=13)
    for rn in range(2, ws.max_row + 1):
        ws.cell(rn, 1).font = Font(bold=True)
        ws.cell(rn, 2).alignment = Alignment(wrap_text=False, vertical="top")
    ws.freeze_panes = "A2"


def _build_handoff(staging: Path, out: Path, meta: dict, parts: dict, actions: dict) -> None:
    """Handoff mode (staged_actions.json present): write the two-workbook deliverable.

    <stem>-actions.xlsx   tab order = work order: Decisions → StatusChanges →
                          AllFillsBackend (fills + paste-ready refs unified) →
                          OperatorsOwners → NewRows/NewRowRefs/MatchedExisting →
                          WikiUpdates → RouteSuggestions → OpenFlags.
    <stem>-evidence.xlsx  ConfirmedAudit → FillDetail → RefWorkDetail → WikiAlignment
                          (non-action remainder) → RouteIntegrity (covered/info) →
                          Flags (covered) → MonitorList → GulfPub (if crosswalked).
    """
    prefix = (meta.get("commodity") or meta.get("scope", {}).get("tracker") or "oil").capitalize()
    suffix = out.suffix or ".xlsx"
    stem = out.name[: -len(out.suffix)] if out.suffix else out.name
    out_a = out.with_name(f"{stem}-actions{suffix}")
    out_b = out.with_name(f"{stem}-evidence{suffix}")
    for p in (out_a, out_b):
        if p.exists():
            sys.exit(f"refusing to overwrite existing {p} (use a fresh <stamp>)")

    backend_header, snapshot_rows = _backend_snapshot(meta)
    wb_a, wb_b = Workbook(), Workbook()
    wb_a.remove(wb_a.active)
    wb_b.remove(wb_b.active)
    readme_a = wb_a.create_sheet("README")
    readme_b = wb_b.create_sheet("README")
    defs_a: list[tuple] = []
    defs_b: list[tuple] = []

    # ---- partition every input into act-on vs evidence ----
    decisions = [_norm_decision(r) for r in actions.get("concerns", [])] + \
                [_norm_decision(r) for r in parts["validity"]]
    open_dec = sorted((r for r in decisions if _is_open_verdict(r["verdict"])),
                      key=lambda r: (0 if r["concern_type"] in _HIGH_CONCERN else 1,
                                     _int_or_zero(r.get("sheet_row"))))
    confirmed = [r for r in decisions if not _is_open_verdict(r["verdict"])]

    own_status_open = [r for r in parts["status"]
                       if (r.get("verdict") or "").strip().lower() != "confirm"]
    status_rows = actions.get("status_changes", []) + \
        [dict(r, source_dir=r.get("source_dir") or _THIS_PACKET) for r in own_status_open]
    own_status_confirms = len(parts["status"]) - len(own_status_open)

    own_fills = [r for r in parts["fill"] if r.get("class_out") == "REFS_ADDED"]
    pending_fills = actions.get("fills", [])
    ref_work = actions.get("ref_work", [])
    own_refunits = parts["tracker"] + parts["oo"]

    is_oo = lambda r: r.get("tab") == "operators_owners" or r.get("ref_col") in OO_PRIMARY
    paste_refs = [r for r in own_refunits + ref_work
                  if r.get("class_out") in ("REFS_ADDED", "DEAD_LINK")]
    # one paste surface: fills first so a fill wins any same-cell tie with ref-only work
    afb_res = sorted(
        [f for f in own_fills + pending_fills if not is_oo(f)] +
        [r for r in paste_refs if not is_oo(r)],
        key=lambda r: (_int_or_zero(r.get("sheet_row")), r.get("ref_col", "")))
    oo_units = [f for f in own_fills + pending_fills if is_oo(f)] + \
               [r for r in paste_refs if is_oo(r)]
    unresolved_refs = [r for r in own_refunits + ref_work
                       if r.get("class_out") == "UNRESOLVED"]

    wikidiff = parts["wikidiff"]
    is_action_wd = lambda r: (r.get("class_out") == "WIKI_UPDATE"
                              and r.get("severity") == "flag")
    wiki_updates = [r for r in wikidiff if is_action_wd(r)]
    wd_unparsed = [r for r in wikidiff if r.get("class_out") == "UNPARSED"]
    wiki_rest = [r for r in wikidiff
                 if not is_action_wd(r) and r.get("class_out") != "UNPARSED"]

    routeqc = parts["routeqc"]
    covered_routeqc = [r for r in routeqc
                       if str(r.get("staged_note") or "").strip()
                       or r.get("severity") != "flag"]

    qcf_path = staging / "qc_flags.json"
    qc_flags = json.loads(qcf_path.read_text()).get("flags", []) if qcf_path.exists() else []
    covered_flags = [f for f in qc_flags if str(f.get("staged_note") or "").strip()]
    open_flags = _collect_open_flags(qc_flags, routeqc, wd_unparsed, unresolved_refs)

    routes_all_raw = parts["route"] + actions.get("routes", [])
    is_route_candidate = lambda r: r.get("class_out") == "ROUTE_CANDIDATE"
    routes_all = [r for r in routes_all_raw if not is_route_candidate(r)]
    route_candidates = [r for r in routes_all_raw if is_route_candidate(r)]
    new_rows_all = actions.get("new_rows", [])
    nr_new = [c for c in new_rows_all if c.get("class") == "new_row"]
    nr_monitor = [c for c in new_rows_all if c.get("class") == "monitor"]
    nr_matched = [c for c in new_rows_all if c.get("class") == "matched_existing"]

    # ---- ACTIONS workbook (tab order = work order) ----
    c_cols = _concerns_columns()
    if open_dec:
        t = f"{prefix}_Decisions"
        _write_sheet(wb_a, t, c_cols, open_dec, _concerns_styler(c_cols))
        defs_a.append((t,
                       f"{len(open_dec)} — READ FIRST: every OPEN validity concern for the scope "
                       "(carried from prior staged packets + this packet's own findings; verdicts "
                       "that confirmed the row are in the evidence file). Red Concern = existence / "
                       "duplicate / classification — the row may be phantom, a relabel of another "
                       "ProjectID, or misclassified: settle these BEFORE any other work on the row. "
                       "Yellow = attribution/spec. Blue 'Also flagged' = this packet's own checks "
                       "hit the same issue. Source packet = the staging dir with the full finding."))
    if status_rows:
        s_cols = _status_columns(with_source=True)
        t = f"{prefix}_StatusChanges"
        _write_sheet(wb_a, t, s_cols, status_rows, _status_styler(s_cols))
        defs_a.append((t,
                       f"{len(status_rows)} — status changes for the scope (carried + this packet's "
                       "own; verdict != confirm — confirms are counts-only in the README). Verdict "
                       "yellow=change (evidence-based new status, refs verified) / red=stale "
                       "(dormancy rule -> inferred shelved/cancelled, ShelvedCancelledType=Presumed, "
                       "no ref by design) or unclear. NOT auto-applied."))
    if afb_res:
        t = f"{prefix}_AllFillsBackend"
        _backend_view(wb_a, t, afb_res, backend_header, snapshot_rows, color_values=True,
                      sheet_row_col=False)
        defs_a.append((t,
                       f"PASTE-READY ({len(afb_res)} cell units) — THE one paste surface for the "
                       "tracker tab: ALL corroborated fills AND all paste-ready reference work "
                       "(new refs + dead-link replacements) for the scope, carried + this packet's "
                       "own, unified in the exact GEM backend layout: FULL column set in sheet "
                       "order, one row per touched segment, current values prefilled. No extra "
                       "locator column — every column aligns 1:1 with the sheet, so cells "
                       "copy-paste with no offset; locate each row by ProjectID. "
                       "A tier-colored VALUE cell = a proposed new value "
                       "(its [ref] cell is colored too — paste them together); a colored [ref] "
                       "cell with an untinted value = ref-only work (value already on the sheet). "
                       "Paste the colored cells only — never the computed/formula columns. "
                       "Verification detail: the evidence file's FillDetail / RefWorkDetail tabs."))
    if oo_units:
        t = f"{prefix}_OperatorsOwners"
        _operators_owners_view(wb_a, t, oo_units)
        defs_a.append((t,
                       f"PASTE-READY ({len(oo_units)} units) — mirror of the separate \"Pipeline "
                       "operators/owners\" backend tab (GID 1489950650), ProjectID-keyed, [ref] "
                       "PRECEDES its values. Owner/Operator fills and refs, carried + this "
                       "packet's own, tier-colored. Paste back by ProjectID — NOT onto a tracker row."))
    if nr_new and backend_header:
        t = f"{prefix}_NewRows"
        _, oo_rows = _new_rows_view(wb_a, t, backend_header, nr_new)
        defs_a.append((t,
                       f"{len(nr_new)} — DISCOVERY candidates (carried from prior staged packets) that "
                       "clear the add-threshold: exact tracker header, one green row per candidate, "
                       "values + verified [ref]s in place. Paste-ready — new rows, not edits."))
        if oo_rows:
            t = f"{prefix}_NewRowRefs"
            _owner_refs_view(wb_a, t, oo_rows)
            defs_a.append((t,
                           f"{len(oo_rows)} — owner/operator [ref]s for the new rows. The main tracker "
                           "has no Owner/Parent [ref] column; apply these on the ProjectID-keyed "
                           "operators/owners tab once each row has an ID. [ref] precedes its value."))
    elif nr_new:
        print(f"  WARN: {len(nr_new)} carried new_row candidates but no tracker snapshot "
              "header — NewRows mirror skipped (still in staged_new.json)")

    def _flag_col2(color):
        def styler(ws, rn, r):
            ws.cell(rn, 2).fill = CONF_FILL[color]
        return styler

    _src_col = ("Source packet", lambda r: r.get("source_dir", ""), 28)
    if nr_matched:
        x_cols = _compact_columns([("Matched ProjectID",
                                    lambda r: r.get("matched_project_id", ""), 16)]) + [_src_col]
        t = f"{prefix}_MatchedExisting"
        _write_sheet(wb_a, t, x_cols, nr_matched, _flag_col2("green"))
        defs_a.append((t,
                       f"{len(nr_matched)} — same physical pipe as an existing GEM row under another "
                       "name: add the candidate name to that row's OtherEnglishNames, no new row."))
    if wiki_updates:
        wu_cols = _wikiupdates_columns()
        t = f"{prefix}_WikiUpdates"
        _write_sheet(wb_a, t, wu_cols, wiki_updates, _wikiupdates_styler(wu_cols))
        defs_a.append((t,
                       f"{len(wiki_updates)} — gem.wiki edits (destination: the WIKI, not the sheet): "
                       "the wiki page lags the verified sheet value — open the page (leftmost link) "
                       "and make the edit in the Action column. Red cell = the stale wiki value to "
                       "replace. gem.wiki is visited, never cited. Info-severity and staged-"
                       "superseded diffs are in the evidence file's WikiAlignment tab."))
    if routes_all:
        carried_routes = actions.get("routes", [])
        rt_cols = _route_columns(with_source=bool(carried_routes))
        t = f"{prefix}_RouteSuggestions"
        _write_sheet(wb_a, t, rt_cols, routes_all, _route_styler(rt_cols))
        defs_a.append((t,
                       f"{len(routes_all)} — route suggestions for low/medium/no-route rows "
                       "(destination: the ROUTES REPO, not the sheet): sourced endpoint coordinates "
                       "+ a corridor description (tier-colored). Yellow coord cells = corridor-only "
                       "(endpoints not both coordinated — no fabricated coordinates, standing rule "
                       "2). These feed a SEPARATE human branch+PR against GOIT-GGIT-pipeline-routes "
                       "— a route is NEVER auto-replaced."))
    if route_candidates:
        rc_cols = _route_candidate_columns()
        t = f"{prefix}_RouteCandidates"
        _write_sheet(wb_a, t, rc_cols, route_candidates, _route_candidate_styler(rc_cols))
        defs_a.append((t,
                       f"{len(route_candidates)} — §8 candidate route GEOMETRY (staged "
                       "<PID>.geojson files, destination: the ROUTES REPO via a human branch+PR). "
                       "Method sets suggested RouteAccuracy (sidecar/gis/osm=high, digitized=medium, "
                       "endpoints=low). License 'ODbL' = OSM-derived, acceptability is Baird's call. "
                       "Red QC result = failed the validation gate (still listed, not applied). "
                       "Yellow Replacement? = a route already exists — review before replacing; a "
                       "route is NEVER auto-replaced."))
    if open_flags:
        of_cols = _openflags_columns()
        t = f"{prefix}_OpenFlags"
        _write_sheet(wb_a, t, of_cols, open_flags, _openflags_styler(of_cols))
        defs_a.append((t,
                       f"{len(open_flags)} — the OPEN residue needing human judgment, nothing staged "
                       "covers it: mechanical-check flags, route-integrity flags, unparseable wiki "
                       "pages, and ref units that could not reach 2 working independent sources. "
                       "Already-covered flags are in the evidence file."))

    # ---- EVIDENCE workbook (audit trail — no action) ----
    if confirmed:
        t = f"{prefix}_ConfirmedAudit"
        _write_sheet(wb_b, t, c_cols, confirmed, _confirmed_styler(c_cols))
        defs_b.append((t,
                       f"{len(confirmed)} — validity checks that CLEARED the row (verdict confirmed, "
                       "possibly with a caveat in the Finding). No action — this is the audit trail "
                       "of what was challenged and survived."))
    fill_detail = pending_fills + \
        [dict(r, source_dir=r.get("source_dir") or _THIS_PACKET) for r in parts["fill"]]
    if fill_detail:
        pf_cols = _fills_columns(with_source=True)
        t = f"{prefix}_FillDetail"
        _write_sheet(wb_b, t, pf_cols, fill_detail, _fills_styler(pf_cols))
        defs_b.append((t,
                       f"{len(fill_detail)} — per-fill evidence for the actions file's paste cells: "
                       "verification status, corroboration tier, notes, Source packet. 'not "
                       "corroborated / dropped' (red) = researched but no ≥2-independent value "
                       "found — left blank, not fabricated."))
    ref_detail = ref_work + \
        [dict(r, source_dir=r.get("source_dir") or _THIS_PACKET) for r in own_refunits]
    if ref_detail:
        rw_cols = _ref_columns(with_source=True)
        rw_cols.insert(5, ("Bucket", lambda r: r.get("class_out", ""), 13))
        t = f"{prefix}_RefWorkDetail"
        _write_sheet(wb_b, t, rw_cols, ref_detail, _refwork_styler(rw_cols))
        defs_b.append((t,
                       f"{len(ref_detail)} — per-ref evidence for the actions file's [ref] overlays: "
                       "REFS_ADDED = blank [ref] filled; DEAD_LINK = red current ref with verified "
                       "replacement; UNRESOLVED = also listed on the actions OpenFlags tab. "
                       "REVERIFIED refs are counts-only in the README (no action)."))
    if wiki_rest:
        wd_cols = _wikidiff_columns()
        t = f"{prefix}_WikiAlignment"
        _write_sheet(wb_b, t, wd_cols, wiki_rest, _wikidiff_styler(wd_cols))
        defs_b.append((t,
                       f"{len(wiki_rest)} — sheet↔wiki diff context that needs NO wiki edit right "
                       "now: SHEET_SUSPECT rows (the sheet side was the suspect — researched via "
                       "Leg 3; outcomes are in the actions file), WIKI_STALE_VS_STAGED (a staged "
                       "sheet correction supersedes the diff — apply that packet first, then align "
                       "the wiki), and info-severity review context. The actionable WIKI_UPDATE "
                       "rows are in the actions file's WikiUpdates tab."))
    if covered_routeqc:
        rq_cols = _routeqc_columns()
        t = f"{prefix}_RouteIntegrity"
        _write_sheet(wb_b, t, rq_cols, covered_routeqc, _routeqc_styler(rq_cols))
        defs_b.append((t,
                       f"{len(covered_routeqc)} — route-integrity flags a staged packet already "
                       "covers (blue note) or info-severity context. OPEN route flags are on the "
                       "actions file's OpenFlags tab. A route is NEVER auto-replaced."))
    if covered_flags:
        t = f"{prefix}_Flags"
        _, _ = _qc_flags_view(wb_b, t, {"flags": covered_flags})
        defs_b.append((t,
                       f"{len(covered_flags)} — mechanical-check flags already covered: blue 'Known "
                       "— staged' = a prior staged packet holds the fix (apply that packet), "
                       "'audited — no existence concern' = the row already passed an existence "
                       "audit. OPEN flags are on the actions file's OpenFlags tab."))
    if nr_monitor:
        m_cols = _compact_columns([("Why monitor (threshold leg failed)",
                                    lambda r: r.get("monitor_reason", ""), 40)]) + [_src_col]
        t = f"{prefix}_MonitorList"
        _write_sheet(wb_b, t, m_cols, nr_monitor, _flag_col2("yellow"))
        defs_b.append((t,
                       f"{len(nr_monitor)} — DISCOVERY candidates below the add-threshold: watch for "
                       "the concrete step, do NOT add yet."))
    cw_path = staging / "gulfpub_crosswalk.json"
    gulfpub_n = 0
    if cw_path.exists():
        crosswalk = json.loads(cw_path.read_text())
        t = f"{prefix}_GulfPub"
        _, gulfpub_n = _gulfpub_view(wb_b, t, crosswalk)
        defs_b.append((t,
                       f"{gulfpub_n} — GulfPub (PE World Map, Tier 2) cross-comparison context. "
                       "Disagreements route to Update's ≥2-independent source search — a single "
                       "Tier-2 value NEVER reaches green alone; nothing here is applied."))

    a_counts = (actions.get("meta") or {}).get("counts", {})
    counts = {
        "decisions_open": len(open_dec), "status_changes": len(status_rows),
        "backend_paste_units": len(afb_res), "oo_paste_units": len(oo_units),
        "new_rows": len(nr_new), "matched_existing": len(nr_matched),
        "wiki_updates": len(wiki_updates), "route_suggestions": len(routes_all),
        "route_candidates": len(route_candidates),
        "open_flags": len(open_flags),
        "confirmed_audits": len(confirmed), "fill_detail": len(fill_detail),
        "ref_detail": len(ref_detail),
        "ref_reverified": a_counts.get("ref_reverified", 0),
        "status_confirms": a_counts.get("status_verdicts", {}).get("confirm", 0)
                           + own_status_confirms,
        "wiki_context": len(wiki_rest), "flags_covered": len(covered_flags),
        "route_covered": len(covered_routeqc), "monitor": len(nr_monitor),
        "gulfpub_crosscompare": gulfpub_n,
    }
    meta = {**meta, "counts": counts}
    _split_readme(readme_a, meta, defs_a, actions_file=True, companion=out_b.name)
    _split_readme(readme_b, meta, defs_b, actions_file=False, companion=out_a.name)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb_a.save(out_a)
    wb_b.save(out_b)
    for p, wb in ((out_a, wb_a), (out_b, wb_b)):
        print(f"wrote {p}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")
    print(f"  counts: {counts}")
    print(f"  next: python scripts/recalc.py {out_a} && python scripts/recalc.py {out_b}")


def _fill_readme(ws, meta, sheet_defs, handoff=False):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    scope = meta.get("scope", {})
    counts = meta.get("counts", {})
    if meta.get("mode") in ("qc", "handoff") or handoff:
        cap = (meta.get("commodity") or scope.get("tracker") or "gas").capitalize()
        if handoff:
            title = "GEM handoff packet — everything staged for this scope"
            work_from = (
                f"FIRST check {cap}_Concerns (the gatekeeper: every carried validity concern — "
                f"settle existence/duplicate/classification doubts before investing work in a row "
                f"GEM may stop tracking), then the carried staged work ({cap}_PendingStatus, "
                f"{cap}_PendingFills, {cap}_RefWork — paste-ready, each row names its Source "
                f"packet), then this run's QC legs ({cap}_WikiAlignment, {cap}_RouteIntegrity, "
                f"{cap}_Flags, {cap}_Findings), and finally the discovery candidates "
                f"({cap}_NewRows / _MonitorList / _MatchedExisting). The "
                f"{cap}_AllFillsBackend tab unifies ALL fills (carried + this packet's own) "
                "in the exact tracker backend layout, tier-colored — paste the colored "
                "cells from there.")
        else:
            title = "GEM QC packet — sheet ↔ wiki ↔ route"
            work_from = (
                f"FIRST check {cap}_Concerns (rows whose very existence or identity is in "
                f"doubt — don't invest wiki/spec work in a row we may stop tracking), then "
                f"the {cap}_WikiAlignment tab (per-field sheet↔wiki diffs, one row per "
                f"mismatch, with an Action column), then {cap}_RouteIntegrity (drawn GeoJSON "
                f"vs the row's own attributes) and {cap}_Flags (mechanical checks). "
                f"{cap}_Findings carries the targeted-research verdicts on flagged rows.")
        rows = [
            (title, ""),
            ("Tracker / commodity", meta.get("commodity", scope.get("tracker", ""))),
            ("Country", scope.get("country", meta.get("country", ""))),
            ("GEM CSV", scope.get("csv", "")),
            ("Prior staged packets", J(meta.get("staged_dirs", []))),
        ("ESCALATIONS", J([f"{e.get('title','')} — {e.get('summary','')} [{e.get('memo','')}]"
                           for e in meta.get("escalations", [])])),
            ("", ""),
            ("Counts", J([f"{k}={v}" for k, v in counts.items()])),
            ("", ""),
            ("Work from", work_from),
            ("Color key", "WikiAlignment Class: red WIKI_UPDATE = wiki lags the sheet → edit the wiki "
                          "page; red SHEET_SUSPECT = the SHEET side is suspect → verify independently, "
                          "then fix the sheet; yellow WIKI_STALE_VS_STAGED = the sheet cell has a staged "
                          "pending correction (apply that first, then align the wiki); yellow Class with "
                          "'info' severity = review context, not a hard flag. All tabs: blue 'Known — "
                          "staged' = a prior staged packet already covers it (apply that packet; no new "
                          "research), and blue 'audited — no existence concern' = the row already passed "
                          "a deep-sweep existence audit. RouteIntegrity/Flags: red = open flag."),
            ("Route scope", "Route-integrity compares each drawn route to the row's OWN attributes "
                            "(length ratio, countries traversed, endpoints, degenerate/null geometry). "
                            "It is NOT the permanently-dropped route/WKT-format QC sheet, and no GulfPub "
                            "route comparison is included (future work). A route is NEVER auto-replaced — "
                            "fixes go via a human branch+PR against GOIT-GGIT-pipeline-routes."),
            ("Standing rules", "gem.wiki is VISITED for the diff but NEVER cited as a source (rule 1). "
                               "Never fabricate a URL (rule 2). Every researched ref passed url_verifier. "
                               "Nothing here is auto-applied — sheet edits are pasted manually and wiki "
                               "edits made manually."),
            ("", ""),
            ("Sheets", ""),
        ]
        for r in rows:
            ws.append(r)
        for name, desc in sheet_defs:
            ws.append((name, desc))
        ws["A1"].font = Font(bold=True, size=13)
        for c in (1, 2):
            ws.cell(1, c).fill = HEADER_FILL
            ws.cell(1, c).font = Font(bold=True, color="FFFFFF", size=13)
        for rn in range(2, ws.max_row + 1):
            ws.cell(rn, 1).font = Font(bold=True)
            ws.cell(rn, 2).alignment = Alignment(wrap_text=False, vertical="top")
        ws.freeze_panes = "A2"
        return
    rows = [
        ("GEM reference sweep", ""),
        ("Tracker / commodity", meta.get("commodity", scope.get("tracker", ""))),
        ("Country", scope.get("country", meta.get("country", ""))),
        ("GEM CSV", scope.get("csv", "")),
        ("Statuses", J(scope.get("statuses", "all"))),
        ("Generated", meta.get("generated", "")),
        ("", ""),
        ("Counts", J([f"{k}={v}" for k, v in counts.items()])),
        ("", ""),
        ("Work from", f"the {meta.get('commodity', scope.get('tracker', 'oil')).capitalize()}_Backend tab "
                      "(mirrors the tracker layout, value next to its [ref]) and the _OperatorsOwners tab "
                      "(owner/operator refs → the separate \"Pipeline operators/owners\" backend tab, "
                      "ProjectID-keyed); the *_Refs_* tabs are supporting detail."),
        ("Color key", "[ref]-cell color = corroboration tier: green=≥2 independent working sources / "
                      "yellow=single / red=low or none. Blue=re-verified existing ref (no action). On the "
                      "*_Refs_DeadLinks tab, a red Current-ref cell = dead/value-missing link."),
        ("Out of scope", "Route/geometry [ref] cells are NOT swept — pipeline geometry is reconciled against "
                         "the GOIT-GGIT-pipeline-routes repo (separate human branch+PR), not media [ref] URLs."),
        ("Standing rules", "Start from the row's gem.wiki page but NEVER cite gem.wiki/globalenergymonitor "
                           "(rule 1). Never theodora. Never fabricate a URL (rule 2). Every Proposed ref "
                           "passed url_verifier (HTTP 200 + value present). Nothing auto-applied — paste manually."),
        ("Target", "Per data point: ≥2 links that both WORK and corroborate each other and contain the "
                   "precise value. Searched in-country languages where needed (see Source language)."),
        ("", ""),
        ("Sheets", ""),
    ]
    for r in rows:
        ws.append(r)
    for name, desc in sheet_defs:
        ws.append((name, desc))
    ws["A1"].font = Font(bold=True, size=13)
    for c in (1, 2):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF", size=13)
    for rn in range(2, ws.max_row + 1):
        ws.cell(rn, 1).font = Font(bold=True)
        ws.cell(rn, 2).alignment = Alignment(wrap_text=False, vertical="top")
    ws.freeze_panes = "A2"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--staging", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    out = Path(args.output)
    if out.exists():
        sys.exit(f"refusing to overwrite existing {out} (use a fresh <stamp>)")

    data = json.loads((Path(args.staging) / "staged_resolutions.json").read_text())
    meta = data.get("meta", {})
    # Optional: class-level escalations that no individual row action can carry (a whole
    # class of GEM values is wrong, an ingest defect, a scope ruling). Without this they
    # live only in notes/ and the researcher working from the workbook never sees them.
    esc_path = Path(args.staging) / "escalations.json"
    if esc_path.exists():
        meta = {**meta, "escalations": json.loads(esc_path.read_text())}
    resolutions = data.get("resolutions", [])
    cmdty = (meta.get("commodity") or meta.get("scope", {}).get("tracker") or "oil")
    prefix = cmdty.capitalize()

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")

    columns = _ref_columns()
    sheet_defs = []

    # full backend header + current rows from the snapshot, so the Backend tab reproduces the
    # tracker's complete column layout (exact order) with current values prefilled
    backend_header, snapshot_rows = _backend_snapshot(meta)

    # Deep-sweep records (class_in VALIDITY / FILL) get their own dedicated tabs — they are
    # NOT ref-pairs, so they must never reach the backend mirror (a VALIDITY record's synthetic
    # `__VALIDITY__` ref_col would otherwise leak in as a phantom column) or the class_out
    # bucket tabs (where they'd be buried among ordinary unresolved refs).
    # Key off the synthetic ref_col sentinel too, not just class_in: an agent occasionally
    # mislabels a validity record's class_in (seen: HAS_REF) while still stamping the
    # `__VALIDITY__` ref_col, and that sentinel must never reach the backend mirror.
    # QC-workflow records (sheet↔wiki diffs, route-integrity flags) are partitioned out
    # FIRST — their synthetic ref_col sentinels must never reach the backend mirror or the
    # ref buckets (same reasoning as __VALIDITY__ below).
    qc_mode = meta.get("mode") in ("qc", "handoff")
    # Handoff extras: the staged_actions.json sidecar (build_qc_staging) carries EVERY
    # pending staged action from prior packets with source_dir provenance. Tolerant
    # detection — an older assembled dir may still say mode "qc" but have the sidecar.
    actions = None
    actions_path = Path(args.staging) / "staged_actions.json"
    if actions_path.exists():
        actions = json.loads(actions_path.read_text())
    is_wikidiff = lambda r: r.get("class_in") == "WIKIDIFF" or r.get("ref_col") == WIKIDIFF_REF
    is_routeqc = lambda r: r.get("class_in") == "ROUTEQC" or r.get("ref_col") == ROUTEQC_REF
    wikidiff_res = [r for r in resolutions if is_wikidiff(r)]
    routeqc_res = [r for r in resolutions if is_routeqc(r) and not is_wikidiff(r)]
    resolutions = [r for r in resolutions if not is_wikidiff(r) and not is_routeqc(r)]

    is_validity = lambda r: r.get("class_in") == "VALIDITY" or r.get("ref_col") == VALIDITY_REF
    is_status = lambda r: r.get("class_in") == "STATUS" or r.get("ref_col") == STATUS_REF
    is_route = lambda r: r.get("class_in") == "ROUTE" or r.get("ref_col") == ROUTE_REF
    validity_res = [r for r in resolutions if is_validity(r)]
    status_res = [r for r in resolutions if is_status(r) and not is_validity(r)]
    route_res = [r for r in resolutions if is_route(r)
                 and not is_validity(r) and not is_status(r)]
    fill_res = [r for r in resolutions if r.get("class_in") == "FILL"
                and not is_validity(r) and not is_status(r) and not is_route(r)]
    ref_res = [r for r in resolutions if r.get("class_in") not in ("VALIDITY", "FILL", "STATUS", "ROUTE")
               and not is_validity(r) and not is_status(r) and not is_route(r)]

    # owner/operator refs land on the separate "Pipeline operators/owners" backend tab, so
    # they get their own paste-ready mirror; everything else mirrors the tracker tab.
    oo_res = [r for r in ref_res if r.get("tab") == "operators_owners"]
    tracker_res = [r for r in ref_res if r.get("tab") != "operators_owners"]

    # HANDOFF (staged_actions.json present): the two-workbook split — actions +
    # evidence — replaces the single-file layout below. Legacy qc dirs (no sidecar)
    # and the sweep modes fall through to the single-workbook path unchanged.
    if actions is not None:
        parts = {"wikidiff": wikidiff_res, "routeqc": routeqc_res,
                 "validity": validity_res, "status": status_res, "route": route_res,
                 "fill": fill_res, "oo": oo_res, "tracker": tracker_res}
        _build_handoff(Path(args.staging), out, meta, parts, actions)
        return

    # HANDOFF gatekeeper tab first (right after README): carried validity concerns of
    # ALL types — settle existence/duplicate/classification doubts BEFORE investing
    # wiki/spec work in a row GEM may stop tracking.
    concerns = (actions or {}).get("concerns", [])
    if actions is None:
        # legacy QC dir predating staged_actions.json: render its existence/duplicate
        # carryover through the same Concerns view (extra columns stay blank)
        legacy_path = Path(args.staging) / "existence_carryover.json"
        if legacy_path.exists():
            concerns = json.loads(legacy_path.read_text()).get("concerns", [])
    if concerns:
        c_cols = _concerns_columns()
        c_title = f"{prefix}_Concerns"
        _write_sheet(wb, c_title, c_cols, concerns, _concerns_styler(c_cols))
        sheet_defs.append((c_title,
                           f"{len(concerns)} — GATEKEEPER: validity concerns carried from prior staged "
                           "packets (read-and-review, nothing applied). Red Concern = existence / "
                           "duplicate / classification — the row may be phantom, a relabel of another "
                           "ProjectID, or misclassified: settle these BEFORE wiki or spec work on the "
                           "row. Yellow = attribution/spec value problems. Blue 'Also flagged' = this "
                           "packet's own wiki/mechanical checks hit the same issue. Source packet = "
                           "the staging dir holding the full finding."))

    # Annual-update StatusReview tab leads the packet when present — the per-row status
    # verdict is what researchers act on first in an update cycle.
    if status_res:
        s_cols = _status_columns()
        s_title = f"{prefix}_StatusReview"
        _write_sheet(wb, s_title, s_cols, status_res, _status_styler(s_cols))
        sheet_defs.append((s_title,
                           f"{len(status_res)} — ANNUAL-UPDATE status verdicts, one per in-dev segment row "
                           "(NOT auto-applied). Verdict green=confirm (status verified, see evidence date) / "
                           "yellow=change (evidence-based new status; Proposed changes lists the exact "
                           "column=value edits, refs verified) / red=stale (dormancy rule -> inferred "
                           "shelved/cancelled, ShelvedCancelledType=Presumed, no ref by design) or unclear."))

    # HANDOFF: status changes carried from prior packets (verdict != confirm) — the
    # confirms are counts-only in the README.
    pending_status = (actions or {}).get("status_changes", [])
    if pending_status:
        ps_cols = _status_columns(with_source=True)
        ps_title = f"{prefix}_PendingStatus"
        _write_sheet(wb, ps_title, ps_cols, pending_status, _status_styler(ps_cols))
        sheet_defs.append((ps_title,
                           f"{len(pending_status)} — status changes carried from prior staged packets "
                           "(verdict change/stale/unclear only — confirms are counted in the README, "
                           "no action). Same verdict color key as StatusReview; NOT auto-applied. "
                           "Source packet = where the evidence is staged."))

    # HANDOFF: carried corroborated fills (read early — the backend-layout tab needs them).
    pending_fills = (actions or {}).get("fills", [])

    # PRIMARY tab first (after README).
    # QC/handoff mode: instead of a full-scope reference Backend mirror, render ALL
    # corroborated fills for the scope — carried pending fills from prior packets PLUS
    # this packet's own Leg-3 fills — unified in the exact backend layout: one row per
    # segment with a fill, current values prefilled, fill values + [ref] overlaid and
    # tier-colored (paste-ready). The individual fill tabs (PendingFills, Fills) stay
    # below as the per-fill detail. Owner/operator fills target the oo tab, not the
    # tracker, so they stay on those list tabs (Target tab column) and never overlay
    # here. Own fills sort before carried on ties, so if the same cell were ever filled
    # in both, the fresher research wins the overlay.
    if qc_mode:
        own_fills = [r for r in fill_res if r.get("class_out") == "REFS_ADDED"]
        afb_res = sorted(
            (f for f in own_fills + pending_fills
             if f.get("tab") != "operators_owners" and f.get("ref_col") not in OO_PRIMARY),
            key=lambda f: (_int_or_zero(f.get("sheet_row")), f.get("ref_col", "")))
        if afb_res:
            afb_title = f"{prefix}_AllFillsBackend"
            _backend_view(wb, afb_title, afb_res, backend_header, snapshot_rows,
                          color_values=True, sheet_row_col=False)
            sheet_defs.append((afb_title,
                               "PASTE-READY — ALL corroborated fills for the scope unified in the exact GEM "
                               "tracker backend layout: the carried pending fills from prior packets PLUS this "
                               "packet's own fills, FULL column set in sheet order, one row per segment that "
                               "has a fill, current values prefilled. No extra locator column — every column "
                               "aligns 1:1 with the sheet (locate rows by ProjectID). Filled "
                               "values AND their paired [ref] cells are overlaid, colored by corroboration "
                               "tier (green=≥2 independent / yellow=single / red=low or none / "
                               "blue=re-verified). Paste the colored cells only — never the computed/formula "
                               "columns. The PendingFills / Fills tabs below hold the per-fill detail; "
                               "owner/operator fills live there + on the oo backend tab, not a tracker row."))
    elif tracker_res:
        backend_title = f"{prefix}_Backend"
        _backend_view(wb, backend_title, tracker_res, backend_header, snapshot_rows)
        sheet_defs.append((backend_title,
                           "PRIMARY — 1:1 paste-ready mirror of the GEM tracker backend: the FULL backend "
                           "column set in exact sheet order, one row per in-scope segment, current values "
                           "prefilled (leading SheetRow = the tracker row locator). Touched [ref] cells carry "
                           "the proposed ref(s), colored by corroboration tier (green=≥2 independent / "
                           "yellow=single / red=low or none / blue=re-verified); proposed values overlaid on "
                           "their cells. Work from THIS tab; the *_Refs_* tabs below are supporting detail."))

    # operators/owners paste-ready mirror (ProjectID-keyed, ref-precedes-values)
    if oo_res:
        oo_title = f"{prefix}_OperatorsOwners"
        _operators_owners_view(wb, oo_title, oo_res)
        sheet_defs.append((oo_title,
                           "PASTE-READY — mirror of the separate \"Pipeline operators/owners\" backend tab "
                           "(GID 1489950650), ProjectID-keyed, where the [ref] column PRECEDES its values. "
                           "Operator [ref] / Owner [ref] cells carry the proposed ref(s), color-coded by "
                           "tier. Paste each back onto that tab by ProjectID — NOT onto a tracker row."))

    # HANDOFF: corroborated fills + actionable ref work carried from prior packets.
    if pending_fills:
        pf_cols = _fills_columns(with_source=True)
        pf_title = f"{prefix}_PendingFills"
        _write_sheet(wb, pf_title, pf_cols, pending_fills, _fills_styler(pf_cols))
        sheet_defs.append((pf_title,
                           f"{len(pending_fills)} — blank-value fills carried from prior staged packets: "
                           "a previously empty data point researched and filled with a paired verified "
                           "ref (tier-colored). Paste value + [ref] together (no orphans). "
                           "Source packet = where the fill is staged."))
    ref_work = (actions or {}).get("ref_work", [])
    if ref_work:
        rw_cols = _ref_columns(with_source=True)
        rw_title = f"{prefix}_RefWork"
        _write_sheet(wb, rw_title, rw_cols, ref_work, _refwork_styler(rw_cols))
        rw_counts = {}
        for r in ref_work:
            rw_counts[r.get("class_out", "")] = rw_counts.get(r.get("class_out", ""), 0) + 1
        sheet_defs.append((rw_title,
                           f"{len(ref_work)} — actionable ref work carried from prior staged packets "
                           f"({J([f'{k}={v}' for k, v in sorted(rw_counts.items())])}): REFS_ADDED = "
                           "blank [ref] filled (tier-colored); DEAD_LINK = red current ref, verified "
                           "replacement proposed; UNRESOLVED = red, manual review. REVERIFIED refs are "
                           "counts-only in the README (no action needed)."))

    # QC-workflow tabs: wiki alignment, route integrity, mechanical flags — placed after
    # the paste-ready mirrors. (The old existence_carryover tab is superseded by the
    # <Cmdty>_Concerns gatekeeper rendered from staged_actions.json.)
    if wikidiff_res:
        wd_cols = _wikidiff_columns()
        wd_title = f"{prefix}_WikiAlignment"
        _write_sheet(wb, wd_title, wd_cols, wikidiff_res, _wikidiff_styler(wd_cols))
        sheet_defs.append((wd_title,
                           f"{len(wikidiff_res)} — SHEET↔WIKI per-field diff, one row per (pipeline, field) "
                           "mismatch. Class red WIKI_UPDATE = wiki lags the sheet → edit the wiki page per "
                           "the Action column; red SHEET_SUSPECT = the sheet side is suspect → verify "
                           "independently then fix the sheet; yellow WIKI_STALE_VS_STAGED = the sheet cell "
                           "has a STAGED pending correction (apply it first, then align the wiki); UNPARSED "
                           "= page missing/unparseable, review manually. gem.wiki is visited, never cited."))
    if routeqc_res:
        rq_cols = _routeqc_columns()
        rq_title = f"{prefix}_RouteIntegrity"
        _write_sheet(wb, rq_title, rq_cols, routeqc_res, _routeqc_styler(rq_cols))
        sheet_defs.append((rq_title,
                           f"{len(routeqc_res)} — ROUTE-integrity flags: the drawn GeoJSON vs the row's own "
                           "attributes (length ratio, countries traversed, endpoint countries, degenerate/"
                           "null geometry). NOT the old route/WKT-format check (permanently dropped). Red "
                           "check = open flag; blue note = a staged packet already covers this row's route. "
                           "A route is NEVER auto-replaced — fixes go via a human routes-repo branch+PR."))
    flags_n = 0
    if qc_mode:
        qcf_path = Path(args.staging) / "qc_flags.json"
        if qcf_path.exists():
            qcf = json.loads(qcf_path.read_text())
            fl_title = f"{prefix}_Flags"
            _, flags_n = _qc_flags_view(wb, fl_title, qcf)
            sheet_defs.append((fl_title,
                               f"{flags_n} — MECHANICAL checks (vocab, date logic, geo consistency, orphan "
                               "refs, ...) scoped to this country. Red Detail = open flag → fix on the "
                               "sheet; blue 'Known — staged' = a prior staged packet already covers it."))

    # Validity tab (deep sweep): existence / duplicate / classification / attribution / spec
    # concerns — read-and-flag only, no proposed edit. Placed prominently after the paste tabs.
    # In QC mode this is the targeted-research verdict tab, retitled *_Findings.
    if validity_res:
        v_cols = _validity_columns()
        v_title = f"{prefix}_Findings" if qc_mode else f"{prefix}_Validity"
        _write_sheet(wb, v_title, v_cols, validity_res, _validity_styler(v_cols))
        sheet_defs.append((v_title,
                           f"{len(validity_res)} — DEEP-SWEEP existence/identity check (NOT ref work, NOT auto-"
                           "applied). Each row flags a concern for human review: Concern type red = "
                           "existence / duplicate / classification (the pipeline may be phantom, a relabel of "
                           "another GEM row, or misclassified — e.g. NGL vs dry gas / not a transmission line). "
                           "Finding is authoritative; Sources back the judgment."))

    # Fills tab (deep sweep): blank value fields researched + filled with a paired ref.
    if fill_res:
        f_cols = _fills_columns()
        f_title = f"{prefix}_Fills"
        _write_sheet(wb, f_title, f_cols, fill_res, _fills_styler(f_cols))
        sheet_defs.append((f_title,
                           f"{len(fill_res)} — DEEP-SWEEP blank-value fills: a previously empty data point "
                           "researched and filled with a paired Proposed ref. Outcome 'filled (corroborated)' = "
                           "ready to paste (tier-colored); 'not corroborated / dropped' (red) = no ≥2-independent "
                           "value found — left blank, not fabricated (standing rule 2)."))

    # RouteSuggestions tab: endpoint coords + corridor for weak-RouteAccuracy rows — a
    # candidate for the routes repo (separate human branch+PR), never applied here. In
    # handoff mode, route suggestions carried from prior packets render on the same tab
    # (Source packet column tells them apart; this run's own rows leave it blank).
    carried_routes = (actions or {}).get("routes", [])
    all_routes_raw = route_res + carried_routes
    is_route_candidate = lambda r: r.get("class_out") == "ROUTE_CANDIDATE"
    all_routes = [r for r in all_routes_raw if not is_route_candidate(r)]
    route_candidates = [r for r in all_routes_raw if is_route_candidate(r)]
    if all_routes:
        rt_cols = _route_columns(with_source=bool(carried_routes))
        rt_title = f"{prefix}_RouteSuggestions"
        _write_sheet(wb, rt_title, rt_cols, all_routes, _route_styler(rt_cols))
        carried_txt = (f" ({len(carried_routes)} carried from prior staged packets — "
                       "see Source packet)" if carried_routes else "")
        sheet_defs.append((rt_title,
                           f"{len(all_routes)} — route suggestions for low/medium/no-route rows: "
                           "sourced endpoint coordinates + a corridor description (tier-colored). Yellow coord "
                           "cells = corridor-only (endpoints not both coordinated — no fabricated coordinates, "
                           "standing rule 2). These feed a SEPARATE human branch+PR against the "
                           f"GOIT-GGIT-pipeline-routes repo — a route is NEVER auto-replaced.{carried_txt}"))
    if route_candidates:
        rc_cols = _route_candidate_columns()
        rc_title = f"{prefix}_RouteCandidates"
        _write_sheet(wb, rc_title, rc_cols, route_candidates, _route_candidate_styler(rc_cols))
        sheet_defs.append((rc_title,
                           f"{len(route_candidates)} — §8 candidate route GEOMETRY (staged <PID>.geojson, "
                           "destination: the ROUTES REPO via a human branch+PR). Method sets suggested "
                           "RouteAccuracy (sidecar/gis/osm=high, digitized=medium, endpoints=low). License "
                           "'ODbL' = OSM-derived (Baird's licensing call). Red QC result = failed the "
                           "validation gate (listed, not applied). Yellow Replacement? = a route already "
                           "exists — a route is NEVER auto-replaced."))

    # GulfPub cross-comparison tab (deep sweep): scoped reconcile of the PE World Map dataset
    # vs GEM, if a crosswalk was generated for this staging dir (build_gulfpub_crosswalk.py).
    cw_path = Path(args.staging) / "gulfpub_crosswalk.json"
    gulfpub_n = 0
    if cw_path.exists():
        crosswalk = json.loads(cw_path.read_text())
        gp_title = f"{prefix}_GulfPub"
        _, gulfpub_n = _gulfpub_view(wb, gp_title, crosswalk)
        cw_counts = (crosswalk.get("meta") or {}).get("counts", {})
        sheet_defs.append((gp_title,
                           f"{gulfpub_n} — DEEP-SWEEP GulfPub (PE World Map / Petroleum Economist, Tier 2) "
                           "cross-comparison: overlaps (matched pairs, Match conf tier-colored), GulfPub-only "
                           "additions (match to an existing GEM row FIRST before treating as a discovery), and "
                           "ambiguous multi-candidate matches. Red Status conflict / yellow Diam or Len flag = "
                           "a disagreement to resolve via Update's ≥2-independent source search — a single "
                           "Tier-2 value NEVER reaches green alone; nothing here is applied. "
                           f"(counts: {J([f'{k}={v}' for k, v in cw_counts.items()])})"))

    # HANDOFF: discovery candidates carried from staged_new.json — paste-ready NewRows
    # mirror (exact tracker header, from the snapshot already loaded for the Backend tab),
    # owner/operator refs, monitor list, matched-existing suggestions.
    new_rows_all = (actions or {}).get("new_rows", [])
    nr_new = [c for c in new_rows_all if c.get("class") == "new_row"]
    nr_monitor = [c for c in new_rows_all if c.get("class") == "monitor"]
    nr_matched = [c for c in new_rows_all if c.get("class") == "matched_existing"]
    if nr_new and backend_header:
        nr_title = f"{prefix}_NewRows"
        _, oo_rows = _new_rows_view(wb, nr_title, backend_header, nr_new)
        sheet_defs.append((nr_title,
                           f"{len(nr_new)} — DISCOVERY candidates (carried from prior staged packets) that "
                           "clear the add-threshold: exact tracker header, one green row per candidate, "
                           "values + verified [ref]s in place. Paste-ready — new rows, not edits."))
        if oo_rows:
            nrr_title = f"{prefix}_NewRowRefs"
            _owner_refs_view(wb, nrr_title, oo_rows)
            sheet_defs.append((nrr_title,
                               f"{len(oo_rows)} — owner/operator [ref]s for the new rows. The main tracker "
                               "has no Owner/Parent [ref] column; apply these on the ProjectID-keyed "
                               "operators/owners tab once each row has an ID. [ref] precedes its value."))
    elif nr_new:
        print(f"  WARN: {len(nr_new)} carried new_row candidates but no tracker snapshot "
              "header — NewRows mirror skipped (still in staged_new.json)")

    def _flag_col2(color):
        def styler(ws, rn, r):
            ws.cell(rn, 2).fill = CONF_FILL[color]
        return styler

    _src_col = ("Source packet", lambda r: r.get("source_dir", ""), 28)
    if nr_monitor:
        m_cols = _compact_columns([("Why monitor (threshold leg failed)",
                                    lambda r: r.get("monitor_reason", ""), 40)]) + [_src_col]
        m_title = f"{prefix}_MonitorList"
        _write_sheet(wb, m_title, m_cols, nr_monitor, _flag_col2("yellow"))
        sheet_defs.append((m_title,
                           f"{len(nr_monitor)} — DISCOVERY candidates below the add-threshold: watch for "
                           "the concrete step, do NOT add yet."))
    if nr_matched:
        x_cols = _compact_columns([("Matched ProjectID",
                                    lambda r: r.get("matched_project_id", ""), 16)]) + [_src_col]
        x_title = f"{prefix}_MatchedExisting"
        _write_sheet(wb, x_title, x_cols, nr_matched, _flag_col2("green"))
        sheet_defs.append((x_title,
                           f"{len(nr_matched)} — same physical pipe as an existing GEM row under another "
                           "name: add the candidate name to that row's OtherEnglishNames, no new row."))

    counts = {}
    if qc_mode or wikidiff_res or routeqc_res:
        counts["wiki_alignment"] = len(wikidiff_res)
        counts["route_integrity"] = len(routeqc_res)
        counts["mechanical_flags"] = flags_n
    if actions:
        counts["carried_concerns"] = len(concerns)
        counts["carried_status_changes"] = len(pending_status)
        counts["carried_fills"] = len(pending_fills)
        counts["carried_ref_work"] = len(ref_work)
        a_counts = (actions.get("meta") or {}).get("counts", {})
        counts["carried_ref_reverified"] = a_counts.get("ref_reverified", 0)
        counts["new_rows"] = len(nr_new)
        counts["monitor"] = len(nr_monitor)
        counts["matched_existing"] = len(nr_matched)
    counts["status_reviews"] = len(status_res)
    counts["validity"] = len(validity_res)
    counts["fills"] = len(fill_res)
    counts["route_suggestions"] = len(all_routes)
    if route_candidates:
        counts["route_candidates"] = len(route_candidates)
    counts["gulfpub_crosscompare"] = gulfpub_n
    for bucket in _ORDER:
        rows = [r for r in ref_res if r.get("class_out") == bucket]
        counts[bucket.lower()] = len(rows)
        if not rows:
            continue
        suffix, blurb = _BUCKETS[bucket]
        title = f"{prefix}_{suffix}"
        _write_sheet(wb, title, columns, rows, _make_styler(columns, bucket))
        sheet_defs.append((title, f"{len(rows)} — {blurb}"))

    meta.setdefault("counts", counts)
    _fill_readme(readme, meta, sheet_defs, handoff=bool(actions))

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")
    print(f"  counts: {counts}")
    print("  next: python scripts/recalc.py " + str(out))


if __name__ == "__main__":
    main()
