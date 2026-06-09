#!/usr/bin/env python3
"""Build the data-health QC workbook for a GEM tracker (rebuild of GOIT_oil_ngl_QC.xlsx).
One sheet per check; flags are REVIEW items, not auto-rejections. The route/WKT-format
sheet is permanently dropped. QC detects; Update fixes.

    python scripts/build_qc_workbook.py --tracker oil [--country "Saudi Arabia"] \
        --output batches/pipelines_batch_<stamp>_ET_<scope>_qc.xlsx
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import normalize as N  # noqa: E402
from match import load_gem_df  # noqa: E402
from ref_pairs import discover_ref_pairs  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FLAG_FILL = PatternFill("solid", fgColor="FFCCCC")

ROUTE_ACCURACY = {"high", "medium", "low", "no route", "very high (within meters)", ""}
PIPELINE_TYPE = {"transmission", "gathering", "distribution", ""}
TITLE_VOCAB = {
    "FIDStatus": {"Pre-FID", "FID", ""},
    "Opposition": {"Yes", "No", ""},
    "Delayed": {"Yes", ""},
    "ShelvedCancelledType": {"Presumed", "Confirmed", ""},
    "DelayType": {"Presumed", "Confirmed", ""},
}
ID_COLS = ["ProjectID", "PipelineName", "SegmentName", "CountriesOrAreas"]


def _base(row):
    return {k: row.get(k, "") for k in ID_COLS}


# --- checks: each returns list of dict rows with ID_COLS + 'Detail' ---------- #
def check_status(df):
    out = []
    for _, r in df.iterrows():
        s = str(r.get("Status", "")).strip()
        if s and s.lower() not in N.GEM_STATUSES:
            out.append({**_base(r), "Detail": f"Status='{s}' not in vocab"})
    return out


def check_route_accuracy(df):
    out = []
    for _, r in df.iterrows():
        v = str(r.get("RouteAccuracy", "")).strip()
        if v.lower() not in ROUTE_ACCURACY and v not in ROUTE_ACCURACY:
            out.append({**_base(r), "Detail": f"RouteAccuracy='{v}' not in ladder"})
    return out


def check_other_vocab(df):
    out = []
    for _, r in df.iterrows():
        pt = str(r.get("PipelineType", "")).strip()
        if pt and pt.lower() not in PIPELINE_TYPE:
            out.append({**_base(r), "Detail": f"PipelineType='{pt}' not in vocab"})
        for col, allowed in TITLE_VOCAB.items():
            v = str(r.get(col, "")).strip()
            if v and v not in allowed:
                out.append({**_base(r), "Detail": f"{col}='{v}' not in {sorted(allowed - {''})}"})
    return out


def check_owner_format(df):
    out = []
    for _, r in df.iterrows():
        o = str(r.get("Owner", "")).strip()
        if not o:  # '--' is a VALID sentinel; truly blank is the flag
            out.append({**_base(r), "Detail": "Owner blank (use '--' if unknown)"})
    return out


def check_date_logic(df):
    out = []
    for _, r in df.iterrows():
        py, cy, sy = (N.parse_year(r.get(c)) for c in ("ProposalYear", "ConstructionYear", "StartYear1"))
        seq = [(y, n) for y, n in ((py, "Proposal"), (cy, "Construction"), (sy, "Start")) if y]
        for i in range(len(seq) - 1):
            if seq[i][0] > seq[i + 1][0]:
                out.append({**_base(r), "Detail": f"{seq[i][1]}Year {seq[i][0]} > {seq[i+1][1]}Year {seq[i+1][0]}"})
                break
        if str(r.get("Status", "")).strip().lower() == "operating" and not sy:
            out.append({**_base(r), "Detail": "Status=operating but no StartYear1"})
    return out


def check_diameter(df):
    out = []
    for _, r in df.iterrows():
        for d in N.parse_diameter_set(r.get("Diameter")):
            if d < 2 or d > 60:
                out.append({**_base(r), "Detail": f"diameter {d}in out of plausible range [2,60]"})
                break
    return out


def check_name_uniqueness(df):
    groups = defaultdict(list)
    for _, r in df.iterrows():
        key = (str(r.get("PipelineNetworkGrouping", "")).strip() or str(r.get("PipelineName", "")).strip(),
               str(r.get("SegmentName", "")).strip())
        groups[key].append(r)
    out = []
    for (grp, seg), rows in groups.items():
        if len(rows) > 1 and seg:
            for r in rows:
                out.append({**_base(r), "Detail": f"duplicate (grouping,segment)=({grp},{seg}) x{len(rows)}"})
    return out


def check_geo_consistency(df):
    out = []
    for _, r in df.iterrows():
        countries = set(N.split_countries(r.get("CountriesOrAreas")))
        for col in ("StartCountryOrArea", "EndCountryOrArea"):
            c = N.normalize_country(r.get(col))
            if c and countries and c not in countries:
                out.append({**_base(r), "Detail": f"{col}='{r.get(col)}' not in CountriesOrAreas"})
    return out


def check_wikilink(df):
    out = []
    for _, r in df.iterrows():
        wiki = str(r.get("Wiki", "")).strip()
        if not wiki:
            out.append({**_base(r), "Detail": "Wiki link blank"})
        elif "gem.wiki" not in wiki and "globalenergymonitor" not in wiki and wiki.startswith("http"):
            out.append({**_base(r), "Detail": f"Wiki not a GEM URL: {wiki[:60]}"})
    return out


def check_broadsweep(df):
    """Orphan-ref sweep over EVERY discovered ref-pair (group-walk from the fresh header,
    via ref_pairs.discover_ref_pairs — single source of truth, shared with the Ref Sweep
    workflow). Flags a `[ref]` cell that is filled while all the value cols it sources are
    blank. The reverse (value present, ref blank) is the Ref Sweep's job, not QC's."""
    out = []
    pairs = [p for p in discover_ref_pairs(list(df.columns)) if p["ref_col"]]
    for _, r in df.iterrows():
        for p in pairs:
            ref = str(r.get(p["ref_col"], "")).strip()
            if not ref:
                continue
            if not any(str(r.get(c, "")).strip() for c in p["value_cols"]):
                cols = "/".join(p["value_cols"])
                out.append({**_base(r), "Detail": f"orphan ref: {p['ref_col']} filled but {cols} all blank"})
    return out


CHECKS = [
    ("Status", check_status), ("RouteAccuracy", check_route_accuracy), ("OtherVocab", check_other_vocab),
    ("Owner_format", check_owner_format), ("WikiLink_health", check_wikilink),
    ("Geo_consistency", check_geo_consistency), ("Name_uniqueness", check_name_uniqueness),
    ("Date_logic", check_date_logic), ("Diameter_OutOfRange", check_diameter),
    ("BroadSweep_Misc", check_broadsweep),
]


def _add_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    cols = ID_COLS + ["Detail"]
    widths = [12, 40, 24, 26, 50]
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
        ws.cell(ws.max_row, len(cols)).fill = FLAG_FILL
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.freeze_panes = "A2"
    return len(rows)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--tracker", required=True, choices=["oil", "gas"])
    ap.add_argument("--country")
    ap.add_argument("--csv", help="GEM CSV (default: latest snapshot for the tracker)")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    csv = args.csv
    if not csv:
        import glob
        import paths
        pat = "GOIT_oil_ngl_snapshot_*.csv" if args.tracker == "oil" else "GGIT_gas_snapshot_*.csv"
        files = sorted(glob.glob(str(paths.repo_root() / "data" / pat)))
        if not files:
            sys.exit("no GEM snapshot found — run ./scripts/refresh_csvs.sh")
        csv = files[-1]

    df = load_gem_df(csv)
    if args.country:
        want = N.normalize_country(args.country)
        df = df[df["CountriesOrAreas"].map(lambda s: want in N.split_countries(s))].reset_index(drop=True)

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")
    summary = []
    for title, fn in CHECKS:
        rows = fn(df)
        if rows:
            n = _add_sheet(wb, title, rows)
            summary.append((title, n))

    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 70
    readme.append(["QC workbook", f"{args.tracker} | {args.country or 'global'} | {Path(csv).name} | {len(df)} rows"])
    readme.append(["", ""])
    readme.append(["Check", "Flagged rows"])
    for title, n in summary:
        readme.append([title, n])
    if not summary:
        readme.append(["(clean)", "no flags raised"])
    readme.append(["", ""])
    readme.append(["Note", "Flags are REVIEW items, not auto-rejections. Route/WKT-format sheet permanently dropped. QC detects → Update fixes."])
    for c in (1, 2):
        readme.cell(1, c).fill, readme.cell(1, c).font = HEADER_FILL, HEADER_FONT
    readme.freeze_panes = "A2"

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")
    print(f"  checks flagged: {dict(summary) if summary else 'none (clean)'}")


if __name__ == "__main__":
    main()
