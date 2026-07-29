#!/usr/bin/env python3
"""Build the reconciliation workbook from match_diff.json (generalizes the Saudi
GulfPub POC to any source/country/commodity).

    python scripts/build_recon_workbook.py --staging batches/<scope>/staging/recon-<source>-<date>/ \
        --output batches/pipelines_batch_<stamp>_ET_<scope>_reconciliation.xlsx

Sheets (per-commodity prefixed): README, <Cmdty>_Overlaps, <Cmdty>_Additions,
<Cmdty>_GEM_only, Status_Conflicts, Routes_WKT, Ambiguous_Clusters. Empty sheets are
omitted. Colors per docs/reference/workbook_conventions.md. Findings are candidates —
nothing here is auto-applied.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from build_recon_crosswalk import DISPOSITION_ACTION  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CONF_FILL = {"green": PatternFill("solid", fgColor="C6EFCE"),
             "yellow": PatternFill("solid", fgColor="FFEB9C"),
             "red": PatternFill("solid", fgColor="FFC7CE")}
ADDITION_FILL = PatternFill("solid", fgColor="E2EFDA")
FLAG_FILL = PatternFill("solid", fgColor="FFFF00")
CELL_MAX = 32000


def J(v):
    if v is None:
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return v


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.freeze_panes = "A2"


def _write_sheet(wb, title, columns, rows, styler=None):
    """columns: list of (header, value_fn, width). styler(ws, rownum, item)."""
    ws = wb.create_sheet(title)
    ws.append([c[0] for c in columns])
    for item in rows:
        ws.append([J(c[1](item)) for c in columns])
        if styler:
            styler(ws, ws.max_row, item)
    _style_header(ws, len(columns))
    for i, c in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = c[2]
    return ws


# --- column specs ---------------------------------------------------------- #
def _overlap_columns():
    r = lambda k: (lambda o: o["ref"].get(k))
    g = lambda k: (lambda o: o["gem"].get(k))
    return [
        ("Confidence", lambda o: o["confidence"], 11),
        ("Coverage", lambda o: o.get("coverage", ""), 11),
        ("Match reason / notes", lambda o: o["reason"], 46),
        ("Ref OID", r("oid"), 9), ("Ref Name", r("name"), 34), ("Ref Status", r("status"), 12),
        ("Ref Start", r("start"), 26), ("Ref End", r("end"), 26),
        ("Ref Diameter (in)", r("diameter"), 12), ("Ref Length (km)", r("length_km"), 12),
        ("Ref Geodesic (km)", r("geodesic_km"), 13), ("Ref Capacity", r("capacity"), 12),
        ("Ref Units", r("capacity_units"), 9), ("Ref Operator", r("operator"), 22),
        ("Ref StartYear", r("start_year"), 11),
        ("GEM ProjectID(s)", lambda o: o["gem_segments"], 16), ("GEM PipelineName", g("pipeline_name"), 34),
        ("GEM SegmentName", g("segment_name"), 22), ("GEM Status", g("status"), 12),
        ("GEM Diameter", g("diameter"), 14), ("GEM Length (km)", g("length_km"), 13),
        ("GEM Start", g("start"), 22), ("GEM End", g("end"), 22), ("GEM Owner", g("owner"), 30),
        ("GEM RouteAccuracy", g("route_accuracy"), 14), ("GEM Wiki", g("wiki"), 30),
        ("Match level", g("kind"), 10), ("Route IoU", lambda o: o["route_iou"], 10),
        ("Route replacement candidate?", lambda o: "YES" if o["route_replacement_candidate"] else "", 16),
    ]


# What a reviewer should DO with each disposition — same table the sweep crosswalk uses.
_DISP_ORDER = ["ROUTE_FOR_EXISTING", "FRAGMENT_OF_EXISTING", "NEAR_MISS", "DISCOVERY_CANDIDATE"]


def _addition_columns():
    r = lambda k: (lambda a: a["ref"].get(k))
    return [
        ("Disposition", lambda a: a.get("disposition", ""), 22),
        ("What to do", lambda a: DISPOSITION_ACTION.get(a.get("disposition", ""), ""), 76),
        ("Trace crosses", lambda a: a.get("trace_footprint", ""), 34),
        ("Ref OID", r("oid"), 9), ("Ref Name", r("name"), 36), ("Ref Status", r("status"), 12),
        ("Ref Start", r("start"), 28), ("Ref End", r("end"), 28),
        ("Ref Diameter (in)", r("diameter"), 12), ("Ref Length (km)", r("length_km"), 12),
        ("Ref Geodesic (km)", r("geodesic_km"), 13), ("Ref Capacity", r("capacity"), 12),
        ("Ref Operator", r("operator"), 24), ("Ref StartYear", r("start_year"), 11),
        ("Ref Description", r("description"), 40), ("Citation", r("report_citation"), 34),
        ("Closest GEM (best guess)", lambda a: (a["best_guess"] or {}).get("name"), 30),
        ("Best score", lambda a: (a["best_guess"] or {}).get("composite"), 10),
        ("Note", lambda a: a["note"], 44),
    ]


def _gem_only_columns():
    return [(h, (lambda k: lambda x: x.get(k))(key), w) for h, key, w in [
        ("GEM ProjectID", "project_id", 13), ("PipelineName", "pipeline_name", 34),
        ("SegmentName", "segment_name", 22), ("Status", "status", 12), ("Owner", "owner", 30),
        ("Diameter", "diameter", 14), ("Length (km)", "length_km", 12),
        ("Start", "start", 24), ("End", "end", 24), ("RouteAccuracy", "route_accuracy", 14),
        ("Wiki", "wiki", 34), ("Note", "note", 30)]]


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--staging", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    diff = json.loads((Path(args.staging) / "match_diff.json").read_text())
    meta = diff["meta"]
    trackers = ["oil", "gas"]

    wb = Workbook()
    wb.remove(wb.active)
    readme = wb.create_sheet("README")  # placeholder; filled last, stays first

    sheet_defs = []

    def by_tracker(rows, t):
        return [x for x in rows if x.get("tracker") == t]

    for t in trackers:
        cap = t.capitalize()
        ov = by_tracker(diff["overlaps"], t)
        if ov:
            def ov_style(ws, rn, o):
                ws.cell(rn, 1).fill = CONF_FILL.get(o["confidence"], PatternFill())
                if o["route_replacement_candidate"]:
                    ws.cell(rn, len(_overlap_columns())).fill = FLAG_FILL
            _write_sheet(wb, f"{cap}_Overlaps", _overlap_columns(), ov, ov_style)
            sheet_defs.append((f"{cap}_Overlaps", f"{len(ov)} matched {t} pairs; Confidence colored "
                               "green/yellow/red; yellow flag = GulfPub route may replace a low-accuracy GEM route."))
        # Sorted by disposition, longest trace first: an unmatched reference route is
        # presumptively REAL pipe, and the handful that are candidate geometry for a
        # routeless GEM row must not be buried under a run of sub-kilometre stubs.
        add = sorted(by_tracker(diff["additions"], t),
                     key=lambda a: (_DISP_ORDER.index(a["disposition"])
                                    if a.get("disposition") in _DISP_ORDER else len(_DISP_ORDER),
                                    -((a.get("ref") or {}).get("geodesic_km") or 0)))
        if add:
            _write_sheet(wb, f"{cap}_Additions", _addition_columns(), add,
                         lambda ws, rn, a, n=len(_addition_columns()): [setattr(ws.cell(rn, c), "fill", ADDITION_FILL) for c in range(1, n + 1)])
            from collections import Counter as _C
            _d = _C(a.get("disposition") or "?" for a in add)
            sheet_defs.append((f"{cap}_Additions", f"{len(add)} {t} reference records with no GEM match "
                               f"({', '.join(f'{k}={v}' for k, v in _d.most_common())}). A reference route is "
                               "presumptively REAL pipe — read Disposition + What to do: ROUTE_FOR_EXISTING = "
                               "candidate geometry for a routeless GEM row (human routes-repo PR, NEVER "
                               "auto-replaced), FRAGMENT_OF_EXISTING = partial trace of a tracked line, "
                               "NEAR_MISS = adjudicate by hand, DISCOVERY_CANDIDATE = match to an existing "
                               "GEM pipeline under another name FIRST."))
        go = by_tracker(diff["gem_only"], t)
        if go:
            _write_sheet(wb, f"{cap}_GEM_only", _gem_only_columns(), go)
            sheet_defs.append((f"{cap}_GEM_only", f"{len(go)} GEM {t} rows with no reference match "
                               "(GEM is more granular than GulfPub — usually expected)."))

    if diff["status_conflicts"]:
        cols = [(h, (lambda k: lambda x: x.get(k))(key), w) for h, key, w in [
            ("Tracker", "tracker", 8), ("Ref Name", "ref_name", 30), ("Ref Status", "ref_status", 12),
            ("GEM ProjectID(s)", "gem_project_ids", 18), ("GEM Name", "gem_name", 30),
            ("GEM Segment", "gem_segment", 20), ("GEM Status", "gem_status", 12),
            ("Recommendation", "recommendation", 44)]]
        _write_sheet(wb, "Status_Conflicts", cols, diff["status_conflicts"])
        sheet_defs.append(("Status_Conflicts", f"{len(diff['status_conflicts'])} status disagreements — verify "
                           "true current status; never auto-flip."))

    if diff["routes"]:
        cols = [
            ("Tracker", lambda x: x["tracker"], 8), ("Ref OID", lambda x: x["oid"], 9),
            ("Ref Name", lambda x: x["name"], 30), ("Ref Status", lambda x: x["status"], 12),
            ("Ref Diameter", lambda x: x["diameter"], 12), ("Ref Length (km)", lambda x: x["length_km"], 12),
            ("Geodesic (km)", lambda x: x["geodesic_km"], 12),
            ("Matched GEM ProjectID(s)", lambda x: x["matched_project_ids"], 20),
            ("GEM RouteAccuracy", lambda x: x["gem_route_accuracy"], 14),
            ("Route IoU", lambda x: x["route_iou"], 10),
            ("Replacement candidate?", lambda x: "YES" if x["replacement_candidate"] else "", 16),
            ("WKT (reference geometry)", lambda x: (x["wkt"] or "")[:CELL_MAX], 60)]

        def route_style(ws, rn, x):
            if x["replacement_candidate"]:
                ws.cell(rn, 11).fill = FLAG_FILL
        _write_sheet(wb, "Routes_WKT", cols, diff["routes"], route_style)
        sheet_defs.append(("Routes_WKT", f"{len(diff['routes'])} reference route geometries (WKT) + matched GEM "
                           "ProjectID + IoU. Yellow = replacement candidate (human review before any GeoJSON swap)."))

    if diff["ambiguous"]:
        cols = [("Tracker", lambda x: x["tracker"], 8), ("Ref Name", lambda x: x["ref_name"], 30),
                ("Candidate 1", lambda x: f"{x['candidates'][0]['name']} ({x['candidates'][0]['composite']})", 40),
                ("Candidate 2", lambda x: f"{x['candidates'][1]['name']} ({x['candidates'][1]['composite']})", 40)]
        _write_sheet(wb, "Ambiguous_Clusters", cols, diff["ambiguous"])
        sheet_defs.append(("Ambiguous_Clusters", f"{len(diff['ambiguous'])} refs whose top-2 GEM candidates "
                           "(different pipelines) score within 10% — resolve by hand."))

    _fill_readme(readme, meta, sheet_defs)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")


def _signal_line(diag: dict) -> str:
    """The per-axis health of the pairing: which signals actually carried the matching."""
    if not diag:
        return ""
    bits = []
    for label, key in (("refs named", "pct_reference_named"),
                       ("with geometry", "pct_reference_with_geometry"),
                       ("geoarea-scored", "pct_reference_geoarea_scored")):
        if diag.get(key) is not None:
            bits.append(f"{diag[key]}% {label}")
    for cmdty, pool in (diag.get("gem_pool") or {}).items():
        if pool.get("pct_with_route") is not None:
            bits.append(f"GEM {cmdty} routes {pool['pct_with_route']}% "
                        f"({pool.get('with_route')}/{pool.get('rows')})")
    if diag.get("overlap_rate") is not None:
        bits.append(f"overlap rate {diag['overlap_rate']}%")
    return "; ".join(bits)


def _fill_readme(ws, meta, sheet_defs):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    diag = meta.get("diagnostics") or {}
    rows = [
        ("GEM ↔ reference reconciliation", ""),
        ("Source", f"{meta.get('display_name')} (tier {meta.get('source_tier')})"),
        ("Country", meta.get("country")), ("Commodity", meta.get("commodity")),
        ("GEM CSV(s)", J([f"{k}: {v}" for k, v in (meta.get('gem_csv') or {}).items()])),
        ("", ""),
        ("Counts", J([f"{k}={v}" for k, v in meta.get("counts", {}).items()])),
    ]
    # Matcher health. A null or thin run is a claim about the MATCHER until its health line is
    # read, and reconcile.py only ever printed these to stdout — so the one person who needed
    # them (whoever opens the workbook) never saw them. Iraq/Egypt gas OSM both matched almost
    # entirely on the province-coarse admin-area signal; that has to travel with the findings.
    escalations = [(e.get("code") or "ESCALATION",
                    " ".join(x for x in (e.get("detail"), e.get("action")) if x))
                   for e in (diag.get("escalations") or [])]
    signal = _signal_line(diag)
    if signal:
        rows.append(("Signal", signal))
    rows.extend(escalations)
    rows.append(("", ""))
    rows += [
        ("Color key", "Confidence: green=strong match / yellow=review / red=weak. "
                      "Green-tint row=reference-only Addition. Yellow cell=route-replacement candidate."),
        ("Hard rule", "Findings are CANDIDATES, never auto-applied. Additions→Discovery; "
                      "value/status disagreements→Update. A single Tier-2 source never settles a value alone."),
        ("", ""),
        ("Sheets", ""),
    ]
    for r in rows:
        ws.append(r)
    for name, desc in sheet_defs:
        ws.append((name, desc))
    ws["A1"].font = Font(bold=True, size=13)
    for c in range(1, 3):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF", size=13)
    for rn in range(2, ws.max_row + 1):
        ws.cell(rn, 1).font = Font(bold=True)
        ws.cell(rn, 2).alignment = Alignment(wrap_text=False, vertical="top")
    # A matcher-health warning that reads like an ordinary README row gets skimmed past, so
    # tint it with the same red the workbook already uses for "weak".
    codes = {c for c, _ in escalations}
    for rn in range(2, ws.max_row + 1):
        if ws.cell(rn, 1).value in codes:
            for c in (1, 2):
                ws.cell(rn, c).fill = CONF_FILL["red"]
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    main()
